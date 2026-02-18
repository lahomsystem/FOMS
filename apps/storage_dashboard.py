"""수납장 대시보드 Blueprint. Phase 2-2 app.py 슬림다운."""
from flask import Blueprint, render_template, request, redirect, url_for, flash, send_file, session, current_app
from sqlalchemy import or_, func, String
from db import get_db
from models import Order
from apps.auth import login_required, role_required, log_access
from constants import CABINET_STATUS, STATUS
import os
import datetime

storage_dashboard_bp = Blueprint('storage_dashboard', __name__)


@storage_dashboard_bp.route('/storage_dashboard')
@login_required
def storage_dashboard():
    """수납장 대시보드"""
    db = get_db()
    search_query = request.args.get('search_query', '').strip()

    base_query = db.query(Order).filter(
        Order.is_cabinet == True,
        Order.status != 'DELETED'
    )

    if search_query:
        search_term = f"%{search_query}%"
        id_conditions = []
        try:
            search_id = int(search_query)
            id_conditions.append(Order.id == search_id)
        except ValueError:
            id_conditions.append(func.cast(Order.id, String).ilike(search_term))

        base_query = base_query.filter(
            or_(
                Order.customer_name.ilike(search_term),
                Order.phone.ilike(search_term),
                Order.address.ilike(search_term),
                Order.product.ilike(search_term),
                Order.notes.ilike(search_term),
                *id_conditions
            )
        )

    all_cabinet_orders = base_query.order_by(Order.id.desc()).all()

    # 카테고리 분류: 접수(RECEIVED), 제작중(IN_PRODUCTION), 발송(SHIPPED)
    received_orders = [o for o in all_cabinet_orders if (o.cabinet_status or 'RECEIVED') == 'RECEIVED']
    in_production_orders = [o for o in all_cabinet_orders if o.cabinet_status == 'IN_PRODUCTION']
    shipped_orders = [o for o in all_cabinet_orders if o.cabinet_status == 'SHIPPED']

    return render_template('storage_dashboard.html',
                           received_orders=received_orders,
                           in_production_orders=in_production_orders,
                           shipped_orders=shipped_orders,
                           search_query=search_query,
                           CABINET_STATUS=CABINET_STATUS,
                           STATUS=STATUS)


@storage_dashboard_bp.route('/api/storage_dashboard/export_excel')
@login_required
@role_required(['ADMIN', 'MANAGER', 'STAFF'])
def export_storage_dashboard_excel():
    """수납장 대시보드 제작중 주문 엑셀 내보내기"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

    db = get_db()
    search_query = request.args.get('search_query', '').strip()

    # 제작중 주문만 조회
    base_query = db.query(Order).filter(
        Order.is_cabinet == True,
        Order.status != 'DELETED',
        Order.cabinet_status == 'IN_PRODUCTION'
    )

    # 검색어 필터 적용
    if search_query:
        search_term = f"%{search_query}%"
        id_conditions = []
        try:
            search_id = int(search_query)
            id_conditions.append(Order.id == search_id)
        except ValueError:
            id_conditions.append(func.cast(Order.id, String).ilike(search_term))

        base_query = base_query.filter(
            or_(
                Order.customer_name.ilike(search_term),
                Order.phone.ilike(search_term),
                Order.address.ilike(search_term),
                Order.product.ilike(search_term),
                Order.regional_memo.ilike(search_term),
                *id_conditions
            )
        )

    orders = base_query.order_by(Order.id.desc()).all()

    if not orders:
        flash('다운로드할 데이터가 없습니다.', 'warning')
        return redirect(url_for('storage_dashboard.storage_dashboard'))

    # 엑셀 워크북 생성
    wb = Workbook()
    ws = wb.active
    ws.title = "제작중 주문"

    # 헤더 정의
    headers = ['번호', '메모', '고객명', '전화번호', '주소', '제품', '배송비', '상태', '설치 예정일']

    # 스타일 정의
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_alignment = Alignment(horizontal='center', vertical='center')

    # 헤더 작성
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center_alignment

    # 데이터 작성
    for row_idx, order in enumerate(orders, start=2):
        ws.cell(row=row_idx, column=1, value=order.id).border = border
        ws.cell(row=row_idx, column=2, value=order.regional_memo or '').border = border
        ws.cell(row=row_idx, column=3, value=order.customer_name).border = border
        ws.cell(row=row_idx, column=4, value=order.phone).border = border
        ws.cell(row=row_idx, column=5, value=order.address).border = border
        ws.cell(row=row_idx, column=6, value=order.product).border = border
        shipping_fee = order.shipping_fee or 0
        ws.cell(row=row_idx, column=7, value=f"{shipping_fee:,}").border = border
        ws.cell(row=row_idx, column=7).alignment = Alignment(horizontal='right', vertical='center')
        status_korean = CABINET_STATUS.get(order.cabinet_status, order.cabinet_status or '')
        ws.cell(row=row_idx, column=8, value=status_korean).border = border
        ws.cell(row=row_idx, column=8).alignment = center_alignment
        ws.cell(row=row_idx, column=9, value=order.scheduled_date or '').border = border

    # 컬럼 너비 자동 조정
    column_widths = [10, 25, 15, 15, 35, 30, 12, 12, 15]
    for col_idx, width in enumerate(column_widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width

    # 파일 저장
    upload_folder = current_app.config.get('UPLOAD_FOLDER', 'uploads')
    os.makedirs(upload_folder, exist_ok=True)
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    excel_filename = f"storage_in_production_{timestamp}.xlsx"
    excel_path = os.path.join(upload_folder, excel_filename)

    wb.save(excel_path)

    # 로그 기록
    log_access(f"수납장 대시보드 제작중 엑셀 다운로드: {excel_filename} ({len(orders)}건)", session.get('user_id'))

    return send_file(excel_path, as_attachment=True, download_name=excel_filename)
