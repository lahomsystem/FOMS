import warnings
# warnings.filterwarnings("ignore", category=DeprecationWarning, module="eventlet")
# import eventlet
# eventlet.monkey_patch()
import os
import hashlib
import datetime
import json
from concurrent.futures import ThreadPoolExecutor
import pandas as pd
import re
from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, g, session, send_file, send_from_directory, current_app
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address
from flask_compress import Compress
from whitenoise import WhiteNoise
from markupsafe import Markup
from werkzeug.utils import secure_filename
from werkzeug.security import generate_password_hash, check_password_hash
from sqlalchemy import or_, and_, text, func, String
from sqlalchemy.orm.attributes import flag_modified
import copy
import json
from datetime import date, timedelta

# 데이터베이스 관련 임포트
from db import get_db, close_db, init_db, db_session
from models import Order, User, SecurityLog, ChatRoom, ChatRoomMember, ChatMessage, ChatAttachment, OrderAttachment, OrderEvent, OrderTask, Notification
from apps.auth import is_password_strong, get_user_by_username
from services.business_calendar import add_business_days
from erp_automation import apply_auto_tasks
from services.erp_policy import (
    recommend_owner_team, 
    get_required_task_keys_for_stage, 
    STAGE_LABELS,
    STAGE_NAME_TO_CODE,
    get_quest_templates,
    get_quest_template_for_stage,
    get_required_approval_teams_for_stage,
    get_next_stage_for_completed_quest,
    check_quest_approvals_complete,
    create_quest_from_template,
    get_stage,
    DEFAULT_OWNER_TEAM_BY_STAGE,
    can_modify_domain,
    get_assignee_ids,
)

# 견적 계산기 독립 데이터베이스 임포트
from wdcalculator_db import get_wdcalculator_db, close_wdcalculator_db, init_wdcalculator_db
from wdcalculator_models import Estimate, EstimateOrderMatch, EstimateHistory

# 백업 시스템 임포트
from simple_backup_system import SimpleBackupSystem

# 지도 시스템 임포트
from foms_address_converter import FOMSAddressConverter
from foms_map_generator import FOMSMapGenerator

# 스토리지 시스템 임포트 (Quest 2)
from services.storage import get_storage
from erp_order_text_parser import parse_order_text
from map_config import KAKAO_REST_API_KEY
from services.business_calendar import business_days_until
from constants import STATUS, BULK_ACTION_STATUS, CABINET_STATUS, UPLOAD_FOLDER, ALLOWED_EXTENSIONS, CHAT_ALLOWED_EXTENSIONS, ERP_MEDIA_ALLOWED_EXTENSIONS

# SocketIO Import (Quest 5)
try:
    from flask_socketio import SocketIO, emit, join_room, leave_room
    SOCKETIO_AVAILABLE = True
except ImportError:
    SOCKETIO_AVAILABLE = False
    print("[WARN] Flask-SocketIO not installed. pip install flask-socketio python-socketio eventlet")

# Initialize Flask app
app = Flask(__name__)

# ==========================================
# Data Optimization (Quest 14)
# ==========================================

# 1. Gzip Compression (Reduce JSON/HTML size by ~70%)
Compress(app)

# 2. WhiteNoise (Fast Static File Serving & Caching)
# Railway/Heroku 배포 시 필수 최적화
_is_production = (os.environ.get('FLASK_ENV') == 'production')
app.wsgi_app = WhiteNoise(
    app.wsgi_app,
    root='static/',
    prefix='static/',
    # 개발 모드에서 정적 파일 변경 시 Content-Length 캐시 불일치 방지
    autorefresh=not _is_production,
    max_age=31536000 if _is_production else 0,
)

# Secret Key from environment variable (CRITICAL: Never hardcode in production!)
app.secret_key = os.environ.get('SECRET_KEY')
if not app.secret_key:
    # Development fallback (MUST set SECRET_KEY in production!)
    if os.environ.get('FLASK_ENV') == 'production':
        raise ValueError("SECRET_KEY environment variable must be set in production!")
    app.secret_key = 'dev-secret-key-CHANGE-IN-PRODUCTION'
    print("[WARN] Using development secret key. Set SECRET_KEY environment variable for production!")

# Session cookie configuration (prevent conflicts with other Flask apps on same domain)
app.config['SESSION_COOKIE_NAME'] = 'session_staging'  # Different from port 5000 (session_dev)

# Fix for Railway/Load Balancer (HTTPS redirect loop)
from werkzeug.middleware.proxy_fix import ProxyFix
app.wsgi_app = ProxyFix(app.wsgi_app, x_for=1, x_proto=1, x_host=1, x_prefix=1)

# Import Apps Blueprints
from apps.auth import auth_bp, login_required, role_required, ROLES, TEAMS, log_access, get_user_by_id
app.register_blueprint(auth_bp)

# ERP Beta Blueprint
from apps.erp import erp_bp, apply_erp_display_fields_to_orders, can_edit_erp
app.register_blueprint(erp_bp)

# API Files Blueprint
from apps.api.files import files_bp
app.register_blueprint(files_bp)

# API Address Blueprint
from apps.api.address import address_bp
app.register_blueprint(address_bp)

# API Orders Blueprint
from apps.api.orders import orders_bp
app.register_blueprint(orders_bp)

# API Notifications Blueprint (ERP 알림, Phase 4-1)
from apps.api.notifications import notifications_bp
app.register_blueprint(notifications_bp)


# Error handler with production safety
@app.errorhandler(500)
def internal_error(error):
    import traceback
    # Only show detailed errors in development
    if app.debug or os.environ.get('FLASK_ENV') != 'production':
        return f"<pre>500 Error: {str(error)}\n\n{traceback.format_exc()}</pre>", 500
    else:
        # Production: Log error but show generic message
        app.logger.error(f"Internal Server Error: {str(error)}\n{traceback.format_exc()}")
        return render_template('error_500.html'), 500


@app.get('/__build')
def build_info():
    return jsonify({
        'build': '20260215-uxfix-03',
        'cwd': os.getcwd(),
        'template': 'templates/layout.html',
    })

# ==========================================
# Security & Scalability Configuration (Quest 14)
# ==========================================

# 1. Redis Configuration
redis_url = os.environ.get('REDIS_URL')

# 2. Rate Limiter (DDoS Protection)
# Redis가 있으면 Redis를 저장소로, 없으면 메모리 사용
# NOTE:
# - Railway/Reverse proxy 환경에서는 get_remote_address()가 내부 IP로 고정될 수 있어
#   사용자들이 같은 rate-limit bucket을 공유하게 됩니다.
# - 인증 사용자(user_id) -> 세션쿠키 해시 -> X-Forwarded-For -> Remote Addr 순서로 key를 선택해
#   불필요한 429를 줄입니다.
def rate_limit_key():
    try:
        uid = session.get('user_id')
        if uid:
            return f"user:{uid}"
    except Exception:
        pass

    # 로그인 세션이 있으나 user_id를 복원하지 못한 경우(예: 세션 초기화 타이밍/리다이렉트 직후)
    # 동일 IP 공유로 bucket 충돌이 발생하지 않도록 세션쿠키 해시를 key로 사용
    try:
        cookie_name = app.config.get('SESSION_COOKIE_NAME', 'session')
        raw_cookie = request.cookies.get(cookie_name, '').strip()
        if raw_cookie:
            cookie_hash = hashlib.sha1(raw_cookie.encode('utf-8')).hexdigest()[:16]
            return f"sess:{cookie_hash}"
    except Exception:
        pass

    xff = request.headers.get('X-Forwarded-For', '')
    if xff:
        client_ip = xff.split(',')[0].strip()
        if client_ip:
            return client_ip

    x_real_ip = request.headers.get('X-Real-IP', '').strip()
    if x_real_ip:
        return x_real_ip

    return get_remote_address()

_default_limits_raw = os.environ.get('FLASK_DEFAULT_RATE_LIMITS', '5000 per day,1200 per hour')
default_limits = [x.strip() for x in _default_limits_raw.split(',') if x.strip()]
if not default_limits:
    default_limits = ["5000 per day", "1200 per hour"]

limiter = Limiter(
    rate_limit_key,
    app=app,
    storage_uri=redis_url or "memory://",
    default_limits=default_limits
)

# 3. SocketIO Initialization with Redis & CORS Control
# Use threading mode for Windows WebSocket support & Stability
if SOCKETIO_AVAILABLE:
    try:
        # CORS 도메인 제한 (환경변수 없으면 모든 도메인 허용 - 개발 편의성)
        allowed_origins = os.environ.get('CORS_ALLOWED_ORIGINS', '*').split(',')
        
        # Redis가 설정되어 있으면 Message Queue로 사용 (Scale-out 지원)
        if redis_url:
            print(f"[INFO] Socket.IO connecting to Redis Message Queue: {redis_url}")
            socketio = SocketIO(
                app, 
                cors_allowed_origins=allowed_origins, 
                async_mode='threading',
                message_queue=redis_url
            )
            print("[INFO] Socket.IO initialized in threading mode with Redis.")
        else:
            print("[WARN] REDIS_URL not found. Socket.IO running in single-worker mode (Memory).")
            socketio = SocketIO(
                app, 
                cors_allowed_origins=allowed_origins, 
                async_mode='threading'
            )
            print("[INFO] Socket.IO initialized in threading mode (Universal Stable).")
            
    except Exception as e:
        # fallback
        print(f"[WARN] Socket.IO init failed: {e}")
        socketio = SocketIO(app, cors_allowed_origins="*", async_mode='threading')
else:
    socketio = None

# 템플릿 캐시 비활성화 (개발 중 변경사항 즉시 반영)
app.config['TEMPLATES_AUTO_RELOAD'] = True

# Extensions config moved to constants.py

app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
from flask import Response
app.config['MAX_CONTENT_LENGTH'] = 500 * 1024 * 1024  # 500MB (Restore video support for Pro Plan)
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

# 데이터베이스 연결 설정
app.teardown_appcontext(close_db)
app.teardown_appcontext(close_wdcalculator_db)  # 견적 계산기 독립 DB

# Function to check if file has allowed extension
def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

# 채팅 파일 확장자 검증 (Quest 3)
def allowed_chat_file(filename):
    """채팅용 파일 확장자 검증"""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in CHAT_ALLOWED_EXTENSIONS

def get_chat_file_max_size(filename):
    """채팅 파일 타입별 최대 크기 제한 (바이트)"""
    if not '.' in filename:
        return 10 * 1024 * 1024  # 기본 10MB
    
    ext = filename.rsplit('.', 1)[1].lower()
    image_exts = ['jpg', 'jpeg', 'png', 'gif', 'webp']
    video_exts = ['mp4', 'mov', 'avi', 'mkv', 'webm']
    
    if ext in image_exts:
        return 10 * 1024 * 1024  # 10MB (사진)
    elif ext in video_exts:
        return 500 * 1024 * 1024  # 500MB (동영상)
    else:
        return 50 * 1024 * 1024  # 50MB (기타 파일)

def allowed_erp_media_file(filename):
    """ERP Beta 첨부(사진/동영상) 확장자 검증"""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ERP_MEDIA_ALLOWED_EXTENSIONS


# 도면 카테고리는 이미지/동영상 외 문서 파일도 허용
DRAWING_ATTACHMENT_EXTRA_EXTENSIONS = {'pdf', 'zip', 'dwg', 'dxf'}


def allowed_erp_attachment_file(filename, category='measurement'):
    """ERP Beta 첨부 확장자 검증 (카테고리별 정책)"""
    if '.' not in filename:
        return False
    ext = filename.rsplit('.', 1)[1].lower()
    allowed_exts = set(ERP_MEDIA_ALLOWED_EXTENSIONS)
    if normalize_attachment_category(category) == 'drawing':
        allowed_exts.update(DRAWING_ATTACHMENT_EXTRA_EXTENSIONS)
    return ext in allowed_exts

def get_erp_media_max_size(filename):
    """ERP Beta 첨부 파일 타입별 최대 크기 (바이트)"""
    if not '.' in filename:
        return 10 * 1024 * 1024
    ext = filename.rsplit('.', 1)[1].lower()
    image_exts = ['jpg', 'jpeg', 'png', 'gif', 'webp']
    video_exts = ['mp4', 'mov', 'avi', 'mkv', 'webm']
    if ext in image_exts:
        return 20 * 1024 * 1024  # 20MB
    if ext in video_exts:
        return 500 * 1024 * 1024  # 500MB
    return 20 * 1024 * 1024

def build_file_view_url(storage_key: str) -> str:
    return f"/api/files/view/{storage_key}"

def build_file_download_url(storage_key: str) -> str:
    return f"/api/files/download/{storage_key}"


# 채팅 이미지 썸네일 비동기 생성기 (요청 스레드 블로킹 방지)
_thumb_workers = int(os.environ.get('CHAT_THUMBNAIL_WORKERS', '2') or 2)
_thumb_workers = max(1, min(_thumb_workers, 4))
chat_thumbnail_executor = ThreadPoolExecutor(max_workers=_thumb_workers)


def _generate_chat_thumbnail_background(storage_key: str):
    """storage_key 기준으로 썸네일 생성 후 ChatAttachment.thumbnail_url 업데이트"""
    if not storage_key:
        return
    try:
        storage = get_storage()
        result = storage.generate_thumbnail_from_storage_key(storage_key)
        if not result.get('success'):
            return

        thumbnail_key = result.get('thumbnail_key')
        if not thumbnail_key:
            return

        attachment_db = db_session()
        try:
            attachment = attachment_db.query(ChatAttachment).filter(
                ChatAttachment.storage_key == storage_key
            ).order_by(ChatAttachment.id.desc()).first()
            if attachment and not attachment.thumbnail_url:
                attachment.thumbnail_url = build_file_view_url(thumbnail_key)
                attachment_db.commit()
        finally:
            attachment_db.close()
            db_session.remove()
    except Exception as e:
        print(f"[ChatThumbnail] background generation error: {e}")


def schedule_chat_thumbnail_generation(storage_key: str):
    """채팅 썸네일 비동기 작업 큐잉"""
    if not storage_key:
        return
    try:
        chat_thumbnail_executor.submit(_generate_chat_thumbnail_background, storage_key)
    except Exception as e:
        print(f"[ChatThumbnail] schedule error: {e}")

# Status constants moved to constants.py

@app.template_filter('parse_json_string')
def parse_json_string(value):
    if not value:
        return {}
    if isinstance(value, dict):
        return value
    try:
        return json.loads(value)
    except (ValueError, TypeError):
        return {}

@app.context_processor
def inject_statuses():
    return dict(
        ALL_STATUS=STATUS,
        BULK_ACTION_STATUS=BULK_ACTION_STATUS
    )

# Auth logic moved to apps/auth.py

# Auth helper functions moved to apps/auth.py

# role_required moved to apps/auth.py

# ============================================
# ERP Process Dashboard (Palantir-style)
# ============================================

def _ensure_dict(data):
    """Ensure data is a dict, properly parsing stringified JSON if needed (migration fix)"""
    if isinstance(data, dict):
        return data
    if isinstance(data, str):
        try:
            return json.loads(data)
        except:
            return {}
    return {}


ATTACHMENT_CATEGORIES = ('measurement', 'drawing', 'construction')


def normalize_attachment_category(raw_category):
    category = (raw_category or 'measurement').strip().lower()
    if category not in ATTACHMENT_CATEGORIES:
        return None
    return category


def parse_attachment_item_index(raw_item_index):
    """Parse optional item index for product-level measurement attachments."""
    if raw_item_index is None:
        return True, None, None
    s = str(raw_item_index).strip().lower()
    if s in ('', 'null', 'none'):
        return True, None, None
    try:
        value = int(s)
    except (TypeError, ValueError):
        return False, None, 'item_index는 0 이상의 정수 또는 null 이어야 합니다.'
    if value < 0:
        return False, None, 'item_index는 0 이상의 정수 또는 null 이어야 합니다.'
    return True, value, None


def ensure_order_attachments_category_column():
    """Ensure category column exists for legacy databases."""
    db = None
    try:
        db = get_db()
        db.execute(text(
            "ALTER TABLE order_attachments "
            "ADD COLUMN IF NOT EXISTS category VARCHAR(50) NOT NULL DEFAULT 'measurement'"
        ))
        db.commit()
        return True
    except Exception as e:
        try:
            if db is not None:
                db.rollback()
        except Exception:
            pass
        print(f"[AUTO-MIGRATION] Failed to ensure order_attachments.category: {e}")
        return False


def ensure_order_attachments_item_index_column():
    """Ensure item_index column exists for product-linked measurement attachments."""
    db = None
    try:
        db = get_db()
        db.execute(text(
            "ALTER TABLE order_attachments "
            "ADD COLUMN IF NOT EXISTS item_index INTEGER NULL"
        ))
        db.commit()
        return True
    except Exception as e:
        try:
            if db is not None:
                db.rollback()
        except Exception:
            pass
        print(f"[AUTO-MIGRATION] Failed to ensure order_attachments.item_index: {e}")
        return False

@app.route('/api/orders/<int:order_id>/attachments', methods=['GET'])
@login_required
def api_order_attachments_list(order_id):
    """주문 첨부 목록(ERP Beta 사진/동영상)"""
    try:
        db = get_db()
        order = db.query(Order).filter(Order.id == order_id).first()
        if not order:
            return jsonify({'success': False, 'message': '주문을 찾을 수 없습니다.'}), 404

        raw_filter_category = request.args.get('category')
        filter_category = normalize_attachment_category(raw_filter_category) if raw_filter_category else None
        if raw_filter_category and not filter_category:
            return jsonify({'success': False, 'message': '유효하지 않은 첨부 카테고리입니다.'}), 400
        raw_filter_item_index = request.args.get('item_index')
        filter_item_index = None
        has_item_filter = raw_filter_item_index is not None
        if has_item_filter:
            ok, filter_item_index, err = parse_attachment_item_index(raw_filter_item_index)
            if not ok:
                return jsonify({'success': False, 'message': err}), 400

        query = db.query(OrderAttachment).filter(OrderAttachment.order_id == order_id)
        if filter_category:
            query = query.filter(OrderAttachment.category == filter_category)
        if has_item_filter:
            if filter_item_index is None:
                query = query.filter(OrderAttachment.item_index.is_(None))
            else:
                query = query.filter(OrderAttachment.item_index == filter_item_index)

        atts = query.order_by(OrderAttachment.created_at.desc()).all()
        items = []
        for a in atts:
            d = a.to_dict()
            d['category'] = normalize_attachment_category(d.get('category')) or 'measurement'
            d['view_url'] = build_file_view_url(a.storage_key)
            d['download_url'] = build_file_download_url(a.storage_key)
            d['thumbnail_view_url'] = build_file_view_url(a.thumbnail_key) if a.thumbnail_key else None
            items.append(d)

        return jsonify({'success': True, 'attachments': items})
    except Exception as e:
        import traceback
        print(f"주문 첨부 목록 오류: {e}")
        print(traceback.format_exc())
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/orders/<int:order_id>/attachments', methods=['POST'])
@login_required
def api_order_attachments_upload(order_id):
    """주문 첨부 업로드(ERP Beta 사진/동영상)"""
    try:
        if 'file' not in request.files:
            return jsonify({'success': False, 'message': '파일이 없습니다.'}), 400
        file = request.files['file']
        if not file or file.filename == '':
            return jsonify({'success': False, 'message': '파일명이 없습니다.'}), 400

        category = normalize_attachment_category(request.form.get('category', 'measurement'))
        if not category:
            return jsonify({'success': False, 'message': '유효하지 않은 첨부 카테고리입니다.'}), 400
        ok, item_index, err = parse_attachment_item_index(request.form.get('item_index'))
        if not ok:
            return jsonify({'success': False, 'message': err}), 400

        if not allowed_erp_attachment_file(file.filename, category):
            allowed_exts = set(ERP_MEDIA_ALLOWED_EXTENSIONS)
            if category == 'drawing':
                allowed_exts.update(DRAWING_ATTACHMENT_EXTRA_EXTENSIONS)
            allowed_exts = ', '.join(sorted(allowed_exts))
            return jsonify({'success': False, 'message': f'허용되지 않은 파일 형식입니다. 지원 형식: {allowed_exts}'}), 400

        # 파일 크기 검증
        file.seek(0, os.SEEK_END)
        file_size = file.tell()
        file.seek(0)
        max_size = get_erp_media_max_size(file.filename)
        if file_size > max_size:
            size_mb = max_size / (1024 * 1024)
            return jsonify({'success': False, 'message': f'파일 크기가 너무 큽니다. 최대 {size_mb:.0f}MB까지 업로드 가능합니다.'}), 400

        db = get_db()
        order = db.query(Order).filter(Order.id == order_id).first()
        if not order:
            return jsonify({'success': False, 'message': '주문을 찾을 수 없습니다.'}), 404

        storage = get_storage()
        folder = f"orders/{order_id}/attachments"

        # 원본 업로드(스토리지가 unique filename을 생성)
        result = storage.upload_file(file, file.filename, folder)
        if not result.get('success'):
            return jsonify({'success': False, 'message': '파일 업로드 실패: ' + result.get('message', '알 수 없는 오류')}), 500

        storage_key = result.get('key')
        filename = file.filename
        file_type = storage._get_file_type(filename)
        if category == 'drawing':
            if file_type not in ['image', 'video', 'file']:
                return jsonify({'success': False, 'message': '지원되지 않는 도면 파일 형식입니다.'}), 400
        else:
            if file_type not in ['image', 'video']:
                return jsonify({'success': False, 'message': '이미지/동영상만 업로드 가능합니다.'}), 400

        # 이미지 썸네일 생성(가능할 때만)
        thumbnail_key = None
        try:
            if file_type == 'image' and hasattr(storage, '_generate_thumbnail'):
                unique_filename = storage_key.rsplit('/', 1)[-1] if storage_key else None
                if unique_filename:
                    file.seek(0)
                    storage._generate_thumbnail(file, unique_filename, folder, 'image', storage_key=storage_key)
                    thumbnail_key = f"{folder}/thumb_{unique_filename}"
        except Exception:
            thumbnail_key = None

        att = OrderAttachment(
            order_id=order_id,
            filename=filename,
            file_type=file_type,
            category=category,
            item_index=item_index,
            file_size=file_size,
            storage_key=storage_key,
            thumbnail_key=thumbnail_key
        )
        db.add(att)
        db.commit()
        db.refresh(att)

        d = att.to_dict()
        d['view_url'] = build_file_view_url(att.storage_key)
        d['download_url'] = build_file_download_url(att.storage_key)
        d['thumbnail_view_url'] = build_file_view_url(att.thumbnail_key) if att.thumbnail_key else None

        return jsonify({'success': True, 'attachment': d})
    except Exception as e:
        db = get_db()
        try:
            db.rollback()
        except Exception:
            pass
        import traceback
        print(f"주문 첨부 업로드 오류: {e}")
        print(traceback.format_exc())
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/orders/<int:order_id>/attachments/<int:attachment_id>', methods=['PATCH'])
@login_required
def api_order_attachments_patch(order_id, attachment_id):
    """주문 첨부 메타 수정(제품 항목 연결/해제)"""
    try:
        payload = request.get_json(silent=True) or {}
        if 'item_index' not in payload:
            return jsonify({'success': False, 'message': 'item_index 필드가 필요합니다.'}), 400
        ok, item_index, err = parse_attachment_item_index(payload.get('item_index'))
        if not ok:
            return jsonify({'success': False, 'message': err}), 400

        db = get_db()
        att = db.query(OrderAttachment).filter(
            OrderAttachment.id == attachment_id,
            OrderAttachment.order_id == order_id
        ).first()
        if not att:
            return jsonify({'success': False, 'message': '첨부파일을 찾을 수 없습니다.'}), 404

        att.item_index = item_index
        db.commit()
        db.refresh(att)

        d = att.to_dict()
        d['category'] = normalize_attachment_category(d.get('category')) or 'measurement'
        d['view_url'] = build_file_view_url(att.storage_key)
        d['download_url'] = build_file_download_url(att.storage_key)
        d['thumbnail_view_url'] = build_file_view_url(att.thumbnail_key) if att.thumbnail_key else None
        return jsonify({'success': True, 'attachment': d})
    except Exception as e:
        db = get_db()
        try:
            db.rollback()
        except Exception:
            pass
        import traceback
        print(f"주문 첨부 수정 오류: {e}")
        print(traceback.format_exc())
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/orders/<int:order_id>/attachments/<int:attachment_id>', methods=['DELETE'])
@login_required
def api_order_attachments_delete(order_id, attachment_id):
    """주문 첨부 삭제(ERP Beta)"""
    try:
        db = get_db()
        att = db.query(OrderAttachment).filter(
            OrderAttachment.id == attachment_id,
            OrderAttachment.order_id == order_id
        ).first()
        if not att:
            return jsonify({'success': False, 'message': '첨부파일을 찾을 수 없습니다.'}), 404

        storage = get_storage()
        try:
            if att.storage_key:
                storage.delete_file(att.storage_key)
            if att.thumbnail_key:
                storage.delete_file(att.thumbnail_key)
        except Exception:
            pass

        db.delete(att)
        db.commit()
        return jsonify({'success': True})
    except Exception as e:
        db = get_db()
        try:
            db.rollback()
        except Exception:
            pass
        import traceback
        print(f"주문 첨부 삭제 오류: {e}")
        print(traceback.format_exc())
        return jsonify({'success': False, 'message': str(e)}), 500

# ============================================
# ERP: Events & Tasks API (Palantir-style)
# ============================================

@app.route('/api/orders/<int:order_id>/events', methods=['GET'])
@login_required
def api_order_events(order_id):
    """주문 이벤트 스트림 조회(최근 N개)"""
    try:
        db = get_db()
        limit = int(request.args.get('limit', 50))
        limit = max(1, min(limit, 200))

        rows = db.query(OrderEvent).filter(OrderEvent.order_id == order_id).order_by(OrderEvent.created_at.desc()).limit(limit).all()
        events = []
        for r in rows:
            events.append({
                'id': r.id,
                'order_id': r.order_id,
                'event_type': r.event_type,
                'payload': r.payload,
                'created_by_user_id': r.created_by_user_id,
                'created_at': r.created_at.strftime('%Y-%m-%d %H:%M:%S') if r.created_at else None
            })
        return jsonify({'success': True, 'events': events})
    except Exception as e:
        import traceback
        print(f"주문 이벤트 조회 오류: {e}")
        print(traceback.format_exc())
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/orders/<int:order_id>/change-events', methods=['GET'])
@login_required
def api_order_change_events(order_id):
    """
    변경 이벤트 로그 조회 (V2 설계)
    - ADMIN: 전체 조회
    - 일반: 본인이 생성한 로그만 조회
    """
    try:
        db = get_db()
        user_id = session.get('user_id')
        user = db.query(User).filter(User.id == user_id).first()
        if not user:
            return jsonify({'success': False, 'message': '사용자를 찾을 수 없습니다.'}), 401
        
        # 기본 쿼리
        query = db.query(OrderEvent).filter(OrderEvent.order_id == order_id)
        
        # 권한 필터링
        if user.role != 'ADMIN':
            query = query.filter(OrderEvent.created_by_user_id == user_id)
        
        # 최신순 정렬
        limit = int(request.args.get('limit', 100))
        limit = max(1, min(limit, 500))
        rows = query.order_by(OrderEvent.created_at.desc()).limit(limit).all()
        
        # 사용자 정보 조인 (효율을 위해 일괄 조회)
        user_ids = list(set([r.created_by_user_id for r in rows if r.created_by_user_id]))
        users_map = {}
        if user_ids:
            users = db.query(User).filter(User.id.in_(user_ids)).all()
            users_map = {u.id: {'name': u.name, 'team': u.team} for u in users}
        
        # 주문 정보 조회 (표시용 고객명)
        order = db.query(Order).filter(Order.id == order_id).first()
        customer_name = get_order_display_name(order) if order else f'주문 #{order_id}'
        
        # 응답 포맷 (누가/무엇을/어떻게/언제)
        events = []
        for r in rows:
            payload = r.payload or {}
            creator = users_map.get(r.created_by_user_id, {'name': 'Unknown', 'team': ''})
            
            # 이벤트 타입 한글화
            event_label = translate_event_type_to_korean(r.event_type)
            
            # 타겟과 값 한글화
            target = payload.get('target', '')
            before = payload.get('before', '')
            after = payload.get('after', '')
            reason = translate_reason_to_korean(payload.get('reason', ''), r.event_type, payload)
            is_override = payload.get('is_override', False)
            
            target_kr = translate_target_to_korean(target)
            before_kr = translate_value_to_korean(target, before)
            after_kr = translate_value_to_korean(target, after)
            
            # 간단하고 이해하기 쉬운 설명 생성
            how_text = generate_change_description(r.event_type, target_kr, before_kr, after_kr, payload)
            
            events.append({
                'id': r.id,
                'when': r.created_at.strftime('%Y-%m-%d %H:%M:%S') if r.created_at else '',
                'who_name': creator['name'],
                'who_team': creator['team'],
                'what_label': event_label,
                'how_text': how_text,
                'reason': reason,
                'is_override': is_override,
                'override_reason': payload.get('override_reason'),
                'event_type': r.event_type,
                'payload': payload,
            })
        
        return jsonify({
            'success': True, 
            'events': events, 
            'total': len(events),
            'customer_name': customer_name,
            'order_id': order_id
        })
        
    except Exception as e:
        import traceback
        print(f"변경 이벤트 조회 오류: {e}")
        print(traceback.format_exc())
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/me/change-events', methods=['GET'])
@login_required
def api_my_change_events():
    """
    본인의 전체 변경 이벤트 로그 조회 (여러 주문 통합)
    """
    try:
        db = get_db()
        user_id = session.get('user_id')
        user = db.query(User).filter(User.id == user_id).first()
        if not user:
            return jsonify({'success': False, 'message': '사용자를 찾을 수 없습니다.'}), 401
        
        # 본인이 생성한 이벤트만 조회
        limit = int(request.args.get('limit', 200))
        limit = max(1, min(limit, 1000))
        
        rows = db.query(OrderEvent).filter(
            OrderEvent.created_by_user_id == user_id
        ).order_by(OrderEvent.created_at.desc()).limit(limit).all()
        
        # 주문 정보 조인 (표시용 고객명 포함)
        order_ids = list(set([r.order_id for r in rows if r.order_id]))
        orders_map = {}
        if order_ids:
            orders = db.query(Order).filter(Order.id.in_(order_ids)).all()
            orders_map = {
                o.id: {
                    'customer_name': get_order_display_name(o),
                    'order_id': o.id
                } 
                for o in orders
            }
        
        # 응답 포맷
        events = []
        for r in rows:
            payload = r.payload or {}
            order_info = orders_map.get(r.order_id, {'customer_name': f'주문 #{r.order_id}', 'order_id': r.order_id})
            
            # 이벤트 타입을 한글로 변환
            event_label = translate_event_type_to_korean(r.event_type)
            
            # payload에서 더 자세한 정보 추출
            action_label = payload.get('action', event_label)
            target = payload.get('target', '')
            before = payload.get('before', '')
            after = payload.get('after', '')
            reason = translate_reason_to_korean(payload.get('reason', ''), r.event_type, payload)
            is_override = payload.get('is_override', False)
            
            # 한글로 변환된 타겟과 값
            target_kr = translate_target_to_korean(target)
            before_kr = translate_value_to_korean(target, before)
            after_kr = translate_value_to_korean(target, after)
            
            # 간단하고 이해하기 쉬운 설명 생성
            how_text = generate_change_description(r.event_type, target_kr, before_kr, after_kr, payload)
            
            events.append({
                'id': r.id,
                'order_id': r.order_id,
                'customer_name': order_info['customer_name'],
                'when': r.created_at.strftime('%Y-%m-%d %H:%M:%S') if r.created_at else '',
                'what_label': event_label,
                'how_text': how_text,
                'reason': reason,
                'is_override': is_override,
                'event_type': r.event_type,
            })
        
        return jsonify({'success': True, 'events': events, 'total': len(events)})
        
    except Exception as e:
        import traceback
        print(f"내 변경 이벤트 조회 오류: {e}")
        print(traceback.format_exc())
        return jsonify({'success': False, 'message': str(e)}), 500


def translate_target_to_korean(target):
    """영어 타겟을 한글로 변환"""
    target_map = {
        'workflow.stage': '진행 단계',
        'workflow.current_quest': '현재 퀘스트',
        'quests': '퀘스트',
        'quest.team_approvals': '팀 승인',
        'quest.assignee_approval': '담당자 승인',
        'assignments.drawing_assignee_user_ids': '도면 담당자',
        'drawings.status': '도면 상태',
        'production.completed': '생산 완료',
        'construction.completed': '시공 완료',
        'cs.completed': 'CS 완료',
        'as.status': 'AS 상태',
    }
    return target_map.get(target, target)


def get_order_display_name(order):
    """로그 카드에 표시할 주문명(고객명)을 최대한 정확히 추출."""
    if not order:
        return ''

    generic_names = {'erp beta', 'erp_beta', 'beta'}

    def _clean_text(value):
        if value is None:
            return ''
        text = str(value).strip()
        if not text:
            return ''
        if text.lower() in generic_names:
            return ''
        return text

    sd = order.structured_data if isinstance(order.structured_data, dict) else {}
    customer = sd.get('customer') if isinstance(sd.get('customer'), dict) else {}
    orderer = sd.get('orderer') if isinstance(sd.get('orderer'), dict) else {}
    contact = sd.get('contact') if isinstance(sd.get('contact'), dict) else {}
    order_node = sd.get('order') if isinstance(sd.get('order'), dict) else {}
    parties = sd.get('parties') if isinstance(sd.get('parties'), dict) else {}
    parties_customer = parties.get('customer') if isinstance(parties.get('customer'), dict) else {}
    parties_orderer = parties.get('orderer') if isinstance(parties.get('orderer'), dict) else {}
    parties_manager = parties.get('manager') if isinstance(parties.get('manager'), dict) else {}

    candidates = [
        parties_customer.get('name'),
        parties_customer.get('customer_name'),
        parties_orderer.get('name'),
        parties_manager.get('name'),
        customer.get('name'),
        customer.get('customer_name'),
        orderer.get('name'),
        contact.get('name'),
        sd.get('client_name'),
        sd.get('client'),
        sd.get('name'),
        sd.get('customer'),
        sd.get('customer_name'),
        sd.get('orderer_name'),
        order_node.get('customer_name'),
        order.customer_name,
    ]
    for candidate in candidates:
        name = _clean_text(candidate)
        if name:
            return name
    return f'주문 #{order.id}'


def translate_event_type_to_korean(event_type):
    """이벤트 타입 영문 코드를 한글 라벨로 변환."""
    labels = {
        # Quest
        'QUEST_APPROVAL_CHANGED': '퀘스트 승인',
        'QUEST_ASSIGNEE_APPROVED': '담당자 승인',
        'QUEST_CREATED': '퀘스트 생성',
        'QUEST_UPDATED': '퀘스트 수정',
        'QUEST_COMPLETED': '퀘스트 완료',
        # Stage
        'STAGE_CHANGED': '단계 변경',
        'STAGE_AUTO_TRANSITIONED': '단계 자동 전환',
        'STAGE_MANUAL_OVERRIDE': '단계 수동 변경',
        # Drawing
        'DRAWING_STATUS_CHANGED': '도면 상태 변경',
        'DRAWING_ASSIGNEE_SET': '도면 담당자 지정',
        'DRAWING_SENT': '도면 전달',
        'DRAWING_CONFIRMED': '도면 확인',
        'DRAWING_REVISION_REQUESTED': '도면 수정 요청',
        # Production / Construction / CS / AS
        'PRODUCTION_STARTED': '생산 시작',
        'PRODUCTION_COMPLETED': '생산 완료',
        'PRODUCTION_DELAYED': '생산 지연',
        'CONSTRUCTION_STARTED': '시공 시작',
        'CONSTRUCTION_COMPLETED': '시공 완료',
        'CONSTRUCTION_SCHEDULED': '시공 예약',
        'CS_STARTED': 'CS 시작',
        'CS_COMPLETED': 'CS 완료',
        'CS_ISSUE_REPORTED': 'CS 이슈 보고',
        'AS_STARTED': 'AS 시작',
        'AS_COMPLETED': 'AS 완료',
        'AS_RECEIVED': 'AS 접수',
        # Measurement / Shipment
        'MEASUREMENT_SCHEDULED': '실측 예약',
        'MEASUREMENT_COMPLETED': '실측 완료',
        'SHIPMENT_SCHEDULED': '출고 예정',
        'SHIPMENT_COMPLETED': '출고 완료',
        # Etc
        'CHANGE_REVERTED': '변경 되돌림',
        'ORDER_CREATED': '주문 생성',
        'ORDER_UPDATED': '주문 수정',
        'ORDER_DELETED': '주문 삭제',
        'ASSIGNMENT_CHANGED': '담당자 변경',
        'STATUS_CHANGED': '상태 변경',
        'FIELD_UPDATED': '필드 수정',
        'COMMENT_ADDED': '메모 추가',
        'ATTACHMENT_ADDED': '첨부파일 추가',
        'ATTACHMENT_DELETED': '첨부파일 삭제',
        'URGENT_CHANGED': '긴급 여부 변경',
    }
    return labels.get(event_type, '기타 변경')


def translate_reason_to_korean(reason, event_type='', payload=None):
    """시스템 reason 코드를 사람이 이해하기 쉬운 한글로 변환."""
    payload = payload or {}
    raw = str(reason or '').strip()
    if not raw:
        raw = str(payload.get('override_reason') or '').strip()
    if not raw and event_type == 'STAGE_AUTO_TRANSITIONED':
        raw = 'quest_approvals_complete'

    if not raw:
        return ''

    reason_map = {
        'quest_approvals_complete': '퀘스트 승인 조건이 충족되어 자동 전환되었습니다.',
        'all_approvals_complete': '모든 승인 조건이 충족되었습니다.',
        'auto_transition_rule_matched': '자동 단계 전환 규칙에 따라 처리되었습니다.',
        'manager_override': '관리자 권한으로 예외 처리되었습니다.',
        'emergency_override': '긴급 권한으로 예외 처리되었습니다.',
        'manual_update': '담당자가 수동으로 변경했습니다.',
    }
    if raw in reason_map:
        return reason_map[raw]

    if '_' in raw and raw.islower():
        return raw.replace('_', ' ')
    return raw


def translate_value_to_korean(target, value):
    """값을 한글로 변환"""
    if not value:
        return '없음'
    
    # Boolean 값 변환
    if isinstance(value, bool):
        return '완료' if value else '미완료'
    
    # 단계(stage) 값 변환
    if 'stage' in target.lower():
        stage_map = {
            'MEASURE': '실측',
            'DRAWING': '도면',
            'CONFIRM': '고객확인',
            'PRODUCTION': '생산',
            'CONSTRUCTION': '시공',
            'CS': 'CS',
            'AS': 'AS',
            'SHIPMENT': '출고',
        }
        return stage_map.get(str(value), value)
    
    # 승인 상태 변환
    if 'approval' in target.lower():
        if isinstance(value, dict):
            if value.get('approved'):
                return f"승인됨 ({value.get('approved_by_name', '담당자')})"
            return '미승인'
        return '승인됨' if value else '미승인'
    
    # 도면 상태 변환
    if 'drawing' in target.lower() and 'status' in target.lower():
        status_map = {
            'pending': '대기중',
            'sent': '전달됨',
            'confirmed': '확인완료',
            'revision_requested': '수정요청',
        }
        return status_map.get(str(value), value)
    
    return str(value)


def generate_change_description(event_type, target_kr, before_kr, after_kr, payload):
    """이벤트 타입에 따라 이해하기 쉬운 설명 생성"""
    
    # 퀘스트 승인 관련
    if event_type == 'QUEST_APPROVAL_CHANGED':
        team = payload.get('team', '')
        team_map = {'CS': 'CS팀', 'SALES': '영업팀', 'DRAWING': '도면팀', 
                   'PRODUCTION': '생산팀', 'CONSTRUCTION': '시공팀', 'SHIPMENT': '출고팀'}
        team_kr = team_map.get(team, team)
        return f"{team_kr}이 퀘스트를 승인했습니다"
    
    if event_type == 'QUEST_ASSIGNEE_APPROVED':
        approved_by = payload.get('approved_by_name', '담당자')
        quest_title = payload.get('quest_title', '')
        return f"{approved_by}님이 '{quest_title}' 퀘스트를 승인했습니다"
    
    # 단계 변경
    if event_type == 'STAGE_CHANGED':
        return f"진행 단계를 '{before_kr}'에서 '{after_kr}'로 변경했습니다"
    
    if event_type == 'STAGE_AUTO_TRANSITIONED':
        return f"퀘스트 완료로 인해 단계가 '{before_kr}'에서 '{after_kr}'로 자동 전환되었습니다"
    
    # 도면 관련
    if event_type == 'DRAWING_ASSIGNEE_SET':
        assignees = payload.get('assignee_names', [])
        if not assignees:
            after_raw = payload.get('after')
            if isinstance(after_raw, str):
                assignees = [x.strip() for x in after_raw.split(',') if x.strip() and x.strip().lower() != 'none']
        if assignees:
            return f"도면 담당자를 {', '.join(assignees)}님으로 지정했습니다"
        return "도면 담당자를 지정했습니다"
    
    if event_type == 'DRAWING_STATUS_CHANGED':
        return f"도면 상태를 '{before_kr}'에서 '{after_kr}'로 변경했습니다"
    
    # 완료 이벤트
    if event_type == 'PRODUCTION_COMPLETED':
        return "생산을 완료 처리했습니다"
    
    if event_type == 'PRODUCTION_STARTED':
        return "생산을 시작했습니다"
    
    if event_type == 'CONSTRUCTION_COMPLETED':
        return "시공을 완료 처리했습니다"
    
    if event_type == 'CONSTRUCTION_STARTED':
        return "시공을 시작했습니다"
    
    if event_type == 'CS_COMPLETED':
        return "CS를 완료 처리했습니다"
    
    if event_type == 'CS_STARTED':
        return "CS를 시작했습니다"
    
    if event_type == 'AS_STARTED':
        return "AS를 시작했습니다"
    
    if event_type == 'AS_COMPLETED':
        return "AS를 완료 처리했습니다"
    
    if event_type == 'AS_RECEIVED':
        return "AS를 접수했습니다"
    
    # 측정 관련
    if event_type == 'MEASUREMENT_SCHEDULED':
        return "실측 일정을 등록했습니다"
    
    if event_type == 'MEASUREMENT_COMPLETED':
        return "실측을 완료했습니다"
    
    # 출고 관련
    if event_type == 'SHIPMENT_SCHEDULED':
        return "출고 일정을 등록했습니다"
    
    if event_type == 'SHIPMENT_COMPLETED':
        return "출고를 완료했습니다"
    
    # 되돌림
    if event_type == 'CHANGE_REVERTED':
        original_type = payload.get('original_event_type', '')
        return f"이전 변경사항을 되돌렸습니다 ({translate_target_to_korean(payload.get('target', ''))})"
    
    # 주문 관련
    if event_type == 'ORDER_CREATED':
        return "주문을 생성했습니다"
    
    if event_type == 'ORDER_UPDATED':
        return "주문 정보를 수정했습니다"
    
    if event_type == 'ORDER_DELETED':
        return "주문을 삭제했습니다"
    
    # 담당자 변경
    if event_type == 'ASSIGNMENT_CHANGED':
        return f"담당자를 변경했습니다"

    if event_type == 'URGENT_CHANGED':
        return "긴급 여부를 변경했습니다"
    
    # 기본 포맷
    if target_kr and before_kr and after_kr:
        return f"{target_kr}를 '{before_kr}'에서 '{after_kr}'로 변경했습니다"
    
    if target_kr:
        return f"{target_kr} 변경"
    
    return "변경 이력"


@app.route('/api/orders/<int:order_id>/change-events/<int:event_id>/revert', methods=['POST'])
@login_required
def api_revert_change_event(order_id, event_id):
    """
    변경 이벤트 되돌리기 (Rollback)
    
    권한:
    - 본인이 생성한 이벤트만 되돌리기 가능
    - ADMIN은 모든 이벤트 되돌리기 가능
    
    제약:
    - 일부 이벤트 타입만 지원 (STAGE_CHANGED, QUEST_APPROVAL, DRAWING_STATUS 등)
    - 되돌리기 시 새로운 OrderEvent를 생성하여 기록
    """
    try:
        data = request.get_json() or {}
        revert_reason = data.get('reason', '').strip()
        
        if not revert_reason:
            return jsonify({'success': False, 'message': '되돌리기 사유를 입력해주세요.'}), 400
        
        db = get_db()
        user_id = session.get('user_id')
        user = db.query(User).filter(User.id == user_id).first()
        if not user:
            return jsonify({'success': False, 'message': '사용자를 찾을 수 없습니다.'}), 401
        
        # 이벤트 조회
        event = db.query(OrderEvent).filter(
            OrderEvent.id == event_id,
            OrderEvent.order_id == order_id
        ).first()
        
        if not event:
            return jsonify({'success': False, 'message': '이벤트를 찾을 수 없습니다.'}), 404
        
        # 권한 확인
        if user.role != 'ADMIN' and event.created_by_user_id != user_id:
            return jsonify({'success': False, 'message': '본인이 생성한 이벤트만 되돌릴 수 있습니다.'}), 403
        
        # 주문 조회
        order = db.query(Order).filter(Order.id == order_id).first()
        if not order:
            return jsonify({'success': False, 'message': '주문을 찾을 수 없습니다.'}), 404
        
        payload = event.payload or {}
        target = payload.get('target', '')
        before_value = payload.get('before')
        after_value = payload.get('after')
        
        if not target or before_value is None:
            return jsonify({'success': False, 'message': '되돌리기 정보가 불완전합니다.'}), 400
        
        # structured_data 업데이트
        import copy
        from sqlalchemy.orm.attributes import flag_modified
        
        sd = order.structured_data or {}
        
        # target 경로 파싱 (예: "workflow.stage" -> ["workflow", "stage"])
        keys = target.split('.')
        
        # 현재 값 확인
        current_obj = sd
        for k in keys[:-1]:
            if k not in current_obj or not isinstance(current_obj[k], dict):
                current_obj[k] = {}
            current_obj = current_obj[k]
        
        last_key = keys[-1]
        current_value = current_obj.get(last_key)

        is_drawing_assignee_target = (target == 'assignments.drawing_assignee_user_ids')

        # 현재 값이 예상값(after)과 같은지 비교 (도메인별 호환 비교)
        is_expected_state = False
        current_display = current_value
        expected_display = after_value

        if is_drawing_assignee_target:
            current_ids = current_value if isinstance(current_value, list) else []
            expected_ids = payload.get('after_ids') if isinstance(payload.get('after_ids'), list) else None

            if expected_ids is not None:
                current_norm = sorted([int(x) for x in current_ids if str(x).isdigit()])
                expected_norm = sorted([int(x) for x in expected_ids if str(x).isdigit()])
                is_expected_state = (current_norm == expected_norm)
                current_display = current_norm
                expected_display = expected_norm
            else:
                # 과거 로그 호환: after가 이름 문자열("김도면")로 저장된 경우
                expected_names = []
                if isinstance(after_value, str):
                    expected_names = [x.strip() for x in after_value.split(',') if x.strip() and x.strip().lower() != 'none']
                if expected_names:
                    users_now = db.query(User).filter(User.id.in_(current_ids)).all() if current_ids else []
                    current_names = [u.name for u in users_now if u.name]
                    is_expected_state = (sorted(current_names) == sorted(expected_names))
                    current_display = ', '.join(current_names) if current_names else 'None'
                    expected_display = ', '.join(expected_names) if expected_names else 'None'
                else:
                    is_expected_state = (str(current_value) == str(after_value))
        else:
            is_expected_state = (str(current_value) == str(after_value))

        if not is_expected_state:
            return jsonify({
                'success': False,
                'message': f'현재 값({current_display})이 예상 값({expected_display})과 다릅니다. 이미 다른 변경이 발생했을 수 있습니다.'
            }), 409

        # 되돌릴 값 계산 (도메인별)
        revert_to_value = before_value
        if is_drawing_assignee_target:
            before_ids = payload.get('before_ids') if isinstance(payload.get('before_ids'), list) else None

            if before_ids is None:
                if isinstance(before_value, list):
                    before_ids = before_value
                elif isinstance(before_value, str):
                    names = [x.strip() for x in before_value.split(',') if x.strip() and x.strip().lower() != 'none']
                    if names:
                        users_prev = db.query(User).filter(User.name.in_(names), User.is_active == True).all()
                        before_ids = [u.id for u in users_prev]
                    else:
                        before_ids = []
                elif before_value in (None, 'None', ''):
                    before_ids = []

            if before_ids is not None:
                revert_to_value = before_ids

        # 되돌리기 실행
        current_obj[last_key] = revert_to_value

        # 도면 담당자 필드는 연관 키를 함께 동기화
        if is_drawing_assignee_target:
            ids_for_sync = revert_to_value if isinstance(revert_to_value, list) else []
            restored_users = db.query(User).filter(User.id.in_(ids_for_sync), User.is_active == True).all() if ids_for_sync else []

            sd['drawing_assignees'] = [
                {'id': u.id, 'name': u.name, 'team': u.team}
                for u in restored_users
            ]
            shipment = sd.get('shipment') or {}
            shipment['drawing_managers'] = [u.name for u in restored_users if u.name]
            sd['shipment'] = shipment
        
        order.structured_data = copy.deepcopy(sd)
        flag_modified(order, "structured_data")
        
        # Revert 이벤트 생성
        revert_payload = {
            'domain': payload.get('domain', 'UNKNOWN'),
            'action': 'REVERTED',
            'target': target,
            'before': after_value,
            'after': before_value,
            'reverted_value': revert_to_value,
            'change_method': 'API_REVERT',
            'source_screen': 'change_log_viewer',
            'reason': revert_reason,
            'reverted_event_id': event_id,
            'original_event_type': event.event_type,
        }
        
        revert_event = OrderEvent(
            order_id=order_id,
            event_type='CHANGE_REVERTED',
            payload=revert_payload,
            created_by_user_id=user_id
        )
        db.add(revert_event)
        
        # 로그
        from models import SecurityLog
        db.add(SecurityLog(
            user_id=user_id,
            message=f"주문 #{order_id} 변경 되돌리기: {target} ({after_value} -> {before_value})"
        ))
        
        db.commit()
        
        return jsonify({
            'success': True,
            'message': '변경이 성공적으로 되돌려졌습니다.',
            'reverted_target': target,
            'new_value': before_value
        })
        
    except Exception as e:
        db = get_db()
        db.rollback()
        import traceback
        print(f"변경 되돌리기 오류: {e}")
        print(traceback.format_exc())
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/orders/<int:order_id>/tasks', methods=['GET'])
@login_required
def api_order_tasks_list(order_id):
    """주문 팔로업(Task) 목록"""
    try:
        db = get_db()
        rows = db.query(OrderTask).filter(OrderTask.order_id == order_id).order_by(OrderTask.updated_at.desc()).all()
        tasks = []
        for t in rows:
            tasks.append({
                'id': t.id,
                'order_id': t.order_id,
                'title': t.title,
                'status': t.status,
                'owner_team': t.owner_team,
                'owner_user_id': t.owner_user_id,
                'due_date': t.due_date,
                'meta': t.meta,
                'created_at': t.created_at.strftime('%Y-%m-%d %H:%M:%S') if t.created_at else None,
                'updated_at': t.updated_at.strftime('%Y-%m-%d %H:%M:%S') if t.updated_at else None,
            })
        return jsonify({'success': True, 'tasks': tasks})
    except Exception as e:
        import traceback
        print(f"주문 Task 목록 오류: {e}")
        print(traceback.format_exc())
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/orders/<int:order_id>/tasks', methods=['POST'])
@login_required
@role_required(['ADMIN', 'MANAGER', 'STAFF'])
def api_order_tasks_create(order_id):
    """주문 팔로업(Task) 생성"""
    try:
        db = get_db()
        payload = request.get_json(silent=True) or {}
        title = (payload.get('title') or '').strip()
        if not title:
            return jsonify({'success': False, 'message': 'title이 필요합니다.'}), 400

        task = OrderTask(
            order_id=order_id,
            title=title,
            status=(payload.get('status') or 'OPEN'),
            owner_team=(payload.get('owner_team') or None),
            owner_user_id=(payload.get('owner_user_id') or None),
            due_date=(payload.get('due_date') or None),
            meta=(payload.get('meta') if isinstance(payload.get('meta'), dict) else None),
            created_at=datetime.datetime.now(),
            updated_at=datetime.datetime.now(),
        )
        db.add(task)
        db.commit()
        db.refresh(task)
        return jsonify({'success': True, 'task_id': task.id})
    except Exception as e:
        db = get_db()
        try:
            db.rollback()
        except Exception:
            pass
        import traceback
        print(f"주문 Task 생성 오류: {e}")
        print(traceback.format_exc())
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/orders/<int:order_id>/tasks/<int:task_id>', methods=['PUT'])
@login_required
@role_required(['ADMIN', 'MANAGER', 'STAFF'])
def api_order_tasks_update(order_id, task_id):
    """주문 팔로업(Task) 수정"""
    try:
        db = get_db()
        task = db.query(OrderTask).filter(OrderTask.id == task_id, OrderTask.order_id == order_id).first()
        if not task:
            return jsonify({'success': False, 'message': 'Task를 찾을 수 없습니다.'}), 404

        payload = request.get_json(silent=True) or {}
        if 'title' in payload:
            task.title = (payload.get('title') or '').strip()
        if 'status' in payload:
            task.status = payload.get('status') or task.status
        if 'owner_team' in payload:
            task.owner_team = payload.get('owner_team') or None
        if 'owner_user_id' in payload:
            task.owner_user_id = payload.get('owner_user_id') or None
        if 'due_date' in payload:
            task.due_date = payload.get('due_date') or None
        if 'meta' in payload and isinstance(payload.get('meta'), dict):
            task.meta = payload.get('meta')

        task.updated_at = datetime.datetime.now()
        db.commit()
        return jsonify({'success': True})
    except Exception as e:
        db = get_db()
        try:
            db.rollback()
        except Exception:
            pass
        import traceback
        print(f"주문 Task 수정 오류: {e}")
        print(traceback.format_exc())
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/orders/<int:order_id>/tasks/<int:task_id>', methods=['DELETE'])
@login_required
@role_required(['ADMIN', 'MANAGER', 'STAFF'])
def api_order_tasks_delete(order_id, task_id):
    """주문 팔로업(Task) 삭제"""
    try:
        db = get_db()
        task = db.query(OrderTask).filter(OrderTask.id == task_id, OrderTask.order_id == order_id).first()
        if not task:
            return jsonify({'success': False, 'message': 'Task를 찾을 수 없습니다.'}), 404
        db.delete(task)
        db.commit()
        return jsonify({'success': True})
    except Exception as e:
        db = get_db()
        try:
            db.rollback()
        except Exception:
            pass
        import traceback
        print(f"주문 Task 삭제 오류: {e}")
        print(traceback.format_exc())
        return jsonify({'success': False, 'message': str(e)}), 500


# -----------------------------
# Quest API (단계별 명확한 퀘스트 시스템)
# -----------------------------

@app.route('/api/orders/<int:order_id>/quest', methods=['GET'])
@login_required
def api_order_quest_get(order_id):
    """현재 단계의 Quest 조회"""
    try:
        db = get_db()
        order = db.query(Order).filter(Order.id == order_id).first()
        if not order:
            return jsonify({'success': False, 'message': '주문을 찾을 수 없습니다.'}), 404
        
        sd = order.structured_data or {}
        current_stage_code = get_stage(sd)  # 영문 코드 (예: 'RECEIVED')
        
        if not current_stage_code:
            return jsonify({'success': True, 'quest': None, 'stage': None})

        # 도면 단계는 퀘스트 승인 흐름을 사용하지 않음 (도면 전달/수령 확정 흐름으로 관리)
        if current_stage_code == 'DRAWING':
            return jsonify({
                'success': True,
                'quest': None,
                'stage': current_stage_code,
                'stage_label': STAGE_LABELS.get(current_stage_code, current_stage_code),
                'message': '도면 단계 퀘스트는 비활성화되었습니다.'
            })
        
        # 영문 코드를 한글 단계명으로 변환 (quest의 stage는 한글 단계명으로 저장될 수 있음)
        CODE_TO_STAGE_NAME = {v: k for k, v in STAGE_NAME_TO_CODE.items()}
        current_stage_name = CODE_TO_STAGE_NAME.get(current_stage_code, current_stage_code)
        
        # 현재 단계의 quest 찾기 (한글 단계명 또는 영문 코드 모두 확인)
        quests = sd.get("quests") or []
        current_quest = None
        for q in quests:
            if isinstance(q, dict):
                quest_stage = q.get("stage")
                # 한글 단계명 또는 영문 코드 모두 확인
                if quest_stage == current_stage_name or quest_stage == current_stage_code:
                    current_quest = q
                    break
        
        # quest가 없으면 템플릿에서 생성하고 DB에 저장 (한글 단계명으로 생성)
        if not current_quest:
            quest_tpl = get_quest_template_for_stage(current_stage_code)
            if quest_tpl:
                owner_person = session.get('username') or ''
                # 영문 코드로 quest 생성 (일관성 유지)
                current_quest = create_quest_from_template(current_stage_code, owner_person, sd)
                if current_quest:
                    # DB에 저장
                    if not sd.get("quests"):
                        sd["quests"] = []
                    sd["quests"].append(current_quest)
                    order.structured_data = sd
                    order.updated_at = datetime.datetime.now()
                    db.commit()
        
        return jsonify({
            'success': True,
            'quest': current_quest,
            'stage': current_stage_code,
            'stage_label': STAGE_LABELS.get(current_stage_code, current_stage_code),
        })
    except Exception as e:
        import traceback
        print(f"Quest 조회 오류: {e}")
        print(traceback.format_exc())
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/orders/<int:order_id>/quest', methods=['POST'])
@login_required
@role_required(['ADMIN', 'MANAGER', 'STAFF'])
def api_order_quest_create(order_id):
    """Quest 생성 (현재 단계 기준)"""
    try:
        db = get_db()
        order = db.query(Order).filter(Order.id == order_id).first()
        if not order:
            return jsonify({'success': False, 'message': '주문을 찾을 수 없습니다.'}), 404
        
        payload = request.get_json(silent=True) or {}
        stage = payload.get('stage') or get_stage(order.structured_data or {})
        
        if not stage:
            return jsonify({'success': False, 'message': '단계가 지정되지 않았습니다.'}), 400

        # 도면 단계는 퀘스트 생성 비활성화
        stage_code = STAGE_NAME_TO_CODE.get(stage, stage)
        if stage_code == 'DRAWING':
            return jsonify({'success': False, 'message': '도면 단계 퀘스트는 비활성화되었습니다.'}), 400
        
        # 이미 해당 단계의 quest가 있는지 확인
        sd = order.structured_data or {}
        if not sd.get("quests"):
            sd["quests"] = []
        
        existing = None
        for q in sd["quests"]:
            if isinstance(q, dict) and q.get("stage") == stage:
                existing = q
                break
        
        if existing:
            return jsonify({'success': False, 'message': '이미 해당 단계의 Quest가 존재합니다.'}), 400
        
        # Quest 생성
        owner_person = payload.get('owner_person') or session.get('username') or ''
        new_quest = create_quest_from_template(stage, owner_person, sd)
        
        if not new_quest:
            return jsonify({'success': False, 'message': 'Quest 템플릿을 찾을 수 없습니다.'}), 400
        
        sd["quests"].append(new_quest)
        order.structured_data = sd
        order.updated_at = datetime.datetime.now()
        db.commit()
        
        return jsonify({'success': True, 'quest': new_quest})
    except Exception as e:
        db = get_db()
        try:
            db.rollback()
        except Exception:
            pass
        import traceback
        print(f"Quest 생성 오류: {e}")
        print(traceback.format_exc())
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/orders/<int:order_id>/quest/approve', methods=['POST'])
@login_required
@role_required(['ADMIN', 'MANAGER', 'STAFF'])
def api_order_quest_approve(order_id):
    """팀별/담당자 Quest 승인 및 자동 단계 전환"""
    try:
        db = get_db()
        order = db.query(Order).filter(Order.id == order_id).first()
        if not order:
            return jsonify({'success': False, 'message': '주문을 찾을 수 없습니다.'}), 404
        
        payload = request.get_json(silent=True) or {}
        team = (payload.get('team') or '').strip()
        emergency_override = payload.get('emergency_override', False)
        override_reason = payload.get('override_reason', '').strip()
        
        user_id = session.get('user_id')
        user = db.query(User).filter(User.id == user_id).first()
        if not user:
            return jsonify({'success': False, 'message': '사용자를 찾을 수 없습니다.'}), 401
        
        # ERP Beta 주문 기본 승인 권한:
        # - 기본: can_edit_erp(user)
        # - 예외: assignee 단계(실측/고객컨펌/도면)는 지정 담당자도 승인 가능
        if getattr(order, 'is_erp_beta', False) and not can_edit_erp(user):
            sd_tmp = order.structured_data or {}
            stage_tmp = get_stage(sd_tmp)
            domain_tmp = None
            if stage_tmp in ('MEASURE', 'CONFIRM'):
                domain_tmp = 'SALES_DOMAIN'
            elif stage_tmp == 'DRAWING':
                domain_tmp = 'DRAWING_DOMAIN'

            can_assignee_override = False
            if domain_tmp:
                can_assignee_override = can_modify_domain(user, order, domain_tmp, emergency_override, override_reason)

                # SALES_DOMAIN 호환 fallback: assignee 미지정 + 주문 담당자명 일치 시 허용
                if (not can_assignee_override) and domain_tmp == 'SALES_DOMAIN':
                    allowed_ids = get_assignee_ids(order, domain_tmp)
                    if not allowed_ids:
                        manager_names = set()
                        parties_tmp = (sd_tmp.get('parties') or {}) if isinstance(sd_tmp, dict) else {}
                        manager_name_sd = ((parties_tmp.get('manager') or {}).get('name') or '').strip()
                        if manager_name_sd:
                            manager_names.add(manager_name_sd.lower())
                        manager_name_col = (order.manager_name or '').strip()
                        if manager_name_col:
                            manager_names.add(manager_name_col.lower())
                        user_name = (user.name or '').strip().lower()
                        user_username = (user.username or '').strip().lower()
                        if user_name in manager_names or user_username in manager_names:
                            can_assignee_override = True

            if not can_assignee_override:
                return jsonify({
                    'success': False,
                    'message': 'ERP Beta 수정 권한이 없습니다. (관리자, 라홈팀, 하우드팀, 영업팀 또는 지정 담당자만 가능)'
                }), 403
        
        sd = order.structured_data or {}
        current_stage_code = get_stage(sd)  # 영문 코드 (예: 'RECEIVED')
        
        if not current_stage_code:
            return jsonify({'success': False, 'message': '현재 단계가 없습니다.'}), 400

        # 도면 단계는 퀘스트 승인 자체를 허용하지 않음
        if current_stage_code == 'DRAWING':
            return jsonify({
                'success': False,
                'message': '도면 단계 퀘스트 승인은 비활성화되었습니다. 도면 전달/수령 확정으로 진행해주세요.'
            }), 400
        
        # 영문 코드를 한글 단계명으로 변환 (quest의 stage는 한글 단계명으로 저장됨)
        CODE_TO_STAGE_NAME = {v: k for k, v in STAGE_NAME_TO_CODE.items()}
        current_stage_name = CODE_TO_STAGE_NAME.get(current_stage_code, current_stage_code)
        
        # 현재 단계의 quest 찾기 (한글 단계명으로 저장된 quest와 비교)
        quests = sd.get("quests") or []
        current_quest = None
        quest_index = -1
        for i, q in enumerate(quests):
            if isinstance(q, dict):
                # quest의 stage는 한글 단계명 또는 영문 코드일 수 있으므로 둘 다 확인
                quest_stage = q.get("stage")
                if quest_stage == current_stage_name or quest_stage == current_stage_code:
                    current_quest = q
                    quest_index = i
                    break
        
        if not current_quest:
            # Quest가 없으면 생성 (한글 단계명으로 생성)
            owner_person = session.get('username') or ''
            current_quest = create_quest_from_template(current_stage_name, owner_person, sd)
            if not current_quest:
                return jsonify({'success': False, 'message': 'Quest 템플릿을 찾을 수 없습니다.'}), 400
            if not sd.get("quests"):
                sd["quests"] = []
            sd["quests"].append(current_quest)
            quest_index = len(sd["quests"]) - 1
        
        username = session.get('username') or ''
        now = datetime.datetime.now()
        
        # 승인 모드 확인
        approval_mode = current_quest.get("approval_mode", "team")
        
        if approval_mode == "assignee":
            # 담당자 기반 승인 (실측/도면/고객컨펌)
            
            # 권한 검사 (영업/도면 단계는 엄격 담당제)
            domain = None
            if current_stage_code in ('MEASURE', 'CONFIRM'):
                domain = 'SALES_DOMAIN'
            elif current_stage_code == 'DRAWING':
                domain = 'DRAWING_DOMAIN'
            
            if domain:
                can_modify = can_modify_domain(user, order, domain, emergency_override, override_reason)

                # 실측/고객컨펌(SALES_DOMAIN) 호환 fallback:
                # assignments.sales_assignee_user_ids가 비어있고 주문 담당자명이 현재 사용자와 일치하면 승인 허용.
                if (not can_modify) and domain == 'SALES_DOMAIN':
                    allowed_ids = get_assignee_ids(order, domain)
                    if not allowed_ids:
                        manager_names = set()
                        parties = (sd.get('parties') or {}) if isinstance(sd, dict) else {}
                        manager_name_sd = ((parties.get('manager') or {}).get('name') or '').strip()
                        if manager_name_sd:
                            manager_names.add(manager_name_sd.lower())

                        manager_name_col = (order.manager_name or '').strip()
                        if manager_name_col:
                            manager_names.add(manager_name_col.lower())

                        owner_person = (current_quest.get('owner_person') or '').strip()
                        if owner_person:
                            manager_names.add(owner_person.lower())

                        user_name = (user.name or '').strip().lower()
                        user_username = (user.username or '').strip().lower()
                        if user_name in manager_names or user_username in manager_names:
                            can_modify = True

                if not can_modify:
                    msg = f'{current_stage_name} 단계는 지정 담당자만 승인할 수 있습니다.'
                    if user.role == 'MANAGER':
                        msg += ' (긴급 오버라이드가 필요합니다.)'
                    return jsonify({'success': False, 'message': msg}), 403
            
            # 담당자 승인 처리
            if "assignee_approval" not in current_quest:
                current_quest["assignee_approval"] = {}
            
            current_quest["assignee_approval"] = {
                "approved": True,
                "approved_by": user_id,
                "approved_by_name": username,
                "approved_at": now.isoformat(),
            }
            
            current_quest["updated_at"] = now.isoformat()
            if current_quest.get("status") == "OPEN":
                current_quest["status"] = "IN_PROGRESS"
            
            # 담당자 기반은 승인 즉시 완료
            is_complete = True
            missing_teams = []
            
            # OrderEvent: 담당자 승인
            quest_event_payload = {
                'domain': domain or f'{current_stage_code}_DOMAIN',
                'action': 'QUEST_ASSIGNEE_APPROVED',
                'target': f'quest.assignee_approval',
                'before': 'not_approved',
                'after': 'approved',
                'change_method': 'API',
                'source_screen': 'erp_dashboard',
                'reason': f'{current_stage_name} 담당자 승인 완료',
                'is_override': emergency_override,
                'override_reason': override_reason if emergency_override else None,
            }
            quest_approval_event = OrderEvent(
                order_id=order.id,
                event_type='QUEST_APPROVAL_CHANGED',
                payload=quest_event_payload,
                created_by_user_id=user_id
            )
            db.add(quest_approval_event)
            
        else:
            # 팀 기반 승인 (기존 로직)
            if not team:
                return jsonify({'success': False, 'message': '팀이 지정되지 않았습니다.'}), 400
            
            # 팀 승인 처리
            if not current_quest.get("team_approvals"):
                current_quest["team_approvals"] = {}
            
            current_quest["team_approvals"][team] = {
                "approved": True,
                "approved_by": user_id,
                "approved_by_name": username,
                "approved_at": now.isoformat(),
            }
            
            current_quest["updated_at"] = now.isoformat()
            if current_quest.get("status") == "OPEN":
                current_quest["status"] = "IN_PROGRESS"
            
            # 모든 필수 팀 승인 완료 확인 (한글 단계명 사용)
            is_complete, missing_teams = check_quest_approvals_complete(sd, current_stage_name)
            
            # OrderEvent: Quest 승인
            quest_event_payload = {
                'domain': f'{current_stage_code}_DOMAIN',
                'action': 'QUEST_APPROVAL_CHANGED',
                'target': f'quest.team_approvals.{team}',
                'before': 'not_approved',
                'after': 'approved',
                'change_method': 'API',
                'source_screen': 'erp_dashboard',
                'reason': f'{team} 팀 승인 완료',
                'is_override': emergency_override,
                'override_reason': override_reason if emergency_override else None,
            }
            quest_approval_event = OrderEvent(
                order_id=order.id,
                event_type='QUEST_APPROVAL_CHANGED',
                payload=quest_event_payload,
                created_by_user_id=user_id
            )
            db.add(quest_approval_event)
        
        # quests 배열 업데이트
        sd["quests"][quest_index] = current_quest
        
        auto_transitioned = False
        if is_complete:
            # Quest 완료 처리
            current_quest["status"] = "COMPLETED"
            current_quest["completed_at"] = now.isoformat()
            sd["quests"][quest_index] = current_quest
            
            # [수정] 고객컨펌(CONFIRM) 단계는 자동 전환하지 않음
            # 생산팀이 "제작 시작" 버튼을 클릭할 때만 PRODUCTION으로 이동
            if current_stage_code != 'CONFIRM':
                # 다음 단계로 자동 전환
                # get_next_stage_for_completed_quest는 영문 코드를 반환함 (템플릿의 next_stage가 영문 코드)
                next_stage_code = get_next_stage_for_completed_quest(current_stage_name)
                if next_stage_code:
                    # 영문 코드를 한글 단계명으로 변환 (quest 생성 시 사용)
                    CODE_TO_STAGE_NAME = {v: k for k, v in STAGE_NAME_TO_CODE.items()}
                    next_stage_name = CODE_TO_STAGE_NAME.get(next_stage_code, next_stage_code)
                    
                    workflow = sd.get("workflow") or {}
                    old_stage = workflow.get("stage")
                    workflow["stage"] = next_stage_code  # 영문 코드로 저장
                    workflow["stage_updated_at"] = now.isoformat()
                    sd["workflow"] = workflow
                    
                    # 다음 단계의 Quest 자동 생성 (한글 단계명으로 생성)
                    next_quest = create_quest_from_template(next_stage_name, username, sd)
                    if next_quest:
                        if not sd.get("quests"):
                            sd["quests"] = []
                        sd["quests"].append(next_quest)
                    
                    # 이벤트 기록
                    ev = OrderEvent(
                        order_id=order.id,
                        event_type='STAGE_AUTO_TRANSITIONED',
                        payload={
                            'from': old_stage,
                            'to': next_stage_code,
                            'reason': 'quest_approvals_complete',
                            'approved_teams': get_required_approval_teams_for_stage(current_stage_name),
                        },
                        created_by_user_id=user_id
                    )
                    db.add(ev)
                    auto_transitioned = True
        
        # 최종 structured_data 저장 및 변경 감지
        order.structured_data = sd
        flag_modified(order, "structured_data")  # JSONB 필드 변경 명시적 표시
        order.updated_at = now
        db.commit()
        
        # 응답용 next_stage는 한글 단계명으로 변환
        next_stage_for_response = None
        if is_complete:
            next_stage_code = get_next_stage_for_completed_quest(current_stage_name)
            if next_stage_code:
                CODE_TO_STAGE_NAME = {v: k for k, v in STAGE_NAME_TO_CODE.items()}
                next_stage_for_response = CODE_TO_STAGE_NAME.get(next_stage_code, next_stage_code)
        
        return jsonify({
            'success': True,
            'quest': current_quest,
            'all_approved': is_complete,
            'missing_teams': missing_teams,
            'auto_transitioned': auto_transitioned,
            'next_stage': next_stage_for_response,
        })
    except Exception as e:
        db = get_db()
        try:
            db.rollback()
        except Exception:
            pass
        import traceback
        print(f"Quest 승인 오류: {e}")
        print(traceback.format_exc())
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/orders/<int:order_id>/quest/status', methods=['PUT'])
@login_required
@role_required(['ADMIN', 'MANAGER', 'STAFF'])
def api_order_quest_update_status(order_id):
    """Quest 상태 수동 업데이트 (OPEN, IN_PROGRESS, COMPLETED)"""
    try:
        db = get_db()
        order = db.query(Order).filter(Order.id == order_id).first()
        if not order:
            return jsonify({'success': False, 'message': '주문을 찾을 수 없습니다.'}), 404
        
        payload = request.get_json(silent=True) or {}
        status = payload.get('status')
        owner_person = payload.get('owner_person')
        
        if status not in ['OPEN', 'IN_PROGRESS', 'COMPLETED']:
            return jsonify({'success': False, 'message': '유효하지 않은 상태입니다.'}), 400
        
        sd = order.structured_data or {}
        current_stage = get_stage(sd)
        
        if not current_stage:
            return jsonify({'success': False, 'message': '현재 단계가 없습니다.'}), 400
        
        # 현재 단계의 quest 찾기
        quests = sd.get("quests") or []
        quest_index = -1
        for i, q in enumerate(quests):
            if isinstance(q, dict) and q.get("stage") == current_stage:
                quest_index = i
                break
        
        if quest_index == -1:
            return jsonify({'success': False, 'message': 'Quest를 찾을 수 없습니다.'}), 404
        
        now = datetime.datetime.now()
        quests[quest_index]["status"] = status
        quests[quest_index]["updated_at"] = now.isoformat()
        
        if owner_person:
            quests[quest_index]["owner_person"] = owner_person
        
        if status == "COMPLETED" and not quests[quest_index].get("completed_at"):
            quests[quest_index]["completed_at"] = now.isoformat()
        
        sd["quests"] = quests
        order.structured_data = sd
        order.updated_at = now
        db.commit()
        
        return jsonify({'success': True, 'quest': quests[quest_index]})
    except Exception as e:
        db = get_db()
        try:
            db.rollback()
        except Exception:
            pass
        import traceback
        print(f"Quest 상태 업데이트 오류: {e}")
        print(traceback.format_exc())
        return jsonify({'success': False, 'message': str(e)}), 500

# 지방 주문 자동 필터링 함수 제거 - 사용자가 직접 선택하도록 변경

# Auth routes moved to apps/auth.py

# Removed email verification and password reset routes

# --- format_options_for_display 함수를 여기로 이동 (또는 적절한 전역 위치) ---
def format_options_for_display(options_json_str):
    if not options_json_str:
        return ""
    try:
        options_data = json.loads(options_json_str)
        key_to_korean = {
            'product_name': '제품명', 'standard': '규격', 'internal': '내부',
            'color': '색상', 'option_detail': '상세옵션', 'handle': '손잡이',
            'misc': '기타', 'quote': '견적내용'
        }
        korean_to_key = {v: k for k, v in key_to_korean.items()}

        if isinstance(options_data, dict):
            if options_data.get("option_type") == "direct" and "details" in options_data:
                details = options_data["details"]
                display_parts = []
                for key, kor_display_name in key_to_korean.items():
                    value = details.get(key)
                    if value:
                        display_parts.append(f"{kor_display_name}: {value}")
                return ", ".join(display_parts) if display_parts else "옵션 정보 없음"
            elif options_data.get("option_type") == "online" and "online_options_summary" in options_data:
                summary = options_data["online_options_summary"]
                # 줄바꿈 문자를 <br> 태그로 변경하여 반환
                return summary.replace('\n', '<br>') if summary else "온라인 옵션 요약 없음"
            elif any(key in options_data for key in key_to_korean.keys()):
                display_parts = []
                for key_eng, value in options_data.items():
                    if value and key_eng in key_to_korean:
                        display_parts.append(f"{key_to_korean[key_eng]}: {value}")
                return ", ".join(display_parts) if display_parts else "옵션 정보 없음 (구)"
            elif any(key_kor in options_data for key_kor in korean_to_key.keys()):
                display_parts = []
                for key_kor, value in options_data.items():
                    if value and key_kor in korean_to_key:
                        display_parts.append(f"{key_kor}: {value}")
                return ", ".join(display_parts) if display_parts else "옵션 정보 없음 (구-한글)"
            else:
                display_parts = []
                for key, value in options_data.items():
                    if isinstance(value, (str, int, float)):
                        display_parts.append(f"{key}: {value}")
                return ", ".join(display_parts) if display_parts else options_json_str
        else:
            return str(options_data)
    except json.JSONDecodeError:
        return options_json_str if options_json_str else "옵션 정보 없음"
    except Exception:
        return options_json_str if options_json_str else "옵션 처리 오류"
# --- format_options_for_display 함수 끝 ---

# Context Processors
@app.context_processor
def inject_status_list():
    """상태 목록과 현재 사용자 정보를 템플릿에 주입"""
    # 삭제됨(DELETED) 상태를 제외한 상태 목록
    display_status = {k: v for k, v in STATUS.items() if k != 'DELETED'}
    
    # 일괄 작업용 상태 목록 (삭제됨 제외)
    bulk_action_status = {k: v for k, v in STATUS.items() if k != 'DELETED'}
    
    # 현재 로그인한 사용자 추가
    current_user = None
    if 'user_id' in session:
        current_user = get_user_by_id(session['user_id'])

    # 관리자 전용: 드롭다운 아이디 이동용 사용자 목록 / 전환 중인지 여부
    admin_switch_users = []
    impersonating_from_id = session.get('impersonating_from')
    if current_user and current_user.role == 'ADMIN':
        db = get_db()
        admin_switch_users = db.query(User).filter(
            User.is_active == True,
            User.id != current_user.id
        ).order_by(User.name).all()

    # ERP Beta 플래그 (기본: 활성) - 필요 시 환경변수로 OFF 가능
    erp_beta_enabled = str(os.getenv('ERP_BETA_ENABLED', 'true')).lower() in ['1', 'true', 'yes', 'y', 'on']
    
    return dict(
        STATUS=display_status, 
        BULK_ACTION_STATUS=bulk_action_status,
        ALL_STATUS=STATUS, 
        ROLES=ROLES,
        current_user=current_user,
        admin_switch_users=admin_switch_users,
        impersonating_from_id=impersonating_from_id,
        erp_beta_enabled=erp_beta_enabled
    )

def parse_json_string(json_string):
    if not json_string:
        return None
    try:
        return json.loads(json_string)
    except json.JSONDecodeError:
        return None

def get_preserved_filter_args(request_args):
    """필터링 상태를 유지하기 위한 URL 매개변수를 반환합니다."""
    redirect_args = {}
    preserved_params = ['search', 'status', 'region', 'page', 'sort', 'direction', 'sort_by', 'sort_order'] + [k for k in request_args.keys() if k.startswith('filter_')]
    for key in preserved_params:
        if key in request_args:
            redirect_args[key] = request_args.get(key)
    return redirect_args

@app.context_processor
def utility_processor():
    return dict(parse_json_string=parse_json_string)



# Routes
@app.route('/favicon.ico')
def favicon():
    """favicon 요청 처리 (404 방지)"""
    return '', 204  # No Content

@app.route('/debug-db')
def debug_db():
    """Database Connection Check Route (Temporary)"""
    try:
        from sqlalchemy import text
        db = get_db()
        # 1. Simple Select
        result = db.execute(text("SELECT 1")).fetchone()
        
        # 2. Check Tables
        tables_query = text("SELECT table_name FROM information_schema.tables WHERE table_schema='public'")
        tables = [row[0] for row in db.execute(tables_query).fetchall()]
        
        status = "SUCCESS"
        message = f"Connected! Result: {result}, Tables: {tables}"
        
        # Check specific table
        if 'users' in tables:
            user_count = db.query(User).count()
            message += f" | Users count: {user_count}"
        else:
            status = "WARNING"
            message += " | 'users' table MISSING"
            
        return jsonify({"status": status, "message": message, "env_db_url_set": bool(os.environ.get('DATABASE_URL'))})
        
    except Exception as e:
        import traceback
        return jsonify({
            "status": "ERROR", 
            "error": str(e), 
            "traceback": traceback.format_exc()
        }), 500

@app.route('/admin/migration', methods=['GET', 'POST'])
@login_required
@role_required(['ADMIN', 'MANAGER'])
def admin_migration():
    """Web-based Data Migration (SQLite Upload -> Postgres)"""
    if request.method == 'POST':
        if 'db_file' not in request.files:
            flash('파일이 없습니다.', 'error')
            return redirect(request.url)
        
        file = request.files['db_file']
        if file.filename == '':
            flash('파일을 선택해주세요.', 'error')
            return redirect(request.url)
        
        if file:
            # 1. Save uploaded file temporarily
            filename = secure_filename(file.filename)
            temp_path = os.path.join(app.config['UPLOAD_FOLDER'], 'temp_migration.db')
            file.save(temp_path)
            
            # 2. Check reset option
            do_reset = (request.form.get('reset') == 'on')

            # 3. Run Migration
            from web_migration import run_web_migration
            db_session = get_db()
            
            success, logs = run_web_migration(temp_path, db_session, reset=do_reset)
            
            # 3. Cleanup
            if os.path.exists(temp_path):
                os.remove(temp_path)
                
            if success:
                flash(f'마이그레이션 완료! ({len(logs)} logs)', 'success')
            else:
                flash('마이그레이션 중 오류가 발생했습니다.', 'error')
                
            return render_template('admin/migration_result.html', logs=logs, success=success)

    return render_template('admin/migration_upload.html')

@app.route('/')
@login_required
def index():
    try:
        db = get_db()
        status_filter = request.args.get('status')
        manager_filter = (request.args.get('manager') or '').strip()
        search_query = request.args.get('search', '').strip()
        sort_column = request.args.get('sort', 'id')
        sort_direction = request.args.get('direction', 'desc')
        page = request.args.get('page', 1, type=int)
        per_page = 100

        status_filter = request.args.get('status')
        region_filter = request.args.get('region')
        sort_by = request.args.get('sort_by', 'id')
        sort_order = request.args.get('sort_order', 'desc')

        filterable_columns = [
            'id', 'received_date', 'received_time', 'customer_name', 'phone',
            'address', 'product', 'options', 'notes', 'status',
            'measurement_date', 'measurement_time', 'completion_date', 'manager_name', 'payment_amount'
        ]

        column_filters = {}
        for col in filterable_columns:
            filter_key = f'filter_{col}'
            if filter_key in request.args:
                column_filters[col] = request.args[filter_key]
        
        active_column_filters = {k: v for k, v in column_filters.items() if v}

        query = db.query(Order).filter(Order.status != 'DELETED')

        if status_filter:
            if status_filter == 'ALL':
                pass
            else:
                query = query.filter(Order.status == status_filter)
        
        if region_filter == 'metro':
            query = query.filter(Order.is_regional == False)
        elif region_filter == 'regional':
            query = query.filter(Order.is_regional == True)
        
        if search_query:
            search_term = f"%{search_query}%"
            query = query.filter(
                or_(
                    Order.id.cast(String).like(search_term),
                    Order.received_date.like(search_term),
                    Order.received_time.like(search_term),
                    Order.customer_name.like(search_term),
                    Order.phone.like(search_term),
                    Order.address.like(search_term),
                    Order.product.like(search_term),
                    Order.options.like(search_term),
                    Order.notes.like(search_term),
                    Order.status.like(search_term),
                    Order.measurement_date.like(search_term),
                    Order.measurement_time.like(search_term),
                    Order.scheduled_date.like(search_term),
                    Order.completion_date.like(search_term),
                    Order.manager_name.like(search_term)
                )
            )

        for column, filter_value in active_column_filters.items():
            if filter_value:
                filter_term = f"%{filter_value}%"
                if column == 'id':
                    query = query.filter(Order.id.cast(String).like(filter_term))
                elif column == 'payment_amount':
                    query = query.filter(Order.payment_amount.cast(String).like(filter_term))
                elif hasattr(Order, column):
                    column_attr = getattr(Order, column)
                    query = query.filter(column_attr.like(filter_term))
        
        # 항상 ID 역순(최신순)으로 정렬 - URL 파라미터 무시
        sort_column = 'id'
        sort_direction = 'desc'

        query = query.order_by(Order.id.desc())

        total_orders = query.count()
        orders_from_db = query.offset((page - 1) * per_page).limit(per_page).all()

        processed_orders = []
        for order_db_item in orders_from_db:
            order_display_data = copy.deepcopy(order_db_item)
            order_display_data.display_options = format_options_for_display(order_db_item.options)
            
            # ERP Beta 주문인 경우 structured_data에서 정보 추출하여 표시
            if order_db_item.is_erp_beta and order_db_item.structured_data:
                sd = _ensure_dict(order_db_item.structured_data)
                
                # 고객명: structured_data.parties.customer.name
                customer_name = ((sd.get('parties') or {}).get('customer') or {}).get('name')
                if customer_name:
                    order_display_data.customer_name = customer_name
                
                # 연락처: structured_data.parties.customer.phone
                phone = ((sd.get('parties') or {}).get('customer') or {}).get('phone')
                if phone:
                    order_display_data.phone = phone
                
                # 주소: structured_data.site.address_full 또는 address_main
                address = ((sd.get('site') or {}).get('address_full') or 
                          (sd.get('site') or {}).get('address_main'))
                if address:
                    order_display_data.address = address
                
                # 제품: structured_data.items에서 첫 번째 제품명 또는 items의 요약
                items = sd.get('items') or []
                if items and len(items) > 0:
                    # 첫 번째 제품의 product_name 또는 name 사용
                    first_item = items[0]
                    product_name = first_item.get('product_name') or first_item.get('name')
                    if product_name:
                        # 여러 제품이 있으면 "제품1 외 N개" 형식으로 표시
                        if len(items) > 1:
                            order_display_data.product = f"{product_name} 외 {len(items) - 1}개"
                        else:
                            order_display_data.product = product_name
                
                # 실측일: structured_data.schedule.measurement.date
                measurement_date = (((sd.get('schedule') or {}).get('measurement') or {}).get('date'))
                if measurement_date:
                    order_display_data.measurement_date = measurement_date
                
                # 실측시간: structured_data.schedule.measurement.time
                # "종일", "오전", "오후" 또는 실제 시간(예: "14:00") 형식 지원
                measurement_time = (((sd.get('schedule') or {}).get('measurement') or {}).get('time'))
                if measurement_time:
                    # 시간 값이 있으면 그대로 사용 (종일, 오전, 오후, 또는 실제 시간)
                    order_display_data.measurement_time = measurement_time
                
                # 시공일: structured_data.schedule.construction.date
                construction_date = (((sd.get('schedule') or {}).get('construction') or {}).get('date'))
                if construction_date:
                    # scheduled_date 필드에 저장 (Order 모델에 construction_date가 없으므로)
                    order_display_data.scheduled_date = construction_date
                
                # 담당자: structured_data.parties.manager.name
                manager_name = ((sd.get('parties') or {}).get('manager') or {}).get('name')
                if manager_name:
                    order_display_data.manager_name = manager_name
            
            processed_orders.append(order_display_data)

        user = None
        if 'user_id' in session:
            user = get_user_by_id(session['user_id'])
        
        return render_template(
            'index.html',
            orders=processed_orders,
            status_list=STATUS,
            STATUS=STATUS,
            current_status=status_filter,
            search_query=search_query,
            sort_column=sort_column,
            sort_direction=sort_direction,
            page=page,
            per_page=per_page,
            total_orders=total_orders,
            active_column_filters=column_filters,
            user=user,
            current_region=region_filter
        )
    except UnicodeDecodeError as e:
        print(f"Index 페이지 로딩 중 인코딩 오류: {str(e)}")
        flash('데이터베이스 연결 중 인코딩 문제가 발생했습니다. 관리자에게 문의하세요.', 'error')
        # 빈 데이터로 페이지 렌더링 시도
        return render_template(
            'index.html',
            orders=[], 
            status_list=STATUS,
            STATUS=STATUS,
            current_status=None,
            search_query='',
            sort_column='id',
            sort_direction='desc',
            page=1,
            per_page=100,
            total_orders=0,
            active_column_filters={},
            user=None,
            current_region=None
        )
    except Exception as e:
        print(f"Index 페이지 로딩 중 오류: {str(e)}")
        flash('페이지 로딩 중 오류가 발생했습니다.', 'error')
        return redirect(url_for('auth.login'))

@app.route('/add', methods=['GET', 'POST'])
@login_required
@role_required(['ADMIN', 'MANAGER', 'STAFF'])
def add_order():
    if request.method == 'POST':
        try:
            db = get_db()

            create_mode = (request.form.get('create_mode') or 'LEGACY').upper().strip()
            
            if create_mode == 'ERP_BETA':
                # ERP Beta 생성: raw/structured 중심. Order의 NOT NULL 컬럼은 fallback으로 채움.
                raw_text = (request.form.get('raw_order_text') or '').strip()
                structured_json = (request.form.get('structured_data_json') or '').strip()
                stage = (request.form.get('erp_stage') or 'RECEIVED').strip()
                owner_team = (request.form.get('erp_owner_team') or '').strip()
                urgent = bool(request.form.get('erp_urgent') == '1')
                urgent_reason = (request.form.get('erp_urgent_reason') or '').strip()
                meas_date = (request.form.get('erp_measurement_date') or '').strip()
                cons_date = (request.form.get('erp_construction_date') or '').strip()

                structured_data = {}
                if structured_json:
                    try:
                        parsed = json.loads(structured_json)
                        if isinstance(parsed, dict):
                            structured_data = parsed
                    except Exception:
                        structured_data = {}

                # 최소 구조 보정
                structured_data.setdefault('workflow', {})
                structured_data['workflow'].setdefault('stage', stage or 'RECEIVED')
                structured_data['workflow']['stage_updated_at'] = datetime.datetime.now().isoformat()

                structured_data.setdefault('assignments', {})
                if owner_team:
                    structured_data['assignments']['owner_team'] = owner_team

                structured_data.setdefault('flags', {})
                if urgent:
                    structured_data['flags']['urgent'] = True
                    if urgent_reason:
                        structured_data['flags']['urgent_reason'] = urgent_reason

                structured_data.setdefault('schedule', {})
                if meas_date:
                    structured_data['schedule'].setdefault('measurement', {})
                    structured_data['schedule']['measurement']['date'] = meas_date
                if cons_date:
                    structured_data['schedule'].setdefault('construction', {})
                    structured_data['schedule']['construction']['date'] = cons_date

                # Order NOT NULL fallback (가능하면 structured에서 뽑고, 없으면 placeholder)
                cust_name = ((structured_data.get('parties') or {}).get('customer') or {}).get('name') or (request.form.get('erp_customer_name') or '').strip() or 'ERP Beta'
                cust_phone = ((structured_data.get('parties') or {}).get('customer') or {}).get('phone') or (request.form.get('erp_customer_phone') or '').strip() or '000-0000-0000'
                addr = ((structured_data.get('site') or {}).get('address_full') or (structured_data.get('site') or {}).get('address_main')) or (request.form.get('erp_address') or '').strip() or '-'
                prod = (request.form.get('erp_product') or '').strip() or 'ERP Beta'

                new_order = Order(
                    received_date=request.form.get('received_date') or datetime.datetime.now().strftime('%Y-%m-%d'),
                    received_time=request.form.get('received_time') or datetime.datetime.now().strftime('%H:%M'),
                    customer_name=cust_name,
                    phone=cust_phone,
                    address=addr,
                    product=prod,
                    options=None,
                    notes=request.form.get('notes') or None,
                    status='RECEIVED',
                    is_erp_beta=True,
                    raw_order_text=raw_text,
                    structured_data=structured_data,
                    structured_schema_version=1,
                    structured_confidence=None,
                    structured_updated_at=datetime.datetime.now(),
                )

                db.add(new_order)
                db.flush()
                db.commit()
                flash('ERP Beta 주문이 성공적으로 추가되었습니다.', 'success')
                return redirect(url_for('index'))

            # ==========================
            # LEGACY(기존 주문) 생성 로직
            # ==========================
            # 필수 필드 검증
            required_fields = ['customer_name', 'phone', 'address', 'product']
            for field in required_fields:
                if not request.form.get(field):
                    flash(f'{field} 필드는 필수입니다.', 'error')
                    return redirect(url_for('add_order'))
            
            # 새 주문 생성
            options_data = None
            option_type = request.form.get('option_type')

            if option_type == 'direct':
                direct_options = {
                    'product_name': request.form.get('direct_product_name'),
                    'standard': request.form.get('direct_standard'),
                    'internal': request.form.get('direct_internal'),
                    'color': request.form.get('direct_color'),
                    'option_detail': request.form.get('direct_option_detail'),
                    'handle': request.form.get('direct_handle'),
                    'misc': request.form.get('direct_misc'),
                    'quote': request.form.get('direct_quote')
                }
                # 비어있지 않은 값들만 필터링하거나, 모든 값을 저장할 수 있습니다.
                # 여기서는 모든 값을 저장합니다.
                options_data = json.dumps(direct_options, ensure_ascii=False)
            else: # 'online' or an undefined type
                options_data = request.form.get('options_online')

            # payment_amount 추가
            payment_amount_str = request.form.get('payment_amount', '').replace(',', '') # 콤마 제거
            payment_amount = None
            if payment_amount_str:
                try:
                    payment_amount = int(payment_amount_str) # 정수로 변환
                except ValueError:
                    flash('결제금액은 숫자만 입력해주세요.', 'error')
                    return render_template('add_order.html')
            else:
                payment_amount = 0 # 값이 없으면 0으로 처리

            # 지방 주문 여부 설정
            is_regional_val = 'is_regional' in request.form
            
            # 자가실측 여부 설정
            is_self_measurement_val = 'is_self_measurement' in request.form
            # 수납장 여부 설정
            is_cabinet_val = 'is_cabinet' in request.form
            
            # 지방 주문일 경우, 체크리스트 항목들도 가져옴
            measurement_completed_val = False
            regional_sales_order_upload_val = False
            regional_blueprint_sent_val = False
            regional_order_upload_val = False
            construction_type_val = None

            if is_regional_val:
                measurement_completed_val = 'measurement_completed' in request.form
                regional_sales_order_upload_val = 'regional_sales_order_upload' in request.form
                regional_blueprint_sent_val = 'regional_blueprint_sent' in request.form
                regional_order_upload_val = 'regional_order_upload' in request.form
                construction_type_val = request.form.get('construction_type')

            new_order = Order(
                received_date=request.form.get('received_date'),
                received_time=request.form.get('received_time'),
                customer_name=request.form.get('customer_name'),
                phone=request.form.get('phone'),
                address=request.form.get('address'),
                product=request.form.get('product'),
                options=options_data,
                notes=request.form.get('notes'),
                status=request.form.get('status', 'RECEIVED'), # Use submitted status or default to RECEIVED
                # Add new fields from the form
                measurement_date=request.form.get('measurement_date'),
                measurement_time=request.form.get('measurement_time'),
                completion_date=request.form.get('completion_date'),
                manager_name=request.form.get('manager_name'),
                payment_amount=payment_amount, # 저장
                # 추가된 상태별 날짜 필드
                scheduled_date=request.form.get('scheduled_date'),
                as_received_date=request.form.get('as_received_date'),
                as_completed_date=request.form.get('as_completed_date'),
                is_regional=is_regional_val,
                is_self_measurement=is_self_measurement_val,
                is_cabinet=is_cabinet_val,
                cabinet_status='RECEIVED' if is_cabinet_val else None,
                measurement_completed=measurement_completed_val,
                regional_sales_order_upload=regional_sales_order_upload_val,
                regional_blueprint_sent=regional_blueprint_sent_val,
                regional_order_upload=regional_order_upload_val,
                construction_type=construction_type_val,
                is_erp_beta=False,
            )
            
            db.add(new_order)
            db.flush() # 새 주문의 ID를 가져오기 위해 flush
            order_id_for_log = new_order.id # ID 저장
            customer_name_for_log = new_order.customer_name
            user_for_log = get_user_by_id(session['user_id'])
            user_name_for_log = user_for_log.name if user_for_log else "Unknown user"
            db.commit() # 커밋은 flush 이후에
            
            log_access(f"주문 #{order_id_for_log} ({customer_name_for_log}) 추가 - 담당자: {user_name_for_log} (ID: {session.get('user_id')})", session.get('user_id'))
            
            flash('주문이 성공적으로 추가되었습니다.', 'success')
            return redirect(url_for('index'))
            
        except Exception as e:
            db.rollback()
            flash(f'오류가 발생했습니다: {str(e)}', 'error')
            return redirect(url_for('add_order'))
    
    today = datetime.datetime.now().strftime('%Y-%m-%d')
    current_time = datetime.datetime.now().strftime('%H:%M')
    
    return render_template('add_order.html', today=today, current_time=current_time)

@app.route('/edit/<int:order_id>', methods=['GET', 'POST'])
@login_required
@role_required(['ADMIN', 'MANAGER', 'STAFF'])
def edit_order(order_id):
    db = get_db()
    
    order = db.query(Order).filter(Order.id == order_id, Order.status != 'DELETED').first()
    
    if not order:
        flash('주문을 찾을 수 없거나 이미 삭제되었습니다.', 'error')
        return redirect(url_for('index'))
    
    # ERP Beta 주문의 경우 수정 권한 검사
    if order.is_erp_beta:
        user = get_user_by_id(session['user_id'])
        if not can_edit_erp(user):
            flash('ERP Beta 주문 수정 권한이 없습니다. (관리자, CS, 영업팀만 가능)', 'error')
            return redirect(url_for('index'))
    
    # 옵션 데이터 처리를 위한 변수 초기화
    option_type = 'online'  # 기본 옵션 타입
    online_options = ""     # 온라인 옵션 텍스트
    direct_options = {      # 직접 입력 옵션 필드
        'product_name': '', 
        'standard': '', 
        'internal': '',
        'color': '',
        'option_detail': '',
        'handle': '',
        'misc': '',
        'quote': ''
    }
    
    # 주문 옵션 데이터 처리
    if order.options:
        try:
            # 옵션 데이터 파싱 시도
            options_data = json.loads(order.options)
            
            # 옵션 데이터가 객체고 option_type 필드가 있는 경우
            if isinstance(options_data, dict):
                # 1. option_type 필드가 있는 경우 
                if 'option_type' in options_data:
                    option_type = options_data['option_type']
                    
                    if option_type == 'direct' and 'details' in options_data:
                        # 새로운 형식: "details" 객체에서 직접 값 추출
                        details = options_data['details']
                        for key in direct_options.keys():
                            if key in details:
                                direct_options[key] = details[key]
                    elif option_type == 'online' and 'online_options_summary' in options_data:
                        online_options = options_data['online_options_summary']
                
                # 2. 구형식 - option_type 없이 직접 키가 있는 경우
                elif any(key in options_data for key in direct_options.keys()):
                    option_type = 'direct'
                    for key in direct_options.keys():
                        if key in options_data:
                            direct_options[key] = options_data[key]
                
                # 3. 한글 키 대응
                elif any(key in options_data for key in ['제품명', '규격', '내부', '색상', '상세옵션', '손잡이', '기타', '견적내용']):
                    option_type = 'direct'
                    key_mapping = {
                        '제품명': 'product_name',
                        '규격': 'standard', 
                        '내부': 'internal',
                        '색상': 'color',
                        '상세옵션': 'option_detail',
                        '손잡이': 'handle',
                        '기타': 'misc',
                        '견적내용': 'quote'
                    }
                    for k_kor, k_eng in key_mapping.items():
                        if k_kor in options_data:
                            direct_options[k_eng] = options_data[k_kor]
                
                # 4. 이외의 경우 online으로 처리하고 문자열로 표시
                else:
                    option_type = 'online'
                    online_options = order.options  # 원래 문자열 그대로 표시
            
            # 객체가 아닌 경우 온라인 옵션으로 처리
            else:
                option_type = 'online'
                online_options = order.options
                
        except json.JSONDecodeError:
            # JSON 파싱 실패 시 온라인 옵션으로 처리
            option_type = 'online'
            online_options = order.options if order.options else ""
    
    if request.method == 'POST':
        try:
            # 폼 데이터 처리
            
            # 폼에서 넘어온 값들을 가져올 때, 해당 필드가 폼에 없으면 기존 order 객체의 값을 기본값으로 사용
            received_date = request.form.get('received_date', order.received_date)
            received_time = request.form.get('received_time', order.received_time)
            customer_name = request.form.get('customer_name', order.customer_name)
            phone = request.form.get('phone', order.phone)
            address = request.form.get('address', order.address)
            product = request.form.get('product', order.product)
            notes = request.form.get('notes', order.notes)
            status = request.form.get('status', order.status) # 상태가 없으면 기존 상태 유지
            
            measurement_date = request.form.get('measurement_date', order.measurement_date)
            measurement_time = request.form.get('measurement_time', order.measurement_time)
            completion_date = request.form.get('completion_date', order.completion_date)
            manager_name = request.form.get('manager_name', order.manager_name)

            # 새로 추가한 필드들 (기존 값 유지)
            scheduled_date = request.form.get('scheduled_date', order.scheduled_date)
            as_received_date = request.form.get('as_received_date', order.as_received_date)
            as_completed_date = request.form.get('as_completed_date', order.as_completed_date)
            shipping_scheduled_date = request.form.get('shipping_scheduled_date', order.shipping_scheduled_date) # 상차 예정일 추가

            # 옵션 데이터 처리 (폼에 option_type이 있을 때만 업데이트)
            options_data_json_to_save = order.options # 기본적으로 기존 옵션 유지
            if 'option_type' in request.form:
                current_option_type = request.form.get('option_type')
            
                if current_option_type == 'direct':
                    direct_details = {
                        'product_name': request.form.get('direct_product_name', ''),
                        'standard': request.form.get('direct_standard', ''),
                        'internal': request.form.get('direct_internal', ''),
                        'color': request.form.get('direct_color', ''),
                        'option_detail': request.form.get('direct_option_detail', ''),
                        'handle': request.form.get('direct_handle', ''),
                        'misc': request.form.get('direct_misc', ''),
                        'quote': request.form.get('direct_quote', '')
                    }
                    options_to_save_dict = {
                        "option_type": "direct",
                        "details": direct_details
                    }
                    options_data_json_to_save = json.dumps(options_to_save_dict, ensure_ascii=False)
                else:  # 'online'
                    online_summary = request.form.get('options_online', '')
                    options_to_save_dict = {
                        "option_type": "online",
                        "online_options_summary": online_summary
                    }
                    options_data_json_to_save = json.dumps(options_to_save_dict, ensure_ascii=False)
            
            # ... (기존 POST 로직의 변경 감지 및 DB 업데이트 부분) ...
            changes = {}
            if order.received_date != received_date: changes['received_date'] = {'old': order.received_date, 'new': received_date}
            if order.received_time != received_time: changes['received_time'] = {'old': order.received_time, 'new': received_time}
            if order.customer_name != customer_name: changes['customer_name'] = {'old': order.customer_name, 'new': customer_name}
            if order.phone != phone: changes['phone'] = {'old': order.phone, 'new': phone}
            if order.address != address: changes['address'] = {'old': order.address, 'new': address}
            if order.product != product: changes['product'] = {'old': order.product, 'new': product}
            if order.options != options_data_json_to_save: changes['options'] = {'old': order.options, 'new': options_data_json_to_save}
            if order.notes != notes: changes['notes'] = {'old': order.notes, 'new': notes}
            if order.status != status: changes['status'] = {'old': order.status, 'new': status}
            if order.measurement_date != measurement_date: changes['measurement_date'] = {'old': order.measurement_date, 'new': measurement_date}
            if order.measurement_time != measurement_time: changes['measurement_time'] = {'old': order.measurement_time, 'new': measurement_time}
            if order.completion_date != completion_date: changes['completion_date'] = {'old': order.completion_date, 'new': completion_date}
            if order.manager_name != manager_name: changes['manager_name'] = {'old': order.manager_name, 'new': manager_name}
            
            # 새 필드들 변경 감지
            if order.scheduled_date != scheduled_date: changes['scheduled_date'] = {'old': order.scheduled_date, 'new': scheduled_date}
            if order.as_received_date != as_received_date: changes['as_received_date'] = {'old': order.as_received_date, 'new': as_received_date}
            if order.as_completed_date != as_completed_date: changes['as_completed_date'] = {'old': order.as_completed_date, 'new': as_completed_date}
            if order.shipping_scheduled_date != shipping_scheduled_date: changes['shipping_scheduled_date'] = {'old': order.shipping_scheduled_date, 'new': shipping_scheduled_date} # 상차 예정일 변경 감지
            
            # 지방 주문 관련 필드 변경 감지
            is_regional_new = 'is_regional' in request.form
            if order.is_regional != is_regional_new: changes['is_regional'] = {'old': order.is_regional, 'new': is_regional_new}
            
            # 자가실측 관련 필드 변경 감지
            is_self_measurement_new = 'is_self_measurement' in request.form
            if order.is_self_measurement != is_self_measurement_new: changes['is_self_measurement'] = {'old': order.is_self_measurement, 'new': is_self_measurement_new}
            
            measurement_completed_new = 'measurement_completed' in request.form
            if order.measurement_completed != measurement_completed_new: changes['measurement_completed'] = {'old': order.measurement_completed, 'new': measurement_completed_new}
            
            construction_type_new = request.form.get('construction_type', order.construction_type)
            if order.construction_type != construction_type_new: changes['construction_type'] = {'old': order.construction_type, 'new': construction_type_new}
            
            # payment_amount 업데이트 및 변경 감지
            new_payment_amount = order.payment_amount # 기본적으로 기존 결제금액 유지
            if 'payment_amount' in request.form:
                payment_amount_str = request.form.get('payment_amount', '').replace(',', '') # 콤마 제거
                if payment_amount_str: # 빈 문자열이 아닌 경우에만 변환 시도
                    try:
                        new_payment_amount = int(payment_amount_str) 
                    except ValueError:
                        flash('결제금액은 숫자만 입력해주세요.', 'error')
                        raise ValueError("Invalid payment amount")
                else: # 빈 문자열로 넘어오면 0으로 처리 (또는 None으로 처리할 수도 있음)
                    new_payment_amount = 0
            
            if order.payment_amount != new_payment_amount:
                changes['payment_amount'] = {'old': order.payment_amount, 'new': new_payment_amount}
            # order.payment_amount = new_payment_amount # 아래에서 한꺼번에 업데이트

            # Update order object
            order.received_date = received_date
            order.received_time = received_time
            order.customer_name = customer_name
            order.phone = phone
            order.address = address
            order.product = product
            order.options = options_data_json_to_save
            order.notes = notes
            order.status = status
            order.measurement_date = measurement_date
            order.measurement_time = measurement_time
            order.completion_date = completion_date
            order.manager_name = manager_name
            
            # 새 필드 값 업데이트
            order.scheduled_date = scheduled_date
            order.as_received_date = as_received_date
            order.as_completed_date = as_completed_date
            order.shipping_scheduled_date = shipping_scheduled_date # 상차 예정일 업데이트
            order.payment_amount = new_payment_amount # 최종 결제금액 업데이트
            
            # 지방 주문 여부 사용자 선택 (체크박스는 체크되지 않으면 폼에 포함되지 않음)
            order.is_regional = is_regional_new
            
            # 자가실측 여부 사용자 선택 (체크박스는 체크되지 않으면 폼에 포함되지 않음)
            order.is_self_measurement = is_self_measurement_new
            
            # 수납장 여부 사용자 선택 (체크박스는 체크되지 않으면 폼에 포함되지 않음)
            is_cabinet_new = 'is_cabinet' in request.form
            if order.is_cabinet != is_cabinet_new:
                changes['is_cabinet'] = {'old': order.is_cabinet, 'new': is_cabinet_new}
            order.is_cabinet = is_cabinet_new
            
            # 수납장 상태 업데이트 (수납장으로 설정되면 기본 상태를 RECEIVED로 설정)
            if is_cabinet_new and not order.cabinet_status:
                order.cabinet_status = 'RECEIVED'
            elif not is_cabinet_new:
                order.cabinet_status = None

            # 시공 구분 업데이트
            order.construction_type = construction_type_new
            
            # 지방 주문 체크박스 필드 업데이트
            if order.is_regional:
                regional_fields = [
                    'measurement_completed',
                    'regional_sales_order_upload',
                    'regional_blueprint_sent',
                    'regional_order_upload',
                    'regional_cargo_sent',
                    'regional_construction_info_sent'
                ]
                
                for field in regional_fields:
                    if field in request.form:
                        # 체크박스는 체크된 경우에만 폼 데이터에 포함됨
                        setattr(order, field, True)
                    else:
                        # 폼에 없으면 체크되지 않은 것으로 간주
                        setattr(order, field, False)
            
            db.commit()
            
            # 필드 레이블 정의 (필드명 -> 한글 레이블 매핑)
            field_labels = {
                'received_date': '접수일',
                'received_time': '접수시간',
                'customer_name': '고객명',
                'phone': '전화번호',
                'address': '주소',
                'product': '제품',
                'options': '옵션 상세',
                'notes': '비고',
                'status': '상태',
                'measurement_date': '실측일',
                'measurement_time': '실측시간',
                'completion_date': '설치완료일',
                'manager_name': '담당자',
                'payment_amount': '결제금액',
                'is_regional': '지방 주문',
                'is_self_measurement': '자가실측',
                'is_cabinet': '수납장',
                'measurement_completed': '실측완료',
                'construction_type': '시공 구분',
                'regional_sales_order_upload': '영업발주 업로드',
                'regional_blueprint_sent': '도면 발송',
                'regional_order_upload': '발주 업로드',
                'regional_cargo_sent': '화물 발송',
                'regional_construction_info_sent': '시공정보 발송',
                'shipping_scheduled_date': '상차 예정일' # 로그용 레이블 추가
            }
            
            # 변경된 필드만 필터링하여 로그 메시지 구성
            change_descriptions = []
            for field, values in changes.items():
                if field in field_labels:
                    # None 값 안전하게 처리
                    old_val = values.get('old', '') or '없음'
                    new_val = values.get('new', '') or '없음'
                    
                    # 옵션은 JSON 문자열이므로 특별 처리
                    if field == 'options':
                        try:
                            old_json = json.loads(old_val) if old_val != '없음' and old_val else None
                            new_json = json.loads(new_val) if new_val != '없음' and new_val else None
                            
                            # 두 JSON이 실질적으로 동일한 값을 가지는지 확인
                            if old_json and new_json:
                                # 온라인 옵션 요약 비교
                                old_option_type = old_json.get('option_type', '')
                                new_option_type = new_json.get('option_type', '')
                                
                                # 타입이 다르면 변경된 것으로 간주
                                if old_option_type != new_option_type:
                                    if old_option_type == 'online':
                                        old_display = old_json.get('online_options_summary', '') or '없음'
                                    elif old_option_type == 'direct':
                                        details = old_json.get('details', {})
                                        old_display = '직접입력: ' + (details.get('product_name', '') or details.get('color', '') or '옵션')
                                    else:
                                        old_display = '옵션 있음'
                                        
                                    if new_option_type == 'online':
                                        new_display = new_json.get('online_options_summary', '') or '없음'
                                    elif new_option_type == 'direct':
                                        details = new_json.get('details', {})
                                        new_display = '직접입력: ' + (details.get('product_name', '') or details.get('color', '') or '옵션')
                                    else:
                                        new_display = '옵션 있음'
                                # 타입이 같고 온라인 옵션인 경우    
                                elif old_option_type == 'online':
                                    old_summary = old_json.get('online_options_summary', '')
                                    new_summary = new_json.get('online_options_summary', '')
                                    
                                    # 내용이 같으면 건너뛰기
                                    if old_summary == new_summary:
                                        continue
                                        
                                    old_display = old_summary or '없음'
                                    new_display = new_summary or '없음'
                                # 타입이 같고 직접 입력 옵션인 경우
                                elif old_option_type == 'direct':
                                    old_details = old_json.get('details', {})
                                    new_details = new_json.get('details', {})
                                    
                                    # 주요 필드만 비교 (product_name, color)
                                    old_key_values = old_details.get('product_name', '') + ' ' + old_details.get('color', '')
                                    new_key_values = new_details.get('product_name', '') + ' ' + new_details.get('color', '')
                                    
                                    # 내용이 같으면 건너뛰기
                                    if old_key_values.strip() == new_key_values.strip():
                                        continue
                                        
                                    old_display = old_details.get('product_name', '') or old_details.get('color', '') or '옵션 있음'
                                    new_display = new_details.get('product_name', '') or new_details.get('color', '') or '옵션 있음'
                                else:
                                    # 기타 경우는 간단하게 표시
                                    old_display = '옵션 있음'
                                    new_display = '옵션 있음'
                                    
                                    # 내용이 같아 보이면 건너뛰기
                                    if json.dumps(old_json, sort_keys=True) == json.dumps(new_json, sort_keys=True):
                                        continue
                            elif not old_json and not new_json:
                                # 둘 다 없거나 빈 값이면 건너뛰기
                                continue
                            else:
                                # 한쪽만 값이 있는 경우 (추가 또는 삭제)
                                if old_json:
                                    if old_json.get('option_type') == 'online':
                                        old_display = old_json.get('online_options_summary', '') or '옵션 있음'
                                    elif old_json.get('option_type') == 'direct':
                                        details = old_json.get('details', {})
                                        old_display = details.get('product_name', '') or details.get('color', '') or '직접입력 옵션'
                                    else:
                                        old_display = '옵션 있음'
                                else:
                                    old_display = '없음'
                                    
                                if new_json:
                                    if new_json.get('option_type') == 'online':
                                        new_display = new_json.get('online_options_summary', '') or '옵션 있음'
                                    elif new_json.get('option_type') == 'direct':
                                        details = new_json.get('details', {})
                                        new_display = details.get('product_name', '') or details.get('color', '') or '직접입력 옵션'
                                    else:
                                        new_display = '옵션 있음'
                                else:
                                    new_display = '없음'
                        except Exception as e:
                            # JSON 파싱 실패 시 원본 값 비교
                            old_display = old_val if old_val != '없음' else '없음'
                            new_display = new_val if new_val != '없음' else '없음'
                            
                            # 값이 같거나 둘 다 빈 값이면 건너뛰기
                            if old_display == new_display or (not old_display.strip() and not new_display.strip()):
                                continue
                    else:
                        # 다른 필드들은 문자열로 변환하여 비교
                        old_display = str(old_val).strip() if old_val != '없음' else '없음'
                        new_display = str(new_val).strip() if new_val != '없음' else '없음'
                        
                        # 값이 같으면 건너뛰기 (공백 제거 후 비교)
                        if old_display == new_display:
                            continue
                        
                        # 상태 코드를 한글 상태명으로 변환
                        if field == 'status':
                            old_display = STATUS.get(old_display, old_display)
                            new_display = STATUS.get(new_display, new_display)
                    
                    # 실제 값이 변경된 경우에만 표시 (위에서 필터링 이미 됨)
                    change_descriptions.append(f"{field_labels[field]}: {old_display} ⇒ {new_display}")
            
            # 주문 번호와 고객명은 필수로 포함
            user_for_log = get_user_by_id(session['user_id'])
            user_name_for_log = user_for_log.name if user_for_log else "Unknown user"
            log_message_prefix = f"주문 #{order_id} ({customer_name}) 수정 - 담당자: {user_name_for_log} (ID: {session.get('user_id')})"
            
            # 변경 내역이 있으면 추가
            if change_descriptions:
                log_message = f"{log_message_prefix} | 변경내용: {'; '.join(change_descriptions)}"
            else:
                log_message = f"{log_message_prefix} | 변경내용 없음"
            
            # 로그 저장
            log_access(log_message, session.get('user_id'))
            
            flash('주문이 성공적으로 수정되었습니다.', 'success')
            
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return jsonify({'status': 'success'})
            
            # 수정 후 원래 보던 페이지로 리디렉션
            referrer = request.form.get('referrer')
            if referrer:
                # Basic check to prevent open redirection vulnerabilities
                from urllib.parse import urlparse
                if urlparse(referrer).netloc == request.host:
                    return redirect(referrer)
            
            # Fallback to the main index page
            return redirect(url_for('index'))
        except ValueError as e:
            db.rollback()
            flash(f'입력 데이터 오류: {str(e)}', 'error')
            
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return jsonify({'status': 'error', 'message': str(e)})
        except Exception as e:
            db.rollback()
            flash(f'주문 수정 중 오류가 발생했습니다: {str(e)}', 'error')
            
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return jsonify({'status': 'error', 'message': '시스템 오류가 발생했습니다.'})
            
            # 오류 발생 시 현재 데이터로 페이지 다시 로드
            return render_template(
                'edit_order.html', 
                order=order,
                option_type=option_type,
                online_options=online_options,
                direct_options=direct_options
            )
    
    # GET 요청에 대한 최종 반환 - 미리 처리된 옵션 데이터를 직접 템플릿에 전달
    # 현재 URL 쿼리 파라미터를 템플릿에 전달하여 필터 상태 유지
    preserved_args = get_preserved_filter_args(request.args)
    return render_template(
        'edit_order.html', 
        order=order,
        option_type=option_type,
        online_options=online_options,
        direct_options=direct_options,
        preserved_args=preserved_args
    )

@app.route('/delete/<int:order_id>')
@login_required
@role_required(['ADMIN', 'MANAGER'])
def delete_order(order_id):
    try:
        db = get_db()
        
        # Get order from database
        order = db.query(Order).filter(Order.id == order_id, Order.status != 'DELETED').first()
        
        if not order:
            flash('주문을 찾을 수 없거나 이미 삭제되었습니다.', 'error')
            return redirect(url_for('index'))
        
        # Save original status before deletion
        original_status = order.status
        deleted_at = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        customer_name_for_log = order.customer_name # 로그용 고객명
        
        # Soft delete by updating status and recording original status
        order.status = 'DELETED'
        order.original_status = original_status
        order.deleted_at = deleted_at
        
        db.commit()
        
        user_for_log = get_user_by_id(session['user_id'])
        user_name_for_log = user_for_log.name if user_for_log else "Unknown user"
        log_access(f"주문 #{order_id} ({customer_name_for_log}) 삭제 - 담당자: {user_name_for_log} (ID: {session.get('user_id')})", session.get('user_id'))
        
        flash('주문이 휴지통으로 이동되었습니다.', 'success')
    except Exception as e:
        db.rollback()
        flash(f'주문 삭제 중 오류가 발생했습니다: {str(e)}', 'error')
    
    # 원래 페이지 필터링 상태 유지
    redirect_args = get_preserved_filter_args(request.args)
    return redirect(url_for('index', **redirect_args))

@app.route('/trash')
@login_required
@role_required(['ADMIN', 'MANAGER'])
def trash():
    search_term = request.args.get('search', '')
    
    db = get_db()
    
    # Base query for deleted orders
    query = db.query(Order).filter(Order.status == 'DELETED')
    
    # Add search filter if provided
    if search_term:
        search_pattern = f"%{search_term}%"
        query = query.filter(
            (Order.customer_name.like(search_pattern)) |
            (Order.phone.like(search_pattern)) |
            (Order.address.like(search_pattern)) |
            (Order.product.like(search_pattern)) |
            (Order.options.like(search_pattern)) |
            (Order.notes.like(search_pattern))
        )
    
    # Order by deleted_at timestamp
    orders = query.order_by(Order.deleted_at.desc()).all()
    
    return render_template('trash.html', orders=orders, search_term=search_term)

@app.route('/restore_orders', methods=['POST'])
@login_required
@role_required(['ADMIN', 'MANAGER'])
def restore_orders():
    selected_ids = request.form.getlist('selected_order')
    
    if not selected_ids:
        flash('복원할 주문을 선택해주세요.', 'warning')
        return redirect(url_for('trash'))
    
    try:
        db = get_db()
        
        for order_id in selected_ids:
            # Get order by id
            order = db.query(Order).filter(Order.id == order_id, Order.status == 'DELETED').first()
            
            if order:
                # Get original status or default to RECEIVED
                original_status = order.original_status if order.original_status else 'RECEIVED'
                
                # Restore order by updating status
                order.status = original_status
                order.original_status = None
                order.deleted_at = None
        
        db.commit()
        
        # 로그 기록 (한글로 변경)
        log_access(f"주문 {len(selected_ids)}개 복원", session.get('user_id'), {"count": len(selected_ids)})
        
        flash(f'{len(selected_ids)}개 주문이 성공적으로 복원되었습니다.', 'success')
    except Exception as e:
        db.rollback()
        flash(f'주문 복원 중 오류가 발생했습니다: {str(e)}', 'error')
    
    return redirect(url_for('trash'))

@app.route('/permanent_delete_orders', methods=['POST'])
@login_required
@role_required(['ADMIN'])
def permanent_delete_orders():
    selected_ids = request.form.getlist('selected_order')
    
    if not selected_ids:
        flash('영구 삭제할 주문을 선택해주세요.', 'warning')
        return redirect(url_for('trash'))
    
    try:
        db = get_db()
        
        for order_id in selected_ids:
            # Get order by id
            order = db.query(Order).filter(Order.id == order_id).first()
            
            if order:
                # Permanently delete order from database
                db.delete(order)
        
        db.commit()
        
        # 주문 ID 재정렬 실행
        reset_order_ids(db)
        
        # 로그 기록 (한글로 변경)
        log_access(f"주문 {len(selected_ids)}개 영구 삭제", session.get('user_id'), {"count": len(selected_ids)})
        
        flash(f'{len(selected_ids)}개의 주문이 영구적으로 삭제되었습니다.', 'success')
    except Exception as e:
        db.rollback()
        flash(f'주문 영구 삭제 중 오류가 발생했습니다: {str(e)}', 'error')
    
    return redirect(url_for('trash'))

@app.route('/permanent_delete_all_orders', methods=['POST'])
@login_required
@role_required(['ADMIN'])
def permanent_delete_all_orders():
    try:
        db = get_db()
        
        # 휴지통에 있는 모든 주문 조회
        deleted_orders = db.query(Order).filter(Order.status == 'DELETED').all()
        
        if not deleted_orders:
            flash('휴지통에 삭제할 주문이 없습니다.', 'warning')
            return redirect(url_for('trash'))
            
        deleted_count = len(deleted_orders)
            
        # 모든 주문 영구 삭제
        for order in deleted_orders:
            db.delete(order)
            
        db.commit()
        
        # 주문 ID 재정렬 실행
        reset_order_ids(db)
            
        # Log the action
        log_access(f"모든 주문 영구 삭제 ({deleted_count}개 항목)", session.get('user_id'), {"count": deleted_count})
            
        flash(f'모든 주문({deleted_count}개)이 영구적으로 삭제되었습니다.', 'success')
    except Exception as e:
        db.rollback()
        flash(f'주문 영구 삭제 중 오류가 발생했습니다: {str(e)}', 'error')
        
    return redirect(url_for('trash'))

def reset_order_ids(db):
    """주문 ID를 1부터 연속적으로 재정렬합니다."""
    try:
        # 임시 테이블 생성
        db.execute(text("CREATE TEMPORARY TABLE temp_order_mapping (old_id INT, new_id INT)"))
        
        # 현재 존재하는 모든 주문 목록 (삭제되지 않은 주문만)
        orders = db.query(Order).filter(Order.status != 'DELETED').order_by(Order.id).all()
        
        # 새로운 ID 값 배정
        new_id = 0
        for new_id, order in enumerate(orders, 1):
            # 이전 ID와 새 ID 매핑 저장
            if order.id != new_id:
                db.execute(text("INSERT INTO temp_order_mapping (old_id, new_id) VALUES (:old_id, :new_id)"), 
                          {"old_id": order.id, "new_id": new_id})
        
        # 실제 ID 업데이트 쿼리 준비
        mapping_exists = db.execute(text("SELECT COUNT(*) FROM temp_order_mapping")).scalar() > 0
        
        # 시퀀스 재설정 준비 (최대 ID 값 + 1로 설정)
        max_id = new_id if orders else 0  # 주문이 없으면 0부터 시작
        
        if mapping_exists:
            # ID 변경이 필요한 경우에만 진행
            # 매핑 테이블을 사용해 주문 ID 업데이트
            db.execute(text("""
                UPDATE orders 
                SET id = (SELECT new_id FROM temp_order_mapping WHERE temp_order_mapping.old_id = orders.id)
                WHERE id IN (SELECT old_id FROM temp_order_mapping)
            """))
            
            # 로그 데이터 업데이트 기능 제거
        
        # 시퀀스 재설정 (PostgreSQL 전용) - 항상 실행
        try:
            # 시퀀스 이름 확인
            seq_query = "SELECT pg_get_serial_sequence('orders', 'id')"
            seq_name = db.execute(text(seq_query)).scalar()
            
            if seq_name:
                # 정확한 시퀀스 이름을 사용하여 재설정
                db.execute(text(f"ALTER SEQUENCE {seq_name} RESTART WITH {max_id + 1}"))
                # 시퀀스 재설정 완료
            else:
                # 이름을 찾지 못한 경우 기본 이름 사용
                db.execute(text(f"ALTER SEQUENCE orders_id_seq RESTART WITH {max_id + 1}"))
                # 기본 시퀀스 재설정 완료
        except Exception as seq_error:
            # 시퀀스 재설정 중 오류 발생 (무시)
            # 기본 이름을 사용해서 시도
            try:
                db.execute(text(f"ALTER SEQUENCE orders_id_seq RESTART WITH {max_id + 1}"))
            except:
                pass
            
        db.commit()
        
        # 임시 테이블 삭제
        db.execute(text("DROP TABLE IF EXISTS temp_order_mapping"))
        
    except Exception as e:
        db.rollback()
        # 오류 발생 시 임시 테이블 제거 시도
        try:
            db.execute(text("DROP TABLE IF EXISTS temp_order_mapping"))
        except:
            pass
        # 주문 ID 재정렬 중 오류 발생 (무시)
        raise e

@app.route('/bulk_action', methods=['POST'])
@login_required
@role_required(['ADMIN', 'MANAGER'])
def bulk_action():
    action = request.form.get('action')
    selected_ids = request.form.getlist('selected_order')
    
    if not selected_ids:
        flash('작업할 주문을 선택해주세요.', 'warning')
        # 원래 페이지 필터링 상태 유지
        redirect_args = get_preserved_filter_args(request.args)
        return redirect(url_for('index', **redirect_args))
    
    if not action:
        flash('수행할 작업을 선택해주세요.', 'warning')
        # 원래 페이지 필터링 상태 유지
        redirect_args = get_preserved_filter_args(request.args)
        return redirect(url_for('index', **redirect_args))

    # db 변수 미리 선언
    db = None
    current_user_id = session.get('user_id')
    processed_count = 0
    failed_count = 0
        
    try:
        db = get_db()
        if action == 'delete':
            for order_id in selected_ids:
                order = db.query(Order).filter(Order.id == order_id, Order.status != 'DELETED').first()
                if order:
                    original_status = order.status
                    deleted_at = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    order.status = 'DELETED'
                    order.original_status = original_status
                    order.deleted_at = deleted_at
                    log_access(f"주문 #{order_id} 삭제 (일괄 작업)", current_user_id, {"order_id": order_id})
                    processed_count += 1
                else:
                    failed_count += 1
        
        # --- 주문 복사 로직 추가 --- 
        elif action == 'copy':
            now = datetime.datetime.now()
            today_str = now.strftime('%Y-%m-%d')
            time_str = now.strftime('%H:%M')
            
            for order_id in selected_ids:
                original_order = db.query(Order).get(order_id)
                if original_order:
                    # Order 객체 복사 (ID 등 자동 생성 필드는 제외)
                    copied_order = Order()
                    
                    # 필드 복사 (수정 필요한 필드 제외)
                    for column in Order.__table__.columns:
                        col_name = column.name
                        if col_name not in ['id', 'status', 'received_date', 'received_time',
                                             'customer_name', 'notes', 'measurement_date', 'measurement_time', 
                                             'completion_date', 'original_status', 'deleted_at']:
                            setattr(copied_order, col_name, getattr(original_order, col_name))
                    
                    # 필드 수정
                    copied_order.status = 'RECEIVED' # 상태는 '접수'로
                    copied_order.received_date = today_str # 접수일은 오늘 날짜
                    copied_order.received_time = time_str # 접수시간은 현재 시간
                    copied_order.customer_name = f"[복사: 원본 #{original_order.id}] {original_order.customer_name}"
                    
                    original_notes = original_order.notes or ""
                    copied_order.notes = f"원본 주문 #{original_order.id} 에서 복사됨.\n---\n" + original_notes
                    
                    # 날짜/시간 정보 초기화
                    copied_order.measurement_date = None
                    copied_order.measurement_time = None
                    copied_order.completion_date = None
                    
                    db.add(copied_order)
                    db.flush() # 새 ID를 가져오기 위해 flush
                    
                    log_access(f"주문 #{original_order.id}를 새 주문 #{copied_order.id}로 복사 (일괄 작업)", 
                               current_user_id, {"original_order_id": original_order.id, "new_order_id": copied_order.id})
                    processed_count += 1
                else:
                    failed_count += 1
        # --- 주문 복사 로직 끝 --- 
            
        elif action.startswith('status_'):
            new_status = action.split('_', 1)[1]
            if new_status in STATUS:
                for order_id in selected_ids:
                    order = db.query(Order).filter(Order.id == order_id, Order.status != 'DELETED').first()
                    if order and order.status != new_status:
                        old_status = order.status
                        order.status = new_status
                        # 상태 한글 변환
                        old_status_kr = STATUS.get(old_status, old_status)
                        new_status_kr = STATUS.get(new_status, new_status)
                        # 한글 로그 메시지
                        log_access(f"주문 #{order_id} 상태 변경: {old_status_kr} => {new_status_kr} (일괄 작업)", 
                                   current_user_id, {"order_id": order_id, "old_status": old_status, "new_status": new_status})
                        processed_count += 1
                    elif not order:
                         failed_count += 1 # 존재하지 않거나 삭제된 주문
                    # 상태가 이미 동일하면 처리하지 않음 (processed_count 증가 안함)
            else:
                 flash("'" + new_status + "'" + '는 유효하지 않은 상태입니다.', 'error')
                 # 원래 페이지 필터링 상태 유지
                 redirect_args = get_preserved_filter_args(request.args)
                 return redirect(url_for('index', **redirect_args))

        # 모든 변경 사항을 한번에 커밋
        db.commit()

        # 성공/실패 메시지 생성
        if action.startswith('status_'):
            status_code = action.split('_', 1)[1]
            status_name = STATUS.get(status_code, status_code)
            action_display_name = f"상태를 '{status_name}'(으)로 변경"
        elif action == 'copy':
            action_display_name = "'복사'"
        elif action == 'delete':
            action_display_name = "'삭제'"
        else:
            action_display_name = f"\'{action}\'"
        
        success_msg = f"{processed_count}개의 주문에 대해 {action_display_name} 작업을 완료했습니다."
        if failed_count > 0:
            warning_msg = f"{failed_count}개의 주문은 처리할 수 없었습니다 (이미 삭제되었거나 존재하지 않음)."
            flash(warning_msg, 'warning')
        
        if processed_count > 0:
             flash(success_msg, 'success')
        elif failed_count == len(selected_ids):
             flash('선택한 주문을 처리할 수 없습니다.', 'error')
        else:
             flash('변경된 사항이 없습니다.', 'info')

    except Exception as e:
        if db:
            db.rollback()
        flash(f'일괄 작업 중 오류 발생: {str(e)}', 'error')
        current_app.logger.error(f"일괄 작업 실패: {e}", exc_info=True)
    
    # 원래 페이지 필터링 상태 유지
    redirect_args = get_preserved_filter_args(request.args)
    return redirect(url_for('index', **redirect_args))

@app.route('/upload', methods=['GET', 'POST'])
@login_required
@role_required(['ADMIN', 'MANAGER'])
def upload_excel():
    if request.method == 'POST':
        # check if the post request has the file part
        if 'excel_file' not in request.files:
            flash('파일이 선택되지 않았습니다.', 'error')
            log_access(f"엑셀 업로드 실패: 파일이 선택되지 않음", session.get('user_id'))
            return redirect(request.url)
        
        file = request.files['excel_file']
        
        if file.filename == '':
            flash('파일이 선택되지 않았습니다.', 'error')
            log_access(f"엑셀 업로드 실패: 빈 파일명", session.get('user_id'))
            return redirect(request.url)
        
        if file and allowed_file(file.filename):
            filename = secure_filename(file.filename)
            file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            file.save(file_path)
            
            db = get_db() # db 변수를 try 블록 외부에서도 사용 가능하도록 이동
            try:
                # Process the Excel file with pandas
                df = pd.read_excel(file_path)
                
                # Check for required columns (한글 컬럼명으로 변경)
                required_columns = ['접수일', '고객명', '전화번호', '주소', '제품']
                missing_columns = [col for col in required_columns if col not in df.columns]
                
                if missing_columns:
                    missing_cols_str = ", ".join(missing_columns)
                    flash(f'엑셀 파일에 필수 컬럼이 누락되었습니다: {missing_cols_str}', 'error')
                    log_access(f"엑셀 업로드 실패: 필수 컬럼 누락 ({missing_cols_str}) - 파일명: {filename}", session.get('user_id'))
                    # 파일 삭제 로직 추가 (오류 시)
                    try:
                        os.remove(file_path)
                    except OSError as e: # 구체적인 에러 타입 명시
                        # 업로드된 파일 삭제 오류 (무시)
                        log_access(f"업로드된 파일 삭제 오류: {file_path} - {e}", session.get('user_id'),
                                   {"filename": file_path, "error": str(e)})
                    return redirect(request.url)
                
                # Connect to database (이미 위에서 get_db() 호출)
                
                # Process each row
                order_count = 0
                added_order_ids = []
                
                for index, row in df.iterrows():
                    # Convert fields to the right format and provide defaults (한글 컬럼명 사용)
                    
                    # 날짜 필드 처리 (pd.to_datetime 사용)
                    received_date_dt = pd.to_datetime(row['접수일'], errors='coerce')
                    received_date = received_date_dt.strftime('%Y-%m-%d') if pd.notna(received_date_dt) else datetime.datetime.now().strftime('%Y-%m-%d')

                    measurement_date_dt = pd.to_datetime(row.get('실측일'), errors='coerce') # .get으로 안전하게 접근
                    measurement_date = measurement_date_dt.strftime('%Y-%m-%d') if pd.notna(measurement_date_dt) else None
                    
                    completion_date_dt = pd.to_datetime(row.get('설치완료일'), errors='coerce') # .get으로 안전하게 접근
                    completion_date = completion_date_dt.strftime('%Y-%m-%d') if pd.notna(completion_date_dt) else None

                    # 시간 필드 처리 (다양한 입력 형식 고려)
                    received_time_raw = row.get('접수시간') # .get으로 안전하게 접근
                    received_time = None
                    if pd.notna(received_time_raw):
                        if isinstance(received_time_raw, datetime.time):
                            received_time = received_time_raw.strftime('%H:%M')
                        elif isinstance(received_time_raw, datetime.datetime): # datetime 객체로 읽혔을 경우
                            received_time = received_time_raw.strftime('%H:%M')
                        elif isinstance(received_time_raw, str):
                            # 간단한 형식 검사 (예: HH:MM) - 필요시 정규식 등으로 강화
                            if re.match(r'^\d{1,2}:\d{2}$', received_time_raw.strip()):
                                received_time = received_time_raw.strip()
                            else:
                                try: # 엑셀에서 0.xxxx 와 같은 숫자형식으로 시간을 읽는 경우 대비
                                    time_float = float(received_time_raw)
                                    hours = int(time_float * 24)
                                    minutes = int((time_float * 24 * 60) % 60)
                                    received_time = f"{hours:02d}:{minutes:02d}"
                                except (ValueError, TypeError):
                                    # Warning: Invalid time format for 접수시간 (using default)
                                    received_time = None # 유효하지 않으면 None
                        # 추가: Excel에서 시간 형식이 숫자로 (예: 0.5 = 12:00 PM) 읽히는 경우 처리
                        elif isinstance(received_time_raw, (int, float)):
                            try:
                                # 소수점 형태의 시간을 HH:MM으로 변환
                                total_seconds = int(received_time_raw * 24 * 60 * 60)
                                hours = total_seconds // 3600
                                minutes = (total_seconds % 3600) // 60
                                received_time = f"{hours:02d}:{minutes:02d}"
                            except Exception:
                                                                   # Warning: Could not convert numeric time for 접수시간 (using default)
                                 received_time = None


                    measurement_time_raw = row.get('실측시간') # .get으로 안전하게 접근
                    measurement_time = None
                    if pd.notna(measurement_time_raw):
                        if isinstance(measurement_time_raw, datetime.time):
                            measurement_time = measurement_time_raw.strftime('%H:%M')
                        elif isinstance(measurement_time_raw, datetime.datetime):
                            measurement_time = measurement_time_raw.strftime('%H:%M')
                        elif isinstance(measurement_time_raw, str):
                            if re.match(r'^\d{1,2}:\d{2}$', measurement_time_raw.strip()):
                                measurement_time = measurement_time_raw.strip()
                            else:
                                try:
                                    time_float = float(measurement_time_raw)
                                    hours = int(time_float * 24)
                                    minutes = int((time_float * 24 * 60) % 60)
                                    measurement_time = f"{hours:02d}:{minutes:02d}"
                                except (ValueError, TypeError):
                                    # Warning: Invalid time format for 실측시간 (using default)
                                    measurement_time = None
                        elif isinstance(measurement_time_raw, (int, float)):
                            try:
                                total_seconds = int(measurement_time_raw * 24 * 60 * 60)
                                hours = total_seconds // 3600
                                minutes = (total_seconds % 3600) // 60
                                measurement_time = f"{hours:02d}:{minutes:02d}"
                            except Exception:
                                 # Warning: Could not convert numeric time for 실측시간 (using default)
                                 measurement_time = None
                    
                    # Handle options column if it exists (한글 컬럼명 '옵션')
                    options_raw = row.get('옵션') # .get으로 안전하게 접근
                    options = str(options_raw) if pd.notna(options_raw) else None # 어떤 형식이든 문자열로 저장
                    
                    # Handle notes column if it exists (한글 컬럼명 '비고')
                    notes_raw = row.get('비고') # .get으로 안전하게 접근
                    notes = str(notes_raw) if pd.notna(notes_raw) else None # 문자열로 저장

                    manager_name_raw = row.get('담당자') # .get으로 안전하게 접근
                    manager_name = str(manager_name_raw) if pd.notna(manager_name_raw) else None

                    # payment_amount 처리 (숫자형으로, 콤마 제거)
                    payment_amount_raw = row.get('결제금액')
                    payment_amount = 0 # 기본값 0
                    if pd.notna(payment_amount_raw):
                        try:
                            # 문자열일 경우 콤마 제거 후 정수 변환
                            if isinstance(payment_amount_raw, str):
                                payment_amount_str = payment_amount_raw.replace(',', '')
                                payment_amount = int(float(payment_amount_str)) # 소수점도 고려하여 float으로 먼저 변환
                            elif isinstance(payment_amount_raw, (int, float)):
                                payment_amount = int(payment_amount_raw)
                        except ValueError:
                            # Warning: Invalid payment amount format for 결제금액 (defaulting to 0)
                            payment_amount = 0 # 변환 실패 시 0

                    new_order = Order(
                        customer_name=str(row['고객명']) if pd.notna(row['고객명']) else '', # 문자열로 명시적 변환
                        phone=str(row['전화번호']) if pd.notna(row['전화번호']) else '', # 문자열로 명시적 변환
                        address=str(row['주소']) if pd.notna(row['주소']) else '', # 문자열로 명시적 변환
                        product=str(row['제품']) if pd.notna(row['제품']) else '', # 문자열로 명시적 변환
                        options=options,
                        notes=notes,
                        received_date=received_date,
                        received_time=received_time,
                        status='RECEIVED',  # Default status
                        measurement_date=measurement_date,
                        measurement_time=measurement_time,
                        completion_date=completion_date,
                        manager_name=manager_name,
                        payment_amount=payment_amount, # 추가
                        # 추가된 상태별 날짜 필드
                        scheduled_date=request.form.get('scheduled_date'),
                        as_received_date=request.form.get('as_received_date'),
                        as_completed_date=request.form.get('as_completed_date'),
                        # 지방 주문 여부 기본값은 False (사용자가 수동으로 변경해야 함)
                        is_regional=False
                    )
                    
                    db.add(new_order)
                    db.flush() # ID 할당을 위해 flush
                    added_order_ids.append(new_order.id) # 추가된 주문 ID 저장
                    order_count += 1
                
                db.commit()
                flash(f'{order_count}개의 주문이 성공적으로 등록되었습니다.', 'success')
                log_access(f"엑셀 업로드 성공: {filename} 파일에서 {order_count}개 주문 추가", session.get('user_id'), 
                           {"filename": filename, "orders_added": order_count, "order_ids": added_order_ids})
                
            except Exception as e:
                if db: # db 객체가 초기화된 경우에만 롤백 시도
                    db.rollback()
                error_message = f'엑셀 파일 처리 중 오류가 발생했습니다: {str(e)}'
                flash(error_message, 'error')
                log_access(f"엑셀 업로드 실패: {filename} 파일 처리 중 오류 - {str(e)}", session.get('user_id'),
                           {"filename": filename, "error": str(e)})
            
            # Delete the file after processing (성공/실패 여부와 관계없이)
            try:
                os.remove(file_path)
            except OSError as e: # 구체적인 에러 타입 명시
                # Error deleting uploaded file (ignored)
                log_access(f"업로드된 파일 삭제 오류: {file_path} - {e}", session.get('user_id'),
                           {"filename": file_path, "error": str(e)})

            return redirect(url_for('index'))
        else:
            flash('허용되지 않은 파일 형식입니다. .xlsx 또는 .xls 파일만 업로드 가능합니다.', 'error')
            log_access(f"엑셀 업로드 실패: 허용되지 않은 파일 형식 - {file.filename}", session.get('user_id'),
                       {"filename": file.filename})
            return redirect(request.url)
    
    return render_template('upload.html')

@app.route('/download_excel')
@login_required
def download_excel():
    db = get_db()
    status_filter = request.args.get('status')
    search_query = request.args.get('search', '').strip()
    sort_column = request.args.get('sort', 'id') # 정렬 기준
    sort_direction = request.args.get('direction', 'desc') # 정렬 방향
    
    # 기본 쿼리 생성 (삭제되지 않은 주문만)
    query = db.query(Order).filter(Order.deleted_at.is_(None))
    
    # 상태 필터 적용
    if status_filter:
        # 접수 탭에서는 RECEIVED와 ON_HOLD 상태를 모두 표시
        if status_filter == 'RECEIVED':
            query = query.filter(Order.status.in_(['RECEIVED', 'ON_HOLD']))
        else:
            query = query.filter(Order.status == status_filter)
    
    # 검색어 필터 적용 (index 함수와 동일하게)
    if search_query:
        search_term = f"%{search_query}%"
        query = query.filter( 
            or_(
                Order.id.cast(String).like(search_term),  # integer 타입을 String으로 캐스팅
                Order.received_date.like(search_term),
                Order.received_time.like(search_term),
                Order.customer_name.like(search_term),
                Order.phone.like(search_term),
                Order.address.like(search_term),
                Order.product.like(search_term),
                Order.options.like(search_term),
                Order.notes.like(search_term),
                Order.status.like(search_term),
                Order.measurement_date.like(search_term),
                Order.measurement_time.like(search_term),
                Order.completion_date.like(search_term),
                Order.manager_name.like(search_term),
                # payment_amount는 숫자형이므로 캐스팅 필요
                func.cast(Order.payment_amount, String).like(search_term)
            )
        )

    # 컬럼별 입력 필터 적용 (index 함수와 동일한 로직으로 변경)
    filterable_columns = [
        'id', 'received_date', 'received_time', 'customer_name', 'phone', 
        'address', 'product', 'options', 'notes', 'status', 
        'measurement_date', 'measurement_time', 'completion_date', 'manager_name', 'payment_amount'
    ]
    for column_name in filterable_columns:
        filter_value = request.args.get(f'filter_{column_name}', '').strip() # get 대신 getlist 사용하지 않음
        if filter_value:
            if hasattr(Order, column_name):
                try:
                    column_attr = getattr(Order, column_name)
                    # 숫자 타입 컬럼일 경우 문자열로 캐스팅 후 LIKE 적용
                    if isinstance(column_attr.type.python_type(), (int, float)):
                         query = query.filter(column_attr.cast(String).like(f"%{filter_value}%"))
                    else:
                         query = query.filter(column_attr.like(f"%{filter_value}%"))
                except AttributeError:
                    # 컬럼이 없거나 LIKE 사용 불가 시 경고 (index 함수와 동일)
                    # Warning: Column not found or cannot be filtered with LIKE in download_excel
                    pass
            else:
                 # Warning: Column not found in Order model in download_excel
                 pass
                
    # 정렬 적용 (루프 바깥으로 이동)
    if hasattr(Order, sort_column):
        column_to_sort = getattr(Order, sort_column)
        if sort_direction == 'asc':
            query = query.order_by(column_to_sort.asc())
        else:
            query = query.order_by(column_to_sort.desc())
    else:
        query = query.order_by(Order.id.desc()) # 기본 정렬

    orders = query.all()
    
    # 다운로드할 주문 ID 목록
    downloaded_order_ids = [order.id for order in orders] if orders else []

    if not orders:
        flash('다운로드할 데이터가 없습니다.', 'warning')
        return redirect(request.referrer or url_for('index'))
    
    # 데이터를 Pandas DataFrame으로 변환
    orders_data = []
    for order in orders:
        order_dict = order.to_dict()
        # 옵션을 한글로 변환하는 로직 추가
        order_dict['options'] = format_options_for_display(order.options) # 전역 함수 호출
        orders_data.append(order_dict)

    df = pd.DataFrame(orders_data)
    
    # 상태 코드를 한글 이름으로 변경
    if 'status' in df.columns:
        df['status'] = df['status'].map(STATUS).fillna(df['status'])
        
    # 필요한 컬럼 선택 및 순서 지정
    excel_columns = [
        'id', 'received_date', 'received_time', 'customer_name', 'phone', 'address', 
        'product', 'options', 'notes', 'payment_amount',
        'measurement_date', 'measurement_time', 'completion_date', 
        'manager_name', 'status'
    ]
    # DataFrame에 없는 컬럼이 excel_columns에 포함되어 있을 경우 KeyError 발생 방지
    df_excel_columns = [col for col in excel_columns if col in df.columns]
    df_excel = df[df_excel_columns]
    
    # 컬럼명 한글로 변경
    column_mapping_korean = {
        'id': '번호', 'received_date': '접수일', 'received_time': '접수시간', 
        'customer_name': '고객명', 'phone': '연락처', 'address': '주소', 
        'product': '제품', 'options': '옵션', 'notes': '비고', 
        'payment_amount': '결제금액', 'measurement_date': '실측일', 
        'measurement_time': '실측시간', 'completion_date': '설치완료일', 
        'manager_name': '담당자', 'status': '상태'
    }
    df_excel.rename(columns=column_mapping_korean, inplace=True)
    
    # 엑셀 파일 생성
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    excel_filename = f"furniture_orders_{timestamp}.xlsx"
    excel_path = os.path.join(app.config['UPLOAD_FOLDER'], excel_filename)
    
    df_excel.to_excel(excel_path, index=False, engine='openpyxl')
    
    # 로그 기록
    log_access(f"엑셀 다운로드: {excel_filename}", session.get('user_id'))
    
    # 파일을 사용자에게 전송 (다운로드 후 서버에서 파일 삭제 옵션 추가 가능)
    return send_file(excel_path, as_attachment=True)

@app.route('/calendar')
@login_required
def calendar():
    return render_template('calendar.html')

@app.route('/wdplanner')
@login_required
def wdplanner():
    """WDPlanner - 붙박이장 3D 설계 프로그램 (FOMS 레이아웃 포함)"""
    # FOMS 레이아웃을 유지하면서 iframe으로 WDPlanner를 표시
    return render_template('wdplanner.html')

# WDPlanner 정적 파일 서빙 - Flask의 기본 static 폴더 활용
@app.route('/wdplanner/app/<path:filename>')
@login_required
def wdplanner_static(filename):
    """WDPlanner 정적 파일 서빙 (JS, CSS, assets 등)"""
    # static/wdplanner/ 경로에서 파일 서빙
    return send_from_directory('static/wdplanner', filename)

@app.route('/wdplanner/app')
@login_required
def wdplanner_app():
    """WDPlanner 앱 자체 (iframe 내부에서 로드)"""
    wdplanner_index = os.path.join('static', 'wdplanner', 'index.html')
    if os.path.exists(wdplanner_index):
        # index.html 파일을 직접 반환 (Content-Type 자동 설정)
        return send_from_directory('static/wdplanner', 'index.html')
    else:
        return render_template('wdplanner_setup.html')

@app.route('/map_view')
@login_required
def map_view():
    """지도 보기 페이지"""
    return render_template('map_view.html')


@app.route('/api/calculate_route')
@login_required
def api_calculate_route():
    """두 지점 간 경로 계산 API"""
    try:
        start_lat = request.args.get('start_lat', type=float)
        start_lng = request.args.get('start_lng', type=float)
        end_lat = request.args.get('end_lat', type=float)
        end_lng = request.args.get('end_lng', type=float)
        
        if not all([start_lat, start_lng, end_lat, end_lng]):
            return jsonify({
                'success': False,
                'error': '출발지와 도착지 좌표가 모두 필요합니다.'
            }), 400
        
        # 주소 변환기 초기화
        address_converter = FOMSAddressConverter()
        
        # 경로 계산
        route_result = address_converter.calculate_route(
            start_lat, start_lng, end_lat, end_lng
        )
        
        if route_result['status'] == 'success':
            return jsonify({
                'success': True,
                'data': route_result
            })
        else:
            return jsonify({
                'success': False,
                'error': route_result['message']
            }), 400
            
    except Exception as e:
        return jsonify({
            'success': False,
            'error': f'경로 계산 중 오류: {str(e)}'
        }), 500

@app.route('/api/address_suggestions')
@login_required
def api_address_suggestions():
    """주소 교정 제안 API"""
    try:
        address = request.args.get('address')
        if not address:
            return jsonify({'success': False, 'error': '주소가 필요합니다.'}), 400
        
        converter = FOMSAddressConverter()
        suggestions = converter.get_address_suggestions(address)
        
        return jsonify({
            'success': True,
            'suggestions': suggestions
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/api/add_address_learning', methods=['POST'])
@login_required
def api_add_address_learning():
    """주소 학습 데이터 추가 API"""
    try:
        data = request.get_json()
        
        original_address = data.get('original_address')
        corrected_address = data.get('corrected_address')
        latitude = data.get('latitude')
        longitude = data.get('longitude')
        
        if not all([original_address, corrected_address, latitude, longitude]):
            return jsonify({
                'success': False, 
                'error': '모든 필드가 필요합니다.'
            }), 400
        
        converter = FOMSAddressConverter()
        converter.add_learning_data(original_address, corrected_address, latitude, longitude)
        
        return jsonify({
            'success': True,
            'message': '학습 데이터가 추가되었습니다.'
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/api/validate_address')
@login_required
def api_validate_address():
    """주소 유효성 검증 API"""
    try:
        address = request.args.get('address')
        if not address:
            return jsonify({'success': False, 'error': '주소가 필요합니다.'}), 400
        
        converter = FOMSAddressConverter()
        validation = converter.validate_address(address)
        
        return jsonify({
            'success': True,
            'validation': validation
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/api/orders')
@login_required
def api_orders():
    start_date = request.args.get('start')
    end_date = request.args.get('end')
    status_filter = request.args.get('status', None)
    limit_raw = request.args.get('limit', '2000')
    
    db = get_db()
    
    # Base query for orders
    query = db.query(Order).filter(Order.status != 'DELETED')
    
    # Add status filter if provided
    if status_filter and status_filter in STATUS:
        # 접수 탭에서는 RECEIVED와 ON_HOLD 상태를 모두 표시
        if status_filter == 'RECEIVED':
            query = query.filter(Order.status.in_(['RECEIVED', 'ON_HOLD']))
        else:
            query = query.filter(Order.status == status_filter)
    
    # Add date range filter if provided
    # ERP Beta 주문의 경우 measurement_date도 고려해야 하므로 OR 조건 사용
    if start_date and end_date:
        from sqlalchemy import or_
        # Handle date and datetime format properly
        if 'T' in start_date:  # ISO format with time (YYYY-MM-DDTHH:MM:SS)
            start_date_only = start_date.split('T')[0]
            end_date_only = end_date.split('T')[0]
            # received_date 또는 measurement_date가 범위 내에 있는 주문 포함
            query = query.filter(
                or_(
                    Order.received_date.between(start_date_only, end_date_only),
                    Order.measurement_date.between(start_date_only, end_date_only)
                )
            )
        else:  # Date only format (YYYY-MM-DD)
            # received_date 또는 measurement_date가 범위 내에 있는 주문 포함
            query = query.filter(
                or_(
                    Order.received_date.between(start_date, end_date),
                    Order.measurement_date.between(start_date, end_date)
                )
            )
    
    # 동시 사용자 증가 시 과도한 메모리/응답 지연 방지를 위한 상한
    try:
        limit = int(limit_raw)
    except (TypeError, ValueError):
        limit = 2000
    limit = max(100, min(limit, 5000))

    orders = query.order_by(Order.id.desc()).limit(limit).all()
    
    # Map status to colors
    status_colors = {
        'RECEIVED': '#3788d8',   # Blue
        'MEASURED': '#f39c12',   # Orange
        'SCHEDULED': '#e74c3c',  # Red
        'SHIPPED_PENDING': '#ff6b35', # Bright Orange
        'COMPLETED': '#2ecc71',  # Green
        'AS_RECEIVED': '#9b59b6', # Purple
        'AS_COMPLETED': '#1abc9c'  # Teal
    }
    
    events = []
    for order in orders:
        # ERP Beta 주문인 경우 structured_data에서 정보 추출
        customer_name = order.customer_name
        phone = order.phone
        address = order.address
        product = order.product
        measurement_date = order.measurement_date
        measurement_time = order.measurement_time
        scheduled_date = order.scheduled_date
        
        if order.is_erp_beta and order.structured_data:
            sd = order.structured_data
            
            # 고객명: structured_data.parties.customer.name
            erp_customer_name = ((sd.get('parties') or {}).get('customer') or {}).get('name')
            if erp_customer_name:
                customer_name = erp_customer_name
            
            # 연락처: structured_data.parties.customer.phone
            erp_phone = ((sd.get('parties') or {}).get('customer') or {}).get('phone')
            if erp_phone:
                phone = erp_phone
            
            # 주소: structured_data.site.address_full 또는 address_main
            erp_address = ((sd.get('site') or {}).get('address_full') or 
                          (sd.get('site') or {}).get('address_main'))
            if erp_address:
                address = erp_address
            
            # 제품: structured_data.items에서 첫 번째 제품명 또는 items의 요약
            items = sd.get('items') or []
            if items and len(items) > 0:
                first_item = items[0]
                product_name = first_item.get('product_name') or first_item.get('name')
                if product_name:
                    if len(items) > 1:
                        product = f"{product_name} 외 {len(items) - 1}개"
                    else:
                        product = product_name
            
            # 실측일: structured_data.schedule.measurement.date
            erp_measurement_date = (((sd.get('schedule') or {}).get('measurement') or {}).get('date'))
            if erp_measurement_date:
                measurement_date = erp_measurement_date
            
            # 실측시간: structured_data.schedule.measurement.time
            erp_measurement_time = (((sd.get('schedule') or {}).get('measurement') or {}).get('time'))
            if erp_measurement_time:
                measurement_time = erp_measurement_time
            
            # 시공일: structured_data.schedule.construction.date
            erp_scheduled_date = (((sd.get('schedule') or {}).get('construction') or {}).get('date'))
            if erp_scheduled_date:
                scheduled_date = erp_scheduled_date
        
        # ERP Beta 주문의 경우 measurement_date가 있으면 우선 사용 (상태와 관계없이)
        # 일반 주문의 경우 상태별 날짜 필드 매핑
        if order.is_erp_beta and measurement_date:
            # ERP Beta 주문이고 measurement_date가 있으면 실측일 기준으로 표시
            start_date = measurement_date
        else:
            # 상태별 날짜 필드 매핑
            status_date_map = {
                'RECEIVED': order.received_date,  
                'MEASURED': measurement_date,  # ERP Beta 주문의 경우 structured_data에서 추출한 날짜 사용
                'SCHEDULED': scheduled_date,  # ERP Beta 주문의 경우 structured_data에서 추출한 날짜 사용
                'SHIPPED_PENDING': scheduled_date,  # 상차 예정도 스케줄된 날짜 사용
                'COMPLETED': order.completion_date,
                'AS_RECEIVED': order.as_received_date,  # AS 접수일 필드 사용
                'AS_COMPLETED': order.as_completed_date  # AS 완료일 필드 사용
            }
            
            # 상태에 맞는 날짜 선택, 없는 경우 기본값으로 received_date 사용
            start_date = status_date_map.get(order.status)
        
        # 날짜 필드가 없는 경우 이벤트를 생성하지 않음
        if not start_date:
            continue
            
        # 시간 필드 매핑 (ERP Beta 주문의 경우 추출한 시간 사용)
        status_time_map = {
            'RECEIVED': order.received_time,
            'MEASURED': measurement_time,  # ERP Beta 주문의 경우 structured_data에서 추출한 시간 사용
            'SCHEDULED': None,  # 설치 예정은 일반적으로 시간 없음
            'SHIPPED_PENDING': None,  # 상차 예정은 일반적으로 시간 없음
            'COMPLETED': None,  # 완료는 일반적으로 시간 없음
            'AS_RECEIVED': None,  # AS는 일반적으로 시간 없음
            'AS_COMPLETED': None  # AS 완료는 일반적으로 시간 없음
        }
        
        time_str = status_time_map.get(order.status)
        
        # '실측' 상태이고 measurement_time이 '종일', '오전', '오후'인 경우 allDay를 true로 설정
        if order.status == 'MEASURED' and measurement_time in ['종일', '오전', '오후']:
            start_datetime = start_date # 날짜만 사용
            all_day = True
        elif time_str: # 기존 시간 처리 로직
            start_datetime = f"{start_date}T{time_str}:00"
            all_day = False
        else: # 시간이 없는 다른 경우 (기존 allDay=True 로직 유지)
            start_datetime = start_date
            all_day = True
            
        color = status_colors.get(order.status, '#3788d8')
        title = f"{customer_name} | {phone} | {product}"
        
        events.append({
            'id': order.id,
            'title': title,
            'start': start_datetime,
            'allDay': all_day,
            'backgroundColor': color,
            'borderColor': color,
            'extendedProps': {
                'customer_name': customer_name,
                'phone': phone,
                'address': address,
                'product': product,
                'options': order.options,
                'notes': order.notes,
                'status': order.status,
                'received_date': order.received_date,
                'received_time': order.received_time,
                'measurement_date': measurement_date,
                'measurement_time': measurement_time,
                'completion_date': order.completion_date,
                'scheduled_date': scheduled_date,
                'as_received_date': order.as_received_date,
                'as_completed_date': order.as_completed_date,
                'manager_name': order.manager_name
            }
        })
    
    return jsonify(events)

# Admin routes for menu management
@app.route('/change-logs')
@login_required
def change_logs():
    """변경 로그 페이지 - 모든 사용자가 본인의 변경 이력 확인 가능"""
    return render_template('change_logs.html')

@app.route('/admin')
@login_required
@role_required(['ADMIN'])
def admin():
    return render_template('admin.html')

@app.route('/admin/update_menu', methods=['POST'])
@login_required
@role_required(['ADMIN'])
def update_menu():
    try:
        menu_config = request.form.get('menu_config')
        if menu_config:
            # Save menu configuration to a file
            with open('menu_config.json', 'w', encoding='utf-8') as f:
                f.write(menu_config)
            
            # Log the action
            # log_access(f"메뉴 설정 업데이트", session.get('user_id')) # 로그 형식 수정 필요
            user_for_log = get_user_by_id(session['user_id'])
            user_name_for_log = user_for_log.name if user_for_log else "Unknown user"
            log_access(f"메뉴 설정 업데이트 - 담당자: {user_name_for_log} (ID: {session.get('user_id')})", session.get('user_id'))
            
            flash('메뉴 구성이 업데이트되었습니다.', 'success')
        else:
            flash('메뉴 구성을 입력해주세요.', 'error')
    except Exception as e:
        flash(f'메뉴 업데이트 중 오류가 발생했습니다: {str(e)}', 'error')
    
    return redirect(url_for('admin'))

def translate_dict_keys(d, key_map):
    if not isinstance(d, dict):
        return d
    new_dict = {}
    for k, v in d.items():
        translated_key = key_map.get(k, k)
        if isinstance(v, dict):
            new_dict[translated_key] = translate_dict_keys(v, key_map)
        elif isinstance(v, list):
            new_dict[translated_key] = [translate_dict_keys(item, key_map) for item in v]
        else:
            new_dict[translated_key] = v
    return new_dict

def format_value_for_log(value):
    if value is None:
        return "없음"
    if isinstance(value, str) and not value.strip(): # 빈 문자열
        return "없음"
    return str(value)

# Profile route for users to manage their own account
@app.route('/profile', methods=['GET', 'POST'])
@login_required
def profile():
    user_id = session.get('user_id')
    db = get_db()
    user = db.query(User).filter(User.id == user_id).first()
    
    if not user:
        session.clear()
        flash('사용자를 찾을 수 없습니다. 다시 로그인해주세요.', 'error')
        return redirect(url_for('auth.login'))
    
    if request.method == 'POST':
        current_password = request.form.get('current_password')
        new_password = request.form.get('new_password')
        confirm_password = request.form.get('confirm_password')
        name = request.form.get('name')
        
        # Validate name
        if not name:
            flash('이름을 입력해주세요.', 'error')
            return render_template('profile.html', user=user)
        
        try:
            # Update name
            user.name = name
            db.commit()
            
            # Handle password change if provided
            if current_password and new_password and confirm_password:
                # Verify current password
                if not check_password_hash(user.password, current_password):
                    flash('현재 비밀번호가 일치하지 않습니다.', 'error')
                    return render_template('profile.html', user=user)
                
                # Check password match
                if new_password != confirm_password:
                    flash('새 비밀번호가 일치하지 않습니다.', 'error')
                    return render_template('profile.html', user=user)
                
                # Check password strength
                if not is_password_strong(new_password):
                    flash('비밀번호는 4자리 이상이어야 합니다.', 'error') # 메시지 수정
                    return render_template('profile.html', user=user)
                
                # Update password
                user.password = generate_password_hash(new_password)
                db.commit()
                
                # Log password change
                log_access("비밀번호 변경 완료", user_id)
                
                flash('비밀번호가 성공적으로 변경되었습니다.', 'success')
            
            flash('프로필이 업데이트되었습니다.', 'success')
            return redirect(url_for('profile'))
                
        except Exception as e:
            db.rollback()
            flash(f'프로필 업데이트 중 오류가 발생했습니다: {str(e)}', 'error')
            return render_template('profile.html', user=user)
    
    return render_template('profile.html', user=user)

def load_menu_config():
    try:
        if os.path.exists('menu_config.json'):
            with open('menu_config.json', 'r', encoding='utf-8') as f:
                return json.load(f)
    except:
        pass
    
    # Default menu configuration
    return {
        'main_menu': [
            {'id': 'calendar', 'name': '캘린더', 'url': '/calendar'},
            {'id': 'order_list', 'name': '전체 주문', 'url': '/'},
            {'id': 'received', 'name': '접수', 'url': '/?status=RECEIVED'},
            {'id': 'measured', 'name': '실측', 'url': '/?status=MEASURED'},
            {'id': 'metro_orders', 'name': '수도권 주문', 'url': '/?region=metro'},
            {'id': 'regional_orders', 'name': '지방 주문', 'url': '/?region=regional'},
            {'id': 'storage_dashboard', 'name': '수납장 대시보드', 'url': '/storage_dashboard'},
            {'id': 'regional_dashboard', 'name': '지방 주문 대시보드', 'url': '/regional_dashboard'},
            {'id': 'self_measurement_dashboard', 'name': '자가실측 대시보드', 'url': '/self_measurement_dashboard'},
            {'id': 'metropolitan_dashboard', 'name': '수도권 주문 대시보드', 'url': '/metropolitan_dashboard'},
            {'id': 'trash', 'name': '휴지통', 'url': '/trash'},
            {'id': 'chat', 'name': '채팅', 'url': '/chat'}
        ],
        'admin_menu': [
            {'id': 'user_management', 'name': '사용자 관리', 'url': '/admin/users'},
            {'id': 'security_logs', 'name': '보안 로그', 'url': '/admin/security-logs'}
        ]
    }

@app.context_processor
def inject_menu():
    menu_config = load_menu_config()

    # ERP Beta: 실측/AS/출고는 ERP 대시보드 서브 내비로 이동 (메인 메뉴에는 추가하지 않음)

    return dict(menu=menu_config)

# Jinja 필터: 메시지 내 "주문 #<번호>"를 클릭 가능한 링크로 변환
@app.template_filter('order_link')
def order_link_filter(s):
    import re
    from flask import url_for
    def repl(m):
        oid = m.group(1)
        link = url_for('edit_order', order_id=oid)
        return Markup(f'<a href="{link}">주문 #{oid}</a>')
    return Markup(re.sub(r'주문 #(\d+)', repl, s))

# 보안 로그 목록 조회 라우트 추가 (관리자 전용)
@app.route('/security_logs')
@login_required
@role_required(['ADMIN'])
def security_logs():
    db = get_db()
    
    page = request.args.get('page', 1, type=int)
    per_page = 50  # 페이지당 로그 수
    
    search_query = request.args.get('search', '')
    
    query = db.query(SecurityLog).order_by(SecurityLog.timestamp.desc())
    
    if search_query:
        # 사용자 이름 또는 메시지 내용으로 검색
        query = query.join(User, User.id == SecurityLog.user_id, isouter=True).filter(
            or_(
                User.name.ilike(f'%{search_query}%'),
                SecurityLog.message.ilike(f'%{search_query}%')
            )
        )
        
    total_logs = query.count()
    logs = query.offset((page - 1) * per_page).limit(per_page).all()
    
    total_pages = (total_logs + per_page - 1) // per_page
    
    return render_template('security_logs.html', 
                           logs=logs, 
                           page=page, 
                           total_pages=total_pages, 
                           search_query=search_query,
                           total_logs=total_logs)

@app.route('/regional_dashboard')
@login_required
def regional_dashboard():
    """지방 주문 관리 대시보드"""
    db = get_db()
    search_query = request.args.get('search_query', '').strip()
    
    # 기본 쿼리
    base_query = db.query(Order).filter(
        Order.is_regional == True,
        Order.status != 'DELETED'
    )
    
    # 검색 기능 적용
    if search_query:
        search_term = f"%{search_query}%"
        # ID 검색을 위한 숫자 체크
        id_conditions = []
        try:
            # 검색어가 숫자인 경우 ID로 검색
            search_id = int(search_query)
            id_conditions.append(Order.id == search_id)
        except ValueError:
            # 숫자가 아닌 경우 ID를 문자열로 캐스팅해서 검색
            id_conditions.append(func.cast(Order.id, String).ilike(search_term))
        
        base_query = base_query.filter(
            or_(
                Order.customer_name.ilike(search_term),
                Order.phone.ilike(search_term),
                Order.address.ilike(search_term),
                Order.product.ilike(search_term),
                Order.regional_memo.ilike(search_term),
                Order.notes.ilike(search_term),
                *id_conditions
            )
        )
    
    # 모든 지방 주문 가져오기
    all_regional_orders = base_query.order_by(Order.id.desc()).all()
    
    # ERP Beta 주문 표시 정보 반영
    apply_erp_display_fields_to_orders(all_regional_orders)
    
    # 오늘 날짜
    today = date.today()
    
        # 완료된 주문 분류
    completed_orders = [
        order for order in all_regional_orders
        if order.status == 'COMPLETED'
    ]

    # 설치예정인 주문 분류
    scheduled_orders = [
        order for order in all_regional_orders
        if order.status == 'SCHEDULED'
    ]

    # 보류 상태 주문 분류
    hold_orders = [
        order for order in all_regional_orders
        if order.status == 'ON_HOLD'
    ]

    # 상차 예정 알림: 실측 완료 + 상차일 지정 + 미완료 상태 + 상차일이 오늘 이후 + 보류 상태 제외
    shipping_alerts = []
    for order in all_regional_orders:
        if (getattr(order, 'measurement_completed', False) and 
            order.shipping_scheduled_date and 
            order.shipping_scheduled_date.strip() and
            order.status not in ['COMPLETED', 'ON_HOLD']):  # 완료된 주문과 보류 상태 제외
            try:
                shipping_date = datetime.datetime.strptime(order.shipping_scheduled_date, '%Y-%m-%d').date()
                # 오늘 이후의 상차일만 포함 (지난 상차일은 제외)
                if shipping_date >= today:
                    shipping_alerts.append(order)
            except (ValueError, TypeError):
                # 날짜 형식이 잘못된 경우 무시
                pass

    # 상차완료: 상차일이 지났지만 완료 처리되지 않은 주문들 + 보류 상태 제외
    shipping_completed_orders = []
    for order in all_regional_orders:
        if (order.shipping_scheduled_date and 
            order.shipping_scheduled_date.strip() and
            order.status not in ['COMPLETED', 'ON_HOLD']):  # 완료된 주문과 보류 상태 제외
            try:
                shipping_date = datetime.datetime.strptime(order.shipping_scheduled_date, '%Y-%m-%d').date()
                # 상차일이 오늘보다 이전인 경우 (지난 상차일)
                if shipping_date < today:
                    shipping_completed_orders.append(order)
            except (ValueError, TypeError):
                # 날짜 형식이 잘못된 경우 무시
                pass

    # 진행 중인 주문: 실측 미완료 + 완료되지 않은 주문 + 상차 예정 알림에 없는 주문 + 상차완료에 없는 주문 + 보류 상태 및 설치예정 상태 제외
    shipping_alert_order_ids = {order.id for order in shipping_alerts}
    shipping_completed_order_ids = {order.id for order in shipping_completed_orders}
    pending_orders = [
        order for order in all_regional_orders
        if (order.status not in ['COMPLETED', 'ON_HOLD', 'SCHEDULED'] and 
            order.id not in shipping_alert_order_ids and
            order.id not in shipping_completed_order_ids and
            (not getattr(order, 'measurement_completed', False) or 
             not order.shipping_scheduled_date or 
             not order.shipping_scheduled_date.strip()))
    ]

    # 상차일 기준으로 정렬 (가까운 날짜부터)
    shipping_alerts.sort(key=lambda x: datetime.datetime.strptime(x.shipping_scheduled_date, '%Y-%m-%d').date())
    shipping_completed_orders.sort(key=lambda x: datetime.datetime.strptime(x.shipping_scheduled_date, '%Y-%m-%d').date())




    today_str = today.strftime('%Y-%m-%d')
    tomorrow_str = (today + timedelta(days=1)).strftime('%Y-%m-%d')
        
    return render_template('regional_dashboard.html', 
                           pending_orders=pending_orders, 
                           scheduled_orders=scheduled_orders,
                           completed_orders=completed_orders,
                           hold_orders=hold_orders,
                           shipping_alerts=shipping_alerts,
                           shipping_completed_orders=shipping_completed_orders,
                           STATUS=STATUS,
                           search_query=search_query,
                           today=today_str,
                           tomorrow=tomorrow_str)

@app.route('/metropolitan_dashboard')
@login_required
def metropolitan_dashboard():
    db = get_db()
    search_query = request.args.get('search_query', '').strip()

    def get_filtered_orders(query):
        if search_query:
            search_term = f"%{search_query}%"
            # ID 검색을 위한 숫자 체크
            id_conditions = []
            try:
                # 검색어가 숫자인 경우 ID로 검색
                search_id = int(search_query)
                id_conditions.append(Order.id == search_id)
            except ValueError:
                # 숫자가 아닌 경우 ID를 문자열로 캐스팅해서 검색
                id_conditions.append(func.cast(Order.id, String).ilike(search_term))
            
            return query.filter(
                or_(
                    Order.customer_name.ilike(search_term),
                    Order.phone.ilike(search_term),
                    Order.address.ilike(search_term),
                    Order.product.ilike(search_term),
                    Order.notes.ilike(search_term),
                    Order.manager_name.ilike(search_term),
                    *id_conditions
                )
            )
        return query

    base_query = db.query(Order).filter(Order.is_regional == False)

    # 쿼리에서 날짜 비교 시 func.date()를 사용하여 타입 일치
    # .all()을 호출하기 전에 필터링이 적용되도록 수정
    urgent_alerts_query = base_query.filter(
        Order.status.in_(['MEASURED']),
        Order.measurement_date != None,
        Order.measurement_date != '',
        func.date(Order.measurement_date) == date.today()
    )
    urgent_alerts = get_filtered_orders(urgent_alerts_query).order_by(Order.measurement_date.asc()).all()

    # 실측 후 미처리: 실측일이 도래했고, 설치일이 없는 경우 (당일 실측 건 제외)
    measurement_alerts_query = base_query.filter(
        Order.status.in_(['MEASURED']),
        Order.measurement_date != None,
        Order.measurement_date != '',
        func.date(Order.measurement_date) < date.today(),  # 당일 제외, 과거 실측일만
        or_(
            Order.scheduled_date == None,
            Order.scheduled_date == ''
        )
    )
    measurement_alerts = get_filtered_orders(measurement_alerts_query).order_by(Order.measurement_date.asc()).all()

    pre_measurement_alerts_query = base_query.filter(
        or_(
            # 실측일이 미래인 경우
            and_(
                Order.status.in_(['RECEIVED', 'MEASURED']),
                Order.measurement_date != None,
                Order.measurement_date != '',
                func.date(Order.measurement_date) > date.today()
            ),
            # 실측일이 없거나 상태가 RECEIVED인 경우 (실측 전 단계)
            and_(
                Order.status == 'RECEIVED',
                or_(
                    Order.measurement_date == None,
                    Order.measurement_date == ''
                )
            )
        )
    )
    pre_measurement_alerts = get_filtered_orders(pre_measurement_alerts_query).order_by(Order.measurement_date.asc()).all()

    installation_alerts_query = base_query.filter(
        Order.status.in_(['SCHEDULED', 'SHIPPED_PENDING']),
        # scheduled_date가 None이 아니고 빈 문자열이 아닌 경우에만 비교
        Order.scheduled_date != None,
        Order.scheduled_date != '',
        func.date(Order.scheduled_date) < date.today()
    )
    installation_alerts = get_filtered_orders(installation_alerts_query).order_by(Order.scheduled_date.asc()).all()

    alert_order_ids = {o.id for o in urgent_alerts + measurement_alerts + pre_measurement_alerts + installation_alerts}

    # AS 관련 주문들 (AS_RECEIVED만 포함)
    as_orders_query = db.query(Order).filter(
        Order.status == 'AS_RECEIVED',
        Order.is_regional == False
    )
    as_orders = get_filtered_orders(as_orders_query).order_by(Order.created_at.desc()).all()

    # 보류 상태 주문들 (ON_HOLD)
    hold_orders_query = db.query(Order).filter(
        Order.status == 'ON_HOLD',
        Order.is_regional == False
    )
    hold_orders = get_filtered_orders(hold_orders_query).order_by(Order.created_at.desc()).all()

    # 정상 진행 중인 주문들 (알림에 포함되지 않은 진행 중인 주문들, 보류 상태 제외)
    normal_orders_query = db.query(Order).filter(
        Order.status.notin_(['COMPLETED', 'DELETED', 'AS_RECEIVED', 'AS_COMPLETED', 'ON_HOLD']),
        ~Order.id.in_(alert_order_ids),
        Order.is_regional == False
    )
    normal_orders = get_filtered_orders(normal_orders_query).order_by(Order.created_at.desc()).limit(20).all()

    # 완료된 주문들 (COMPLETED와 AS_COMPLETED 포함)
    completed_orders_query = db.query(Order).filter(
        Order.status.in_(['COMPLETED', 'AS_COMPLETED']),
        Order.is_regional == False
    )
    completed_orders = get_filtered_orders(completed_orders_query).order_by(Order.completion_date.desc()).limit(50).all()
    
    # 모든 주문 리스트 수집
    all_metro_orders = urgent_alerts + measurement_alerts + pre_measurement_alerts + installation_alerts + as_orders + hold_orders + normal_orders + completed_orders
    
    # ERP Beta 주문 표시 정보 반영
    apply_erp_display_fields_to_orders(all_metro_orders)
        
    return render_template('metropolitan_dashboard.html', 
                           urgent_alerts=urgent_alerts,
                           measurement_alerts=measurement_alerts,
                           pre_measurement_alerts=pre_measurement_alerts,
                           installation_alerts=installation_alerts,
                           as_orders=as_orders,
                           hold_orders=hold_orders,
                           normal_orders=normal_orders,
                           completed_orders=completed_orders,
                           STATUS=STATUS,
                           search_query=search_query)

# 백업 시스템 라우트들
@app.route('/api/simple_backup', methods=['POST'])
@login_required
@role_required(['ADMIN'])
def execute_simple_backup():
    """간단한 2단계 백업 실행"""
    try:
        backup_system = SimpleBackupSystem()
        results = backup_system.execute_backup()
        
        # 결과 요약
        success_count = sum(1 for r in results.values() if r["success"])
        success_rate = success_count * 50  # 2단계이므로 50%씩
        
        # 로그 기록
        log_access(f"백업 실행 - 성공률: {success_rate}%", session.get('user_id'), {
            "tier1_success": results["tier1"]["success"],
            "tier2_success": results["tier2"]["success"]
        })
        
        return jsonify({
            "success": True,
            "message": f"백업 완료! 성공률: {success_rate}%",
            "results": results,
            "success_count": success_count,
            "total_tiers": 2
        })
        
    except Exception as e:
        log_access(f"백업 실행 실패: {str(e)}", session.get('user_id'))
        return jsonify({
            "success": False,
            "message": f"백업 실행 중 오류가 발생했습니다: {str(e)}"
        }), 500

@app.route('/api/backup_status')
@login_required
@role_required(['ADMIN'])
def check_backup_status():
    """백업 상태 확인"""
    try:
        backup_system = SimpleBackupSystem()
        
        status = {
            "tier1": {
                "path": backup_system.tier1_path,
                "exists": os.path.exists(backup_system.tier1_path),
                "latest_backup": None
            },
            "tier2": {
                "path": backup_system.tier2_path,
                "exists": os.path.exists(backup_system.tier2_path),
                "latest_backup": None
            }
        }
        
        # 각 티어의 최신 백업 정보 조회
        for tier_name, tier_info in status.items():
            if tier_info["exists"]:
                try:
                    info_file = os.path.join(tier_info["path"], "backup_info.json")
                    if os.path.exists(info_file):
                        with open(info_file, 'r', encoding='utf-8') as f:
                            backup_info = json.load(f)
                            tier_info["latest_backup"] = backup_info
                except Exception as e:
                    tier_info["error"] = str(e)
        
        return jsonify({
            "success": True,
            "status": status
        })
        
    except Exception as e:
        return jsonify({
            "success": False,
            "message": f"백업 상태 확인 중 오류: {str(e)}"
        }), 500

@app.route('/self_measurement_dashboard')
@login_required
def self_measurement_dashboard():
    """자가실측 대시보드"""
    db = get_db()
    search_query = request.args.get('search_query', '').strip()
    
    # 기본 쿼리
    base_query = db.query(Order).filter(
        Order.is_self_measurement == True,
        Order.status != 'DELETED'
    )
    
    # 검색 기능 적용
    if search_query:
        search_term = f"%{search_query}%"
        # ID 검색을 위한 숫자 체크
        id_conditions = []
        try:
            # 검색어가 숫자인 경우 ID로 검색
            search_id = int(search_query)
            id_conditions.append(Order.id == search_id)
        except ValueError:
            # 숫자가 아닌 경우 ID를 문자열로 캐스팅해서 검색
            id_conditions.append(func.cast(Order.id, String).ilike(search_term))
        
        base_query = base_query.filter(
            or_(
                Order.customer_name.ilike(search_term),
                Order.phone.ilike(search_term),
                Order.address.ilike(search_term),
                Order.product.ilike(search_term),
                Order.notes.ilike(search_term),
                *id_conditions
            )
        )
    
    # 모든 자가실측 주문 가져오기
    all_self_measurement_orders = base_query.order_by(Order.id.desc()).all()
    
    # ERP Beta 주문 표시 정보 반영
    apply_erp_display_fields_to_orders(all_self_measurement_orders)
    
    # AS 접수된 주문 분류
    as_orders = [
        order for order in all_self_measurement_orders
        if order.status == 'AS_RECEIVED'
    ]

    # 완료된 주문 분류
    completed_orders = [
        order for order in all_self_measurement_orders
        if order.status in ['COMPLETED', 'AS_COMPLETED']
    ]
    
    # 설치예정인 주문 분류
    scheduled_orders = [
        order for order in all_self_measurement_orders
        if order.status == 'SCHEDULED'
    ]
    
    # 진행 중인 주문 분류 (완료/설치예정/AS접수 제외)
    pending_orders = [
        order for order in all_self_measurement_orders
        if order.status not in ['COMPLETED', 'AS_COMPLETED', 'SCHEDULED', 'AS_RECEIVED']
    ]
    
    return render_template('self_measurement_dashboard.html',
                           pending_orders=pending_orders,
                           scheduled_orders=scheduled_orders,
                           as_orders=as_orders,
                           completed_orders=completed_orders,
                           search_query=search_query,
                           STATUS=STATUS)

@app.route('/storage_dashboard')
@login_required
def storage_dashboard():
    """수납장 대시보드"""
    db = get_db()
    search_query = request.args.get('search_query', '').strip()

    base_query = db.query(Order).filter(
        Order.is_cabinet == True,
        Order.status != 'DELETED'
    )

    if search_query:
        search_term = f"%{search_query}%"
        id_conditions = []
        try:
            search_id = int(search_query)
            id_conditions.append(Order.id == search_id)
        except ValueError:
            id_conditions.append(func.cast(Order.id, String).ilike(search_term))

        base_query = base_query.filter(
            or_(
                Order.customer_name.ilike(search_term),
                Order.phone.ilike(search_term),
                Order.address.ilike(search_term),
                Order.product.ilike(search_term),
                Order.notes.ilike(search_term),
                *id_conditions
            )
        )

    all_cabinet_orders = base_query.order_by(Order.id.desc()).all()

    # 카테고리 분류: 접수(RECEIVED), 제작중(IN_PRODUCTION), 발송(SHIPPED)
    received_orders = [o for o in all_cabinet_orders if (o.cabinet_status or 'RECEIVED') == 'RECEIVED']
    in_production_orders = [o for o in all_cabinet_orders if o.cabinet_status == 'IN_PRODUCTION']
    shipped_orders = [o for o in all_cabinet_orders if o.cabinet_status == 'SHIPPED']

    return render_template('storage_dashboard.html',
                           received_orders=received_orders,
                           in_production_orders=in_production_orders,
                           shipped_orders=shipped_orders,
                           search_query=search_query,
                           CABINET_STATUS=CABINET_STATUS,
                           STATUS=STATUS)

@app.route('/api/storage_dashboard/export_excel')
@login_required
@role_required(['ADMIN', 'MANAGER', 'STAFF'])
def export_storage_dashboard_excel():
    """수납장 대시보드 제작중 주문 엑셀 내보내기"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    
    db = get_db()
    search_query = request.args.get('search_query', '').strip()
    
    # 제작중 주문만 조회
    base_query = db.query(Order).filter(
        Order.is_cabinet == True,
        Order.status != 'DELETED',
        Order.cabinet_status == 'IN_PRODUCTION'
    )
    
    # 검색어 필터 적용
    if search_query:
        search_term = f"%{search_query}%"
        id_conditions = []
        try:
            search_id = int(search_query)
            id_conditions.append(Order.id == search_id)
        except ValueError:
            id_conditions.append(func.cast(Order.id, String).ilike(search_term))
        
        base_query = base_query.filter(
            or_(
                Order.customer_name.ilike(search_term),
                Order.phone.ilike(search_term),
                Order.address.ilike(search_term),
                Order.product.ilike(search_term),
                Order.regional_memo.ilike(search_term),
                *id_conditions
            )
        )
    
    orders = base_query.order_by(Order.id.desc()).all()
    
    if not orders:
        flash('다운로드할 데이터가 없습니다.', 'warning')
        return redirect(url_for('storage_dashboard'))
    
    # 엑셀 워크북 생성
    wb = Workbook()
    ws = wb.active
    ws.title = "제작중 주문"
    
    # 헤더 정의
    headers = ['번호', '메모', '고객명', '전화번호', '주소', '제품', '배송비', '상태', '설치 예정일']
    
    # 스타일 정의
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")  # 파란색 배경
    header_font = Font(bold=True, color="FFFFFF")  # 흰색 볼드
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_alignment = Alignment(horizontal='center', vertical='center')
    
    # 헤더 작성
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center_alignment
    
    # 데이터 작성
    for row_idx, order in enumerate(orders, start=2):
        # 번호
        ws.cell(row=row_idx, column=1, value=order.id).border = border
        
        # 메모
        ws.cell(row=row_idx, column=2, value=order.regional_memo or '').border = border
        
        # 고객명
        ws.cell(row=row_idx, column=3, value=order.customer_name).border = border
        
        # 전화번호
        ws.cell(row=row_idx, column=4, value=order.phone).border = border
        
        # 주소
        ws.cell(row=row_idx, column=5, value=order.address).border = border
        
        # 제품
        ws.cell(row=row_idx, column=6, value=order.product).border = border
        
        # 배송비 (콤마 포맷팅)
        shipping_fee = order.shipping_fee or 0
        ws.cell(row=row_idx, column=7, value=f"{shipping_fee:,}").border = border
        ws.cell(row=row_idx, column=7).alignment = Alignment(horizontal='right', vertical='center')
        
        # 상태 (한글 변환)
        status_korean = CABINET_STATUS.get(order.cabinet_status, order.cabinet_status or '')
        ws.cell(row=row_idx, column=8, value=status_korean).border = border
        ws.cell(row=row_idx, column=8).alignment = center_alignment
        
        # 설치 예정일
        ws.cell(row=row_idx, column=9, value=order.scheduled_date or '').border = border
    
    # 컬럼 너비 자동 조정
    column_widths = [10, 25, 15, 15, 35, 30, 12, 12, 15]
    for col_idx, width in enumerate(column_widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width
    
    # 파일 저장
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    excel_filename = f"storage_in_production_{timestamp}.xlsx"
    excel_path = os.path.join(app.config['UPLOAD_FOLDER'], excel_filename)
    
    wb.save(excel_path)
    
    # 로그 기록
    log_access(f"수납장 대시보드 제작중 엑셀 다운로드: {excel_filename} ({len(orders)}건)", session.get('user_id'))
    
    # 파일 다운로드
    return send_file(excel_path, as_attachment=True, download_name=excel_filename)

# ==================== WDCalculator (가구 견적 계산기) ====================

# JSON 파일 경로
WD_CALCULATOR_DATA_PATH = os.path.join(os.path.dirname(__file__), 'data', 'products.json')
ERP_SHIPMENT_SETTINGS_PATH = os.path.join(os.path.dirname(__file__), 'data', 'erp_shipment_settings.json')
WD_ADDITIONAL_OPTIONS_PATH = os.path.join(os.path.dirname(__file__), 'data', 'additional_options.json')
WD_NOTES_CATEGORIES_PATH = os.path.join(os.path.dirname(__file__), 'data', 'notes_categories.json')
DEFAULT_ERP_WORKER_CAPACITY = 10

def clean_categories_data(categories):
    """카테고리 데이터에서 JSON 직렬화 불가능한 값 제거 및 id 자동 생성"""
    if not categories:
        return []
    
    cleaned = []
    base_option_id = 1000  # id가 null인 옵션에 부여할 시작 ID
    
    for cat_idx, category in enumerate(categories):
        if category is None:
            continue
        
        cleaned_category = {
            'id': category.get('id') if category.get('id') is not None else None,
            'name': category.get('name') or '',
            'options': []
        }
        
        # 옵션 정리 및 id 자동 생성
        options = category.get('options')
        if options and isinstance(options, list) and len(options) > 0:
            # 먼저 기존 옵션들 중 유효한 id 찾기
            existing_ids = [o.get('id') for o in options if o and isinstance(o, dict) and o.get('id') is not None]
            max_existing_id = max(existing_ids + [0]) if existing_ids else 0
            next_id = max(max_existing_id + 1, base_option_id + (cat_idx * 100))
            
            for opt_idx, option in enumerate(options):
                if option is None or not isinstance(option, dict):
                    print(f"[DEBUG] clean_categories_data: 유효하지 않은 옵션 건너뜀 (인덱스 {opt_idx})")
                    continue
                
                # id가 null이면 자동으로 생성
                option_id = option.get('id')
                if option_id is None:
                    option_id = next_id
                    next_id += 1
                    print(f"[DEBUG] clean_categories_data: 옵션 '{option.get('name')}'에 id {option_id} 자동 생성")
                
                cleaned_option = {
                    'id': option_id,
                    'name': str(option.get('name') or '').strip(),
                    'price': float(option.get('price', 0)) if option.get('price') is not None else 0.0
                }
                cleaned_category['options'].append(cleaned_option)
        elif options is not None and not isinstance(options, list):
            print(f"[DEBUG] clean_categories_data: 카테고리 {category.get('id')}의 options가 리스트가 아님: {type(options)}")
        
        cleaned.append(cleaned_category)
    
    return cleaned

def load_additional_option_categories():
    """추가 옵션 카테고리 데이터를 JSON 파일에서 로드"""
    try:
        if os.path.exists(WD_ADDITIONAL_OPTIONS_PATH):
            # 먼저 UTF-8로 시도
            try:
                with open(WD_ADDITIONAL_OPTIONS_PATH, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                    categories = data.get('categories', [])
                    return clean_categories_data(categories)
            except UnicodeDecodeError:
                # 실패 시 CP949로 시도 (Windows 호환)
                print("UTF-8 decoding failed for additional options, trying CP949...")
                with open(WD_ADDITIONAL_OPTIONS_PATH, 'r', encoding='cp949') as f:
                    data = json.load(f)
                    categories = data.get('categories', [])
                    return clean_categories_data(categories)
        return []
    except Exception as e:
        print(f"Error loading additional option categories: {e}")
        return []

def save_additional_option_categories(categories):
    """추가 옵션 카테고리 데이터를 JSON 파일에 저장"""
    try:
        os.makedirs(os.path.dirname(WD_ADDITIONAL_OPTIONS_PATH), exist_ok=True)
        data = {'categories': categories}
        # 저장할 때는 UTF-8 (ensure_ascii=False) 또는 ASCII (ensure_ascii=True)로 통일
        # 여기서는 호환성을 위해 ensure_ascii=True로 변경하여 인코딩 문제 원천 차단
        with open(WD_ADDITIONAL_OPTIONS_PATH, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=True, indent=2)
        return True
    except Exception as e:
        print(f"Error saving additional option categories: {e}")
        return False

def load_notes_categories():
    """비고 카테고리 데이터를 JSON 파일에서 로드"""
    try:
        print(f"[DEBUG] load_notes_categories 호출됨")
        print(f"[DEBUG] 파일 경로: {WD_NOTES_CATEGORIES_PATH}")
        print(f"[DEBUG] 파일 존재 여부: {os.path.exists(WD_NOTES_CATEGORIES_PATH)}")
        
        if os.path.exists(WD_NOTES_CATEGORIES_PATH):
            try:
                with open(WD_NOTES_CATEGORIES_PATH, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                    categories = data.get('categories', [])
                    print(f"[DEBUG] UTF-8로 로드 성공 - 카테고리 수: {len(categories)}")
                    cleaned = clean_categories_data(categories)
                    print(f"[DEBUG] 정리 후 카테고리 수: {len(cleaned)}")
                    return cleaned
            except UnicodeDecodeError:
                print("[DEBUG] UTF-8 decoding failed for notes categories, trying CP949...")
                with open(WD_NOTES_CATEGORIES_PATH, 'r', encoding='cp949') as f:
                    data = json.load(f)
                    categories = data.get('categories', [])
                    print(f"[DEBUG] CP949로 로드 성공 - 카테고리 수: {len(categories)}")
                    cleaned = clean_categories_data(categories)
                    print(f"[DEBUG] 정리 후 카테고리 수: {len(cleaned)}")
                    return cleaned
        else:
            print(f"[DEBUG] 파일이 존재하지 않음 - 빈 배열 반환")
        return []
    except Exception as e:
        print(f"[ERROR] Error loading notes categories: {e}")
        import traceback
        print(traceback.format_exc())
        return []

def save_notes_categories(categories):
    """비고 카테고리 데이터를 JSON 파일에 저장"""
    try:
        print(f"[DEBUG] save_notes_categories 호출됨")
        print(f"[DEBUG] 저장할 카테고리 수: {len(categories) if categories else 0}")
        print(f"[DEBUG] 저장할 카테고리 데이터: {json.dumps(categories, ensure_ascii=False, indent=2)}")
        
        os.makedirs(os.path.dirname(WD_NOTES_CATEGORIES_PATH), exist_ok=True)
        data = {'categories': categories}
        with open(WD_NOTES_CATEGORIES_PATH, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=True, indent=2)
        
        print(f"[DEBUG] 파일 저장 완료: {WD_NOTES_CATEGORIES_PATH}")
        
        # 저장 후 확인
        if os.path.exists(WD_NOTES_CATEGORIES_PATH):
            with open(WD_NOTES_CATEGORIES_PATH, 'r', encoding='utf-8') as f:
                saved_data = json.load(f)
                print(f"[DEBUG] 저장된 파일 확인 - 카테고리 수: {len(saved_data.get('categories', []))}")
        
        return True
    except Exception as e:
        print(f"[ERROR] Error saving notes categories: {e}")
        import traceback
        print(traceback.format_exc())
        return False

def load_products():
    """제품 데이터를 JSON 파일에서 로드"""
    try:
        if os.path.exists(WD_CALCULATOR_DATA_PATH):
            # 먼저 UTF-8로 시도
            try:
                with open(WD_CALCULATOR_DATA_PATH, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                    return data.get('products', [])
            except UnicodeDecodeError:
                # 실패 시 CP949로 시도
                print("UTF-8 decoding failed for products, trying CP949...")
                with open(WD_CALCULATOR_DATA_PATH, 'r', encoding='cp949') as f:
                    data = json.load(f)
                    return data.get('products', [])
        return []
    except Exception as e:
        print(f"Error loading products: {e}")
        return []

def save_products(products):
    """제품 데이터를 JSON 파일에 저장"""
    try:
        os.makedirs(os.path.dirname(WD_CALCULATOR_DATA_PATH), exist_ok=True)
        data = {'products': products}
        # 저장할 때는 인코딩 문제 방지를 위해 ensure_ascii=True 사용
        with open(WD_CALCULATOR_DATA_PATH, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=True, indent=2)
        return True
    except Exception as e:
        print(f"Error saving products: {e}")
        return False





def calculate_estimate(product, width_mm, additional_options=None):
    """견적 계산 함수"""
    if not product:
        return 0
    
    base_price = 0
    
    if product['pricing_type'] == '1m':
        # 1m 옵션: (가로넓이 / 1000) * 1m 비용
        meters = width_mm / 1000
        base_price = meters * product.get('price_1m', 0)
    elif product['pricing_type'] == '30cm':
        # 30cm 옵션: (가로넓이 / 300) * 30cm 비용 + (나머지 / 10) * 1cm 비용
        # 예: 3100mm = 10 * 300mm + 100mm → (10 * 30cm 설정가) + (10 * 1cm 설정가)
        # 예: 3300mm = 11 * 300mm + 0mm → (11 * 30cm 설정가)
        units_30cm = width_mm // 300
        remainder_mm = width_mm % 300
        # 나머지를 10으로 나눈 개수만큼 1cm 설정가 적용
        units_1cm = remainder_mm // 10
        base_price = (units_30cm * product.get('price_30cm', 0)) + (units_1cm * product.get('price_1cm', 0))
    
    # 추가 옵션 가격 합산
    additional_price = 0
    if additional_options:
        for option in additional_options:
            if isinstance(option, dict) and 'price' in option:
                additional_price += float(option.get('price', 0))
    
    total_price = base_price + additional_price
    return total_price

def apply_coupon(total_price, coupon_type, coupon_value):
    """쿠폰가 적용"""
    if coupon_type == 'percentage':
        # 할인율 적용
        discount = total_price * (float(coupon_value) / 100)
        return total_price - discount
    elif coupon_type == 'fixed':
        # 고정 금액 할인
        return max(0, total_price - float(coupon_value))
    return total_price

@app.route('/wdcalculator')
@login_required
def wdcalculator():
    """견적 계산 메인 페이지"""
    try:
        categories = load_additional_option_categories()
        if categories is None:
            categories = []
        # 추가로 한 번 더 정리 (안전장치)
        categories = clean_categories_data(categories)
    except Exception as e:
        print(f"Error loading categories: {e}")
        categories = []
    
    try:
        notes_categories = load_notes_categories()
        if notes_categories is None:
            notes_categories = []
        # 추가로 한 번 더 정리 (안전장치)
        notes_categories = clean_categories_data(notes_categories)
    except Exception as e:
        print(f"Error loading notes categories: {e}")
        notes_categories = []
    
    return render_template('wdcalculator/calculator.html', categories=categories, notes_categories=notes_categories)

@app.route('/wdcalculator/product-settings')
@login_required
def wdcalculator_product_settings():
    """제품 설정 페이지"""
    try:
        products = load_products()
        if products is None:
            products = []
    except Exception as e:
        print(f"Error loading products: {e}")
        products = []
    
    try:
        categories = load_additional_option_categories()
        if categories is None:
            categories = []
        # 추가로 한 번 더 정리 (안전장치)
        categories = clean_categories_data(categories)
    except Exception as e:
        print(f"Error loading categories: {e}")
        categories = []
    
    try:
        notes_categories = load_notes_categories()
        if notes_categories is None:
            notes_categories = []
        # 추가로 한 번 더 정리 (안전장치)
        notes_categories = clean_categories_data(notes_categories)
    except Exception as e:
        print(f"Error loading notes categories: {e}")
        notes_categories = []
    
    return render_template('wdcalculator/product_settings.html', products=products, categories=categories, notes_categories=notes_categories)

@app.route('/api/wdcalculator/products', methods=['GET'])
@login_required
def api_wdcalculator_get_products():
    """제품 목록 조회"""
    products = load_products()
    return jsonify({'success': True, 'products': products})

@app.route('/api/wdcalculator/products', methods=['POST'])
@login_required
def api_wdcalculator_save_product():
    """제품 추가/수정"""
    try:
        data = request.get_json()
        products = load_products()
        
        product_id = data.get('id')
        
        if product_id:
            # 수정
            for i, product in enumerate(products):
                if product['id'] == product_id:
                    products[i] = data
                    break
        else:
            # 추가
            new_id = max([p['id'] for p in products], default=0) + 1
            data['id'] = new_id
            products.append(data)
        
        if save_products(products):
            return jsonify({'success': True, 'message': '제품이 저장되었습니다.'})
        else:
            return jsonify({'success': False, 'message': '제품 저장에 실패했습니다.'})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})

@app.route('/api/wdcalculator/products/<int:product_id>', methods=['DELETE'])
@login_required
def api_wdcalculator_delete_product(product_id):
    """제품 삭제"""
    try:
        products = load_products()
        products = [p for p in products if p['id'] != product_id]
        
        if save_products(products):
            return jsonify({'success': True, 'message': '제품이 삭제되었습니다.'})
        else:
            return jsonify({'success': False, 'message': '제품 삭제에 실패했습니다.'})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})

@app.route('/api/wdcalculator/calculate', methods=['POST'])
@login_required
def api_wdcalculator_calculate():
    """견적 계산 API"""
    try:
        data = request.get_json()
        product_id = data.get('product_id')
        width_mm = float(data.get('width_mm', 0))
        additional_options = data.get('additional_options', [])
        coupon_type = data.get('coupon_type', 'percentage')
        coupon_value = data.get('coupon_value', 0)
        
        products = load_products()
        product = next((p for p in products if p['id'] == product_id), None)
        
        if not product:
            return jsonify({'success': False, 'message': '제품을 찾을 수 없습니다.'})
        
        total_price = calculate_estimate(product, width_mm, additional_options)
        final_price = apply_coupon(total_price, coupon_type, coupon_value)
        
        return jsonify({
            'success': True,
            'base_price': total_price,
            'final_price': final_price
        })
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})

# 추가 옵션 카테고리 관리 API
@app.route('/api/wdcalculator/additional-options/categories', methods=['GET'])
@login_required
def api_wdcalculator_get_categories():
    """추가 옵션 카테고리 목록 조회"""
    categories = load_additional_option_categories()
    return jsonify({'success': True, 'categories': categories})

@app.route('/api/wdcalculator/additional-options/categories', methods=['POST'])
@login_required
def api_wdcalculator_save_category():
    """추가 옵션 카테고리 추가/수정"""
    try:
        data = request.get_json()
        if not data:
            return jsonify({'success': False, 'message': '데이터가 없습니다.'})
        
        # 필수 필드 검증
        if not data.get('name'):
            return jsonify({'success': False, 'message': '카테고리명을 입력해주세요.'})
        
        categories = load_additional_option_categories()
        
        category_id = data.get('id')
        
        # 데이터 정리
        category_data = {
            'name': data.get('name', '').strip(),
            'options': data.get('options', [])
        }
        
        if category_id:
            # 수정
            category_data['id'] = category_id
            found = False
            for i, category in enumerate(categories):
                if category.get('id') == category_id:
                    # 카테고리명만 업데이트 (options는 기존 것 유지)
                    category['name'] = category_data['name']
                    # options가 명시적으로 전달된 경우에만 업데이트 (카테고리 추가 시에만 사용)
                    if 'options' in category_data and category_data['options'] is not None:
                    # 기존 옵션은 유지하고 새로운 옵션만 추가
                        existing_options = category.get('options', [])
                        # 기존 옵션 ID 유지
                        for new_option in category_data['options']:
                            if 'id' not in new_option or not new_option.get('id'):
                                # 새 옵션 ID 생성
                                option_ids = [o.get('id') or 0 for o in existing_options if o.get('id')]
                                new_option_id = max(option_ids, default=0) + 1
                                new_option['id'] = new_option_id
                                existing_options.append(new_option)
                        category['options'] = existing_options
                    # options가 없으면 기존 옵션 유지 (카테고리명만 변경)
                    found = True
                    break
            if not found:
                return jsonify({'success': False, 'message': '카테고리를 찾을 수 없습니다.'})
        else:
            # 추가 - 카테고리명으로 기존 카테고리 찾기
            existing_category = next((c for c in categories if c.get('name') == category_data['name']), None)
            if existing_category:
                # 기존 카테고리에 옵션 추가
                if 'options' in category_data and category_data['options']:
                    existing_options = existing_category.get('options', [])
                    for new_option in category_data['options']:
                        if 'id' not in new_option or not new_option.get('id'):
                            option_ids = [o.get('id') or 0 for o in existing_options if o.get('id')]
                            new_option_id = max(option_ids, default=0) + 1
                            new_option['id'] = new_option_id
                            existing_options.append(new_option)
                    existing_category['options'] = existing_options
            else:
                # 새 카테고리 생성
                new_id = max([c.get('id', 0) for c in categories], default=0) + 1
                category_data['id'] = new_id
                if 'options' not in category_data:
                    category_data['options'] = []
                # 옵션에 ID 부여
                for option in category_data['options']:
                    if 'id' not in option or not option.get('id'):
                        # 모든 카테고리의 모든 옵션에서 최대 ID 찾기
                        all_option_ids = []
                        for cat in categories:
                            if cat.get('options'):
                                all_option_ids.extend([o.get('id') or 0 for o in cat['options'] if o.get('id')])
                        # 현재 카테고리의 옵션 ID도 확인
                        current_option_ids = [o.get('id') or 0 for o in category_data['options'] if o.get('id')]
                        all_option_ids.extend(current_option_ids)
                        option_id = max(all_option_ids, default=0) + 1
                        option['id'] = option_id
                categories.append(category_data)
        
        # 데이터 정리 후 저장
        cleaned_categories = clean_categories_data(categories)
        if save_additional_option_categories(cleaned_categories):
            # 수정 모드인 경우 업데이트된 category 객체 반환, 추가 모드인 경우 category_data 반환
            if category_id:
                updated_category = next((c for c in cleaned_categories if c.get('id') == category_id), None)
                if updated_category:
                    # 카테고리 객체를 복사하여 반환 (참조 문제 방지)
                    return_category_copy = {
                        'id': updated_category.get('id'),
                        'name': updated_category.get('name', ''),
                        'options': updated_category.get('options', [])[:] if updated_category.get('options') else []
                    }
                    return jsonify({'success': True, 'message': '카테고리가 저장되었습니다.', 'category': return_category_copy})
            return jsonify({'success': True, 'message': '카테고리가 저장되었습니다.'})
        else:
            return jsonify({'success': False, 'message': '카테고리 저장에 실패했습니다.'})
    except Exception as e:
        import traceback
        print(f"Error saving category: {e}")
        print(traceback.format_exc())
        return jsonify({'success': False, 'message': f'오류가 발생했습니다: {str(e)}'})

@app.route('/api/wdcalculator/additional-options/categories/<int:category_id>', methods=['DELETE'])
@login_required
def api_wdcalculator_delete_category(category_id):
    """추가 옵션 카테고리 삭제"""
    try:
        categories = load_additional_option_categories()
        categories = [c for c in categories if c['id'] != category_id]
        
        if save_additional_option_categories(categories):
            return jsonify({'success': True, 'message': '카테고리가 삭제되었습니다.'})
        else:
            return jsonify({'success': False, 'message': '카테고리 삭제에 실패했습니다.'})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})

@app.route('/api/wdcalculator/additional-options/categories/<int:category_id>/options', methods=['POST'])
@login_required
def api_wdcalculator_save_option(category_id):
    """카테고리 내 옵션 추가/수정"""
    try:
        data = request.get_json()
        if not data:
            return jsonify({'success': False, 'message': '데이터가 없습니다.'})
        
        # 필수 필드 검증
        if not data.get('name'):
            return jsonify({'success': False, 'message': '옵션명을 입력해주세요.'})
        if data.get('price') is None:
            return jsonify({'success': False, 'message': '가격을 입력해주세요.'})
        
        categories = load_additional_option_categories()
        
        category = next((c for c in categories if c.get('id') == category_id), None)
        if not category:
            return jsonify({'success': False, 'message': '카테고리를 찾을 수 없습니다.'})
        
        # 데이터 정리
        option_data = {
            'name': data.get('name', '').strip(),
            'price': int(float(data.get('price', 0)))
        }
        
        option_id = data.get('id')
        
        if option_id:
            # 수정
            option_data['id'] = option_id
            found = False
            for i, option in enumerate(category.get('options', [])):
                if option.get('id') == option_id:
                    category['options'][i] = option_data
                    found = True
                    break
            if not found:
                return jsonify({'success': False, 'message': '옵션을 찾을 수 없습니다.'})
        else:
            # 추가
            if 'options' not in category:
                category['options'] = []
            # None 값을 제외한 ID 목록 생성
            option_ids = [o.get('id') or 0 for o in category['options'] if o.get('id')]
            new_id = max(option_ids, default=0) + 1
            option_data['id'] = new_id
            category['options'].append(option_data)
        
        # 데이터 정리 후 저장
        cleaned_categories = clean_categories_data(categories)
        if save_additional_option_categories(cleaned_categories):
            return jsonify({'success': True, 'message': '옵션이 저장되었습니다.'})
        else:
            return jsonify({'success': False, 'message': '옵션 저장에 실패했습니다.'})
    except Exception as e:
        import traceback
        print(f"Error saving option: {e}")
        print(traceback.format_exc())
        return jsonify({'success': False, 'message': f'오류가 발생했습니다: {str(e)}'})

@app.route('/api/wdcalculator/additional-options/categories/<int:category_id>/options/<int:option_id>', methods=['DELETE'])
@login_required
def api_wdcalculator_delete_option(category_id, option_id):
    """카테고리 내 옵션 삭제"""
    try:
        categories = load_additional_option_categories()
        
        category = next((c for c in categories if c['id'] == category_id), None)
        if not category:
            return jsonify({'success': False, 'message': '카테고리를 찾을 수 없습니다.'})
        
        category['options'] = [o for o in category['options'] if o.get('id') != option_id]
        
        if save_additional_option_categories(categories):
            return jsonify({'success': True, 'message': '옵션이 삭제되었습니다.'})
        else:
            return jsonify({'success': False, 'message': '옵션 삭제에 실패했습니다.'})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})

# 비고 카테고리 관리 API
@app.route('/api/wdcalculator/notes/categories', methods=['GET'])
@login_required
def api_wdcalculator_get_notes_categories():
    """비고 카테고리 목록 조회"""
    categories = load_notes_categories()
    return jsonify({'success': True, 'categories': categories})

@app.route('/api/wdcalculator/notes/categories', methods=['POST'])
@login_required
def api_wdcalculator_save_notes_category():
    """비고 카테고리 추가/수정"""
    try:
        data = request.get_json()
        print(f"[DEBUG] api_wdcalculator_save_notes_category 호출됨")
        print(f"[DEBUG] 받은 데이터: {json.dumps(data, ensure_ascii=False, indent=2)}")
        
        if not data:
            return jsonify({'success': False, 'message': '데이터가 없습니다.'})
        
        if not data.get('name'):
            return jsonify({'success': False, 'message': '카테고리명을 입력해주세요.'})
        
        categories = load_notes_categories()
        print(f"[DEBUG] 기존 카테고리 수: {len(categories)}")
        
        category_id = data.get('id')
        
        if category_id:
            # 수정
            print(f"[DEBUG] 카테고리 수정 모드 - category_id: {category_id}")
            category = next((c for c in categories if c.get('id') == category_id), None)
            if category:
                # 카테고리명만 업데이트 (options는 기존 것 유지)
                category['name'] = data.get('name', '').strip()
                # options가 명시적으로 전달된 경우에만 업데이트 (카테고리 추가 시에만 사용)
                if 'options' in data and data['options'] is not None:
                    category['options'] = data['options']
                    print(f"[DEBUG] 카테고리 수정 완료 - name: {category['name']}, options 수: {len(category.get('options', []))}")
                else:
                    print(f"[DEBUG] 카테고리 수정 완료 - name: {category['name']}, 기존 options 유지 (수: {len(category.get('options', []))})")
            else:
                print(f"[DEBUG] 카테고리를 찾을 수 없음 - category_id: {category_id}")
                return jsonify({'success': False, 'message': '카테고리를 찾을 수 없습니다.'})
        else:
            # 추가
            print(f"[DEBUG] 카테고리 추가 모드")
            category_data = {
                'name': data.get('name', '').strip(),
                'options': data.get('options', [])
            }
            max_id = max([c.get('id', 0) for c in categories] + [0])
            category_data['id'] = max_id + 1
            categories.append(category_data)
            print(f"[DEBUG] 새 카테고리 추가 - id: {category_data['id']}, name: {category_data['name']}, options: {len(category_data.get('options', []))}")
        
        if save_notes_categories(categories):
            print(f"[DEBUG] 저장 성공")
            # 수정 모드인 경우 업데이트된 category 객체 반환, 추가 모드인 경우 category_data 반환
            return_category = category if category_id else category_data
            # 카테고리 객체를 복사하여 반환 (참조 문제 방지)
            if return_category:
                return_category_copy = {
                    'id': return_category.get('id'),
                    'name': return_category.get('name'),
                    'options': return_category.get('options', [])[:] if return_category.get('options') else []
                }
                return jsonify({'success': True, 'message': '비고 카테고리가 저장되었습니다.', 'category': return_category_copy})
            return jsonify({'success': True, 'message': '비고 카테고리가 저장되었습니다.', 'category': return_category})
        else:
            print(f"[DEBUG] 저장 실패")
            return jsonify({'success': False, 'message': '비고 카테고리 저장에 실패했습니다.'})
    except Exception as e:
        print(f"[ERROR] api_wdcalculator_save_notes_category 오류: {e}")
        import traceback
        print(traceback.format_exc())
        return jsonify({'success': False, 'message': str(e)})

@app.route('/api/wdcalculator/notes/categories/<int:category_id>', methods=['DELETE'])
@login_required
def api_wdcalculator_delete_notes_category(category_id):
    """비고 카테고리 삭제"""
    try:
        categories = load_notes_categories()
        categories = [c for c in categories if c.get('id') != category_id]
        
        if save_notes_categories(categories):
            return jsonify({'success': True, 'message': '비고 카테고리가 삭제되었습니다.'})
        else:
            return jsonify({'success': False, 'message': '비고 카테고리 삭제에 실패했습니다.'})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})

@app.route('/api/wdcalculator/notes/categories/<int:category_id>/options', methods=['POST'])
@login_required
def api_wdcalculator_save_notes_option(category_id):
    """비고 카테고리 내 옵션 추가/수정"""
    try:
        data = request.get_json()
        print(f"[DEBUG] api_wdcalculator_save_notes_option 호출됨")
        print(f"[DEBUG] category_id: {category_id}")
        print(f"[DEBUG] 받은 데이터: {json.dumps(data, ensure_ascii=False, indent=2)}")
        
        if not data:
            return jsonify({'success': False, 'message': '데이터가 없습니다.'})
        
        if not data.get('name'):
            return jsonify({'success': False, 'message': '옵션명을 입력해주세요.'})
        
        categories = load_notes_categories()
        print(f"[DEBUG] 기존 카테고리 수: {len(categories)}")
        
        category = next((c for c in categories if c.get('id') == category_id), None)
        if not category:
            print(f"[DEBUG] 카테고리를 찾을 수 없음 - category_id: {category_id}")
            return jsonify({'success': False, 'message': '카테고리를 찾을 수 없습니다.'})
        
        print(f"[DEBUG] 카테고리 찾음 - name: {category.get('name')}, 기존 옵션 수: {len(category.get('options', []))}")
        
        option_id = data.get('id')
        option_data = {
            'name': data.get('name', '').strip(),
            'price': 0  # 비고는 가격이 없음
        }
        
        if option_id:
            # 수정
            print(f"[DEBUG] 옵션 수정 모드 - option_id: {option_id}")
            option = next((o for o in category.get('options', []) if o.get('id') == option_id), None)
            if option:
                option.update(option_data)
                print(f"[DEBUG] 옵션 수정 완료")
            else:
                print(f"[DEBUG] 옵션을 찾을 수 없음 - option_id: {option_id}")
                return jsonify({'success': False, 'message': '옵션을 찾을 수 없습니다.'})
        else:
            # 추가
            print(f"[DEBUG] 옵션 추가 모드")
            # 기존 옵션들에서 id가 null이 아닌 것들만 확인
            existing_ids = [o.get('id') for o in category.get('options', []) if o and o.get('id') is not None]
            max_id = max(existing_ids + [0])
            option_data['id'] = max_id + 1
            if 'options' not in category:
                category['options'] = []
            category['options'].append(option_data)
            print(f"[DEBUG] 새 옵션 추가 - id: {option_data['id']}, name: {option_data['name']}")
            print(f"[DEBUG] 추가 후 옵션 수: {len(category.get('options', []))}")
            print(f"[DEBUG] 추가 후 옵션들: {[{'id': o.get('id'), 'name': o.get('name')} for o in category.get('options', [])]}")
            
            # 저장 전에 모든 옵션에 id가 있는지 확인하고 없으면 생성
            for opt in category.get('options', []):
                if opt and opt.get('id') is None:
                    print(f"[DEBUG] id가 null인 옵션 발견, 자동 생성: {opt.get('name')}")
                    existing_ids = [o.get('id') for o in category.get('options', []) if o and o.get('id') is not None]
                    opt['id'] = max(existing_ids + [0]) + 1
                    print(f"[DEBUG] 생성된 id: {opt['id']}")
        
        if save_notes_categories(categories):
            print(f"[DEBUG] 저장 성공")
            return jsonify({'success': True, 'message': '비고 옵션이 저장되었습니다.', 'option': option_data})
        else:
            print(f"[DEBUG] 저장 실패")
            return jsonify({'success': False, 'message': '비고 옵션 저장에 실패했습니다.'})
    except Exception as e:
        print(f"[ERROR] api_wdcalculator_save_notes_option 오류: {e}")
        import traceback
        print(traceback.format_exc())
        return jsonify({'success': False, 'message': str(e)})

@app.route('/api/wdcalculator/notes/categories/<int:category_id>/options/<int:option_id>', methods=['DELETE'])
@login_required
def api_wdcalculator_delete_notes_option(category_id, option_id):
    """비고 카테고리 내 옵션 삭제"""
    try:
        categories = load_notes_categories()
        
        category = next((c for c in categories if c.get('id') == category_id), None)
        if not category:
            return jsonify({'success': False, 'message': '카테고리를 찾을 수 없습니다.'})
        
        original_options = category.get('options', [])
        if not original_options:
            return jsonify({'success': False, 'message': '삭제할 옵션이 없습니다.'})
        
        print(f"[DEBUG] 삭제 시작: category_id={category_id}, option_id={option_id}")
        print(f"[DEBUG] 삭제 전 옵션 수: {len(original_options)}")
        print(f"[DEBUG] 삭제 전 옵션 목록: {[{'id': o.get('id'), 'name': o.get('name')} for o in original_options]}")
        
        # 옵션 ID가 없는 옵션들에 대해 자동으로 ID 생성
        for opt in original_options:
            if opt and opt.get('id') is None:
                existing_ids = [o.get('id') for o in original_options if o and o.get('id') is not None]
                opt['id'] = max(existing_ids + [0]) + 1
                print(f"[DEBUG] 옵션 ID 자동 생성: name={opt.get('name')}, id={opt['id']}")
        
        # 옵션 삭제: ID로 찾거나, ID가 없으면 인덱스로 찾기
        remaining_options = []
        found = False
        
        # 먼저 ID로 찾기 시도
        for i, opt in enumerate(original_options):
            if not opt:
                continue
            
            opt_id = opt.get('id')
            if opt_id is not None:
                # 타입 변환하여 비교
                opt_id = int(opt_id) if isinstance(opt_id, (int, str)) else opt_id
                if opt_id == option_id:
                    found = True
                    # 이 옵션은 삭제 대상이므로 추가하지 않음
                    print(f"[DEBUG] ID로 옵션 찾음: id={opt_id}, name={opt.get('name')}, index={i}")
                    continue
            
            # 삭제 대상이 아니면 남은 옵션에 추가
            remaining_options.append(opt)
        
        # ID로 찾지 못했으면, 인덱스로 찾기 시도
        if not found and 0 <= option_id < len(original_options):
            print(f"[DEBUG] 인덱스로 옵션 찾기 시도: option_id={option_id}, 총 옵션 수={len(original_options)}")
            remaining_options = [opt for i, opt in enumerate(original_options) if i != option_id]
            found = True
            print(f"[DEBUG] 인덱스로 옵션 찾음: 삭제할 인덱스={option_id}, 남은 옵션 수={len(remaining_options)}")
        
        if not found:
            print(f"[DEBUG] 삭제할 옵션을 찾을 수 없음: category_id={category_id}, option_id={option_id}")
            print(f"[DEBUG] 옵션 목록: {[{'id': o.get('id'), 'name': o.get('name')} for o in original_options]}")
            return jsonify({'success': False, 'message': f'삭제할 옵션을 찾을 수 없습니다. (option_id: {option_id})'})
        
        category['options'] = remaining_options
        print(f"[DEBUG] 옵션 삭제 성공: category_id={category_id}, option_id={option_id}")
        print(f"[DEBUG] 삭제 전 옵션 수: {len(original_options)}, 삭제 후 옵션 수: {len(remaining_options)}")
        print(f"[DEBUG] 남은 옵션: {[{'id': o.get('id'), 'name': o.get('name')} for o in remaining_options]}")
        
        if save_notes_categories(categories):
            return jsonify({'success': True, 'message': '비고 옵션이 삭제되었습니다.'})
        else:
            return jsonify({'success': False, 'message': '비고 옵션 삭제에 실패했습니다.'})
    except Exception as e:
        import traceback
        print(f"[ERROR] api_wdcalculator_delete_notes_option 오류: {e}")
        print(traceback.format_exc())
        return jsonify({'success': False, 'message': str(e)})

# 견적 저장 API
@app.route('/api/wdcalculator/save-estimate', methods=['POST'])
@login_required
def api_wdcalculator_save_estimate():
    """견적 저장 (신규 생성 또는 업데이트)"""
    try:
        data = request.get_json()
        
        if not data:
            return jsonify({'success': False, 'message': '요청 데이터가 없습니다.'})
        
        # estimate_id가 있으면 업데이트, 없으면 신규 생성
        estimate_id = data.get('estimate_id')
        customer_name = data.get('customer_name', '').strip()
        estimate_data = data.get('estimate_data', {})
        
        if not customer_name:
            return jsonify({'success': False, 'message': '고객명을 입력해주세요.'})
        
        if not estimate_data:
            return jsonify({'success': False, 'message': '견적 데이터가 없습니다.'})
        
        db = get_wdcalculator_db()
        
        if estimate_id:
            # 기존 견적 업데이트
            estimate = db.query(Estimate).filter(Estimate.id == estimate_id).first()
            if not estimate:
                return jsonify({'success': False, 'message': '수정할 견적을 찾을 수 없습니다.'})
            
            # 히스토리 저장 (변경 전 데이터)
            try:
                history = EstimateHistory(
                    estimate_id=estimate.id,
                    estimate_data=estimate.estimate_data
                )
                db.add(history)
            except Exception as history_error:
                print(f"Warning: Failed to save estimate history: {str(history_error)}")
                # 히스토리 저장 실패해도 견적 업데이트는 진행
            
            # 견적 업데이트
            # customer_name: 일반 문자열
            # estimate_data: 딕셔너리(JSON 객체) 그대로 저장 (SQLAlchemy + JSONB가 자동 처리)
            estimate.customer_name = customer_name
            estimate.estimate_data = estimate_data
            # updated_at은 onupdate=func.now()에 의해 자동 갱신됨
            
            message = '견적이 수정되었습니다.'
        else:
            # 새 견적 생성 (JSONB 사용으로 인코딩 문제 완전 차단)
            estimate = Estimate(
                customer_name=customer_name,
                estimate_data=estimate_data
            )
            db.add(estimate)
            message = '견적이 저장되었습니다.'
            
        db.commit()
        
        return jsonify({
            'success': True,
            'message': message,
            'estimate_id': estimate.id
        })
    except Exception as e:
        db = get_wdcalculator_db()
        db.rollback()
        import traceback
        print(f"Error saving estimate: {str(e)}")
        print(traceback.format_exc())
        return jsonify({'success': False, 'message': f'견적 저장 중 오류: {str(e)}'})

# 견적 검색 API (전체 목록 조회 포함)
@app.route('/api/wdcalculator/search-estimates', methods=['GET'])
@login_required
def api_wdcalculator_search_estimates():
    """고객명으로 견적 검색 (파라미터 없으면 전체 목록)"""
    try:
        customer_name = request.args.get('customer_name', '').strip()
        
        db = get_wdcalculator_db()
        
        query = db.query(Estimate)
        
        if customer_name:
            query = query.filter(Estimate.customer_name.ilike(f'%{customer_name}%'))
            
        # 최신순 정렬
        estimates = query.order_by(Estimate.created_at.desc()).limit(50).all()
        
        # JSONB 사용으로 인해 복잡한 파싱 로직 불필요
        estimates_list = [est.to_dict() for est in estimates]
        
        return jsonify({
            'success': True,
            'estimates': estimates_list,
            'count': len(estimates_list)
        })
    except Exception as e:
        import traceback
        print(f"Error in search_estimates: {str(e)}")
        print(traceback.format_exc())
        return jsonify({'success': False, 'message': f'견적 검색 중 오류: {str(e)}'})

# 견적 조회 및 삭제 API
@app.route('/api/wdcalculator/estimate/<int:estimate_id>', methods=['GET', 'DELETE'])
@login_required
def api_wdcalculator_estimate(estimate_id):
    """견적 ID로 단일 견적 조회 또는 삭제"""
    try:
        db = get_wdcalculator_db()
        estimate = db.query(Estimate).filter(Estimate.id == estimate_id).first()
        
        if not estimate:
            return jsonify({'success': False, 'message': '견적을 찾을 수 없습니다.'})
        
        # DELETE 메서드인 경우 삭제
        if request.method == 'DELETE':
            db.delete(estimate)
            db.commit()
            return jsonify({'success': True, 'message': '견적이 삭제되었습니다.'})
        
        # GET 메서드인 경우 조회
        return jsonify({
            'success': True,
            'estimate': estimate.to_dict()
        })
    except Exception as e:
        db = get_wdcalculator_db()
        db.rollback()
        import traceback
        print(f"Error in estimate API: {str(e)}")
        print(traceback.format_exc())
        if request.method == 'DELETE':
            return jsonify({'success': False, 'message': f'견적 삭제 중 오류: {str(e)}'})
        return jsonify({'success': False, 'message': f'견적 조회 중 오류: {str(e)}'})

# 주문과 견적 매칭 API
@app.route('/api/wdcalculator/match-order', methods=['POST'])
@login_required
def api_wdcalculator_match_order():
    """견적과 FOMS 주문 매칭"""
    try:
        data = request.get_json()
        estimate_id = data.get('estimate_id')
        order_id = data.get('order_id')
        
        if not estimate_id or not order_id:
            return jsonify({'success': False, 'message': '견적 ID와 주문 ID가 필요합니다.'})
        
        # 견적 존재 확인 (독립 DB)
        wd_db = get_wdcalculator_db()
        estimate = wd_db.query(Estimate).filter(Estimate.id == estimate_id).first()
        if not estimate:
            return jsonify({'success': False, 'message': '견적을 찾을 수 없습니다.'})
        
        # FOMS 주문 존재 확인 (읽기 전용, FOMS DB 직접 조회)
        foms_db = get_db()
        order = foms_db.query(Order).filter(Order.id == order_id).first()
        if not order:
            return jsonify({'success': False, 'message': '주문을 찾을 수 없습니다.'})
        
        # 이미 매칭되어 있는지 확인
        existing_match = wd_db.query(EstimateOrderMatch).filter(
            EstimateOrderMatch.estimate_id == estimate_id,
            EstimateOrderMatch.order_id == order_id
        ).first()
        
        if existing_match:
            return jsonify({'success': False, 'message': '이미 매칭된 주문입니다.'})
        
        # 매칭 생성
        match = EstimateOrderMatch(
            estimate_id=estimate_id,
            order_id=order_id
        )
        
        wd_db.add(match)
        wd_db.commit()
        
        return jsonify({
            'success': True,
            'message': '견적과 주문이 매칭되었습니다.',
            'match_id': match.id
        })
    except Exception as e:
        wd_db = get_wdcalculator_db()
        wd_db.rollback()
        return jsonify({'success': False, 'message': f'매칭 중 오류: {str(e)}'})

# 주문별 견적 조회 API
@app.route('/api/wdcalculator/order-estimates/<int:order_id>', methods=['GET'])
@login_required
def api_wdcalculator_get_order_estimates(order_id):
    """특정 주문에 매칭된 견적 조회"""
    try:
        # FOMS 주문 존재 확인 (읽기 전용)
        foms_db = get_db()
        order = foms_db.query(Order).filter(Order.id == order_id).first()
        if not order:
            return jsonify({'success': False, 'message': '주문을 찾을 수 없습니다.'})
        
        # 매칭된 견적 조회 (독립 DB)
        wd_db = get_wdcalculator_db()
        matches = wd_db.query(EstimateOrderMatch).filter(
            EstimateOrderMatch.order_id == order_id
        ).all()
        
        estimates = []
        for match in matches:
            estimate = wd_db.query(Estimate).filter(Estimate.id == match.estimate_id).first()
            if estimate:
                # JSONB 사용으로 인해 복잡한 파싱 로직 불필요
                estimates.append(estimate.to_dict())
        
        return jsonify({
            'success': True,
            'estimates': estimates,
            'count': len(estimates)
        })
    except Exception as e:
        import traceback
        print(f"Error in get_order_estimates: {str(e)}")
        print(traceback.format_exc())
        return jsonify({'success': False, 'message': f'견적 조회 중 오류: {str(e)}'})

# FOMS 주문 검색 API (고객명으로, 견적 매칭용)
@app.route('/api/wdcalculator/search-orders', methods=['GET'])
@login_required
def api_wdcalculator_search_orders():
    """고객명으로 FOMS 주문 검색 (견적 매칭용, 읽기 전용)"""
    try:
        customer_name = request.args.get('customer_name', '').strip()
        
        if not customer_name:
            return jsonify({'success': False, 'message': '고객명을 입력해주세요.'})
        
        # FOMS 주문 검색 (읽기 전용)
        foms_db = get_db()
        orders = foms_db.query(Order).filter(
            Order.customer_name.ilike(f'%{customer_name}%')
        ).order_by(Order.created_at.desc()).limit(50).all()
        
        orders_list = [{
            'id': order.id,
            'customer_name': order.customer_name,
            'phone': order.phone,
            'address': order.address,
            'product': order.product,
            'status': order.status,
            'received_date': order.received_date,
            'created_at': order.created_at.strftime('%Y-%m-%d %H:%M:%S') if order.created_at else None
        } for order in orders]
        
        return jsonify({
            'success': True,
            'orders': orders_list,
            'count': len(orders_list)
        })
    except Exception as e:
        return jsonify({'success': False, 'message': f'주문 검색 중 오류: {str(e)}'})

# ============================================
# 채팅 시스템 API (Quest 3)
# ============================================

@app.route('/api/chat/upload', methods=['POST'])
@login_required
def api_chat_upload():
    """채팅 파일 업로드 API (Quest 3)"""
    try:
        if 'file' not in request.files:
            return jsonify({'success': False, 'message': '파일이 선택되지 않았습니다.'}), 400
        
        file = request.files['file']
        room_id = request.form.get('room_id')  # 선택사항 (임시 업로드 시 None 가능)
        
        if file.filename == '':
            return jsonify({'success': False, 'message': '파일명이 없습니다.'}), 400
        
        # 파일 확장자 검증
        if not allowed_chat_file(file.filename):
            allowed_exts = ', '.join(sorted(CHAT_ALLOWED_EXTENSIONS))
            return jsonify({
                'success': False,
                'message': f'허용되지 않은 파일 형식입니다. 지원 형식: {allowed_exts}'
            }), 400
        
        # 파일 크기 검증
        file.seek(0, os.SEEK_END)
        file_size = file.tell()
        file.seek(0)
        
        max_size = get_chat_file_max_size(file.filename)
        if file_size > max_size:
            size_mb = max_size / (1024 * 1024)
            return jsonify({
                'success': False,
                'message': f'파일 크기가 너무 큽니다. 최대 {size_mb:.0f}MB까지 업로드 가능합니다.'
            }), 400
        
        # 스토리지에 업로드
        storage = get_storage()
        
        # 임시 메시지 ID 생성 (실제 메시지 ID는 나중에 생성됨)
        temp_id = f"temp_{int(datetime.datetime.now().timestamp() * 1000)}"
        if room_id:
            temp_id = f"room_{room_id}_{temp_id}"
        
        # 파일 업로드
        # NOTE:
        # - 요청 지연/스레드 점유를 줄이기 위해 업로드 요청에서는 썸네일을 동기 생성하지 않는다.
        # - 이미지 썸네일은 메시지 저장 후 백그라운드에서 생성한다.
        result = storage.upload_chat_file(file, file.filename, temp_id, generate_thumbnail=False)
        
        if not result.get('success'):
            return jsonify({
                'success': False,
                'message': f'파일 업로드 실패: {result.get("error", "알 수 없는 오류")}'
            }), 500
        
        # ⚠️ presigned URL 만료 문제 방지:
        # DB/클라이언트에 presigned URL을 저장/전달하지 않고, key 기반 view/download 엔드포인트를 사용한다.
        storage_key = result.get('key')
        file_url = build_file_view_url(storage_key)
        thumbnail_key = result.get('thumbnail_key')
        thumbnail_url = build_file_view_url(thumbnail_key) if thumbnail_key else None
        file_info = {
            'filename': file.filename,
            'url': file_url,
            'storage_url': file_url,  # 호환성을 위해 추가
            'thumbnail_url': thumbnail_url,
            'file_type': result.get('file_type'),
            'size': file_size,
            'key': storage_key,
            'download_url': f"/api/chat/download/{storage_key}"
        }
        
        # 로그 기록
        log_access(
            f"채팅 파일 업로드: {file.filename} ({result.get('file_type')}, {file_size / 1024 / 1024:.2f}MB)",
            session.get('user_id')
        )
        
        return jsonify({
            'success': True,
            'message': '파일이 성공적으로 업로드되었습니다.',
            'file_info': file_info
        })
        
    except Exception as e:
        import traceback
        print(f"채팅 파일 업로드 오류: {e}")
        print(traceback.format_exc())
        return jsonify({
            'success': False,
            'message': f'파일 업로드 중 오류가 발생했습니다: {str(e)}'
        }), 500

@app.route('/api/chat/download/<path:storage_key>', methods=['GET'])
@login_required
def api_chat_download(storage_key):
    """채팅 파일 다운로드 API (Quest 4)"""
    try:
        # 보안: 경로 탐색 공격 방지
        if '..' in storage_key or storage_key.startswith('/'):
            return jsonify({'success': False, 'message': '잘못된 파일 경로입니다.'}), 400
        
        storage = get_storage()
        
        # 서명된 URL 생성 (클라우드 스토리지) 또는 직접 경로 반환 (로컬)
        if storage.storage_type in ['r2', 's3']:
            # 클라우드 스토리지: 서명된 URL로 리다이렉트
            url = storage.get_download_url(storage_key, expires_in=3600)
            if url:
                log_access(f"채팅 파일 다운로드 요청: {storage_key}", session.get('user_id'))
                return redirect(url)
            else:
                return jsonify({'success': False, 'message': '파일을 찾을 수 없습니다.'}), 404
        else:
            # 로컬 저장소: 직접 파일 전송
            file_path = os.path.join(storage.upload_folder, storage_key)
            if not os.path.exists(file_path):
                return jsonify({'success': False, 'message': '파일을 찾을 수 없습니다.'}), 404
            
            log_access(f"채팅 파일 다운로드: {storage_key}", session.get('user_id'))
            return send_file(file_path, as_attachment=True)
            
    except Exception as e:
        import traceback
        print(f"파일 다운로드 오류: {e}")
        print(traceback.format_exc())
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/chat/preview/<path:storage_key>', methods=['GET'])
@login_required
def api_chat_preview(storage_key):
    """채팅 파일 미리보기 API (Quest 4)"""
    try:
        # 보안: 경로 탐색 공격 방지
        if '..' in storage_key or storage_key.startswith('/'):
            return jsonify({'success': False, 'message': '잘못된 파일 경로입니다.'}), 400
        
        storage = get_storage()
        
        # 파일 타입 확인
        filename = storage_key.rsplit('/', 1)[-1] if '/' in storage_key else storage_key
        file_type = storage._get_file_type(filename)
        
        if file_type == 'image':
            # 이미지: 썸네일 또는 원본 반환
            if storage.storage_type in ['r2', 's3']:
                # 썸네일이 있으면 썸네일 URL 반환
                # 실제로는 DB에서 thumbnail_url을 조회해야 하지만, 여기서는 간단히 처리
                url = storage.get_download_url(storage_key, expires_in=3600)
                if url:
                    return redirect(url)
                else:
                    return jsonify({'success': False, 'message': '파일을 찾을 수 없습니다.'}), 404
            else:
                # 로컬: 직접 파일 전송
                file_path = os.path.join(storage.upload_folder, storage_key)
                if os.path.exists(file_path):
                    return send_file(file_path)
                else:
                    return jsonify({'success': False, 'message': '파일을 찾을 수 없습니다.'}), 404
        
        elif file_type == 'video':
            # 동영상: 서명된 URL 반환 (브라우저에서 재생)
            if storage.storage_type in ['r2', 's3']:
                url = storage.get_download_url(storage_key, expires_in=3600)
                if url:
                    return jsonify({
                        'success': True,
                        'type': 'video',
                        'url': url
                    })
                else:
                    return jsonify({'success': False, 'message': '파일을 찾을 수 없습니다.'}), 404
            else:
                # 로컬: 직접 파일 전송
                file_path = os.path.join(storage.upload_folder, storage_key)
                if os.path.exists(file_path):
                    return send_file(file_path)
                else:
                    return jsonify({'success': False, 'message': '파일을 찾을 수 없습니다.'}), 404
        
        else:
            # 일반 파일: 다운로드 링크 반환
            return jsonify({
                'success': False,
                'message': '미리보기를 지원하지 않는 파일 형식입니다.',
                'type': 'file'
            }), 400
            
    except Exception as e:
        import traceback
        print(f"파일 미리보기 오류: {e}")
        print(traceback.format_exc())
        return jsonify({'success': False, 'message': str(e)}), 500

# ============================================
# 채팅방 관리 API (Quest 6)
# ============================================

@app.route('/api/chat/rooms', methods=['GET'])
@login_required
def api_chat_rooms_list():
    """채팅방 목록 조회 API (Quest 6)"""
    try:
        db = get_db()
        user_id = session.get('user_id')

        # 1) 사용자가 속한 채팅방 + 멤버 메타(마지막 읽은 시각)를 한 번에 조회
        memberships = db.query(ChatRoomMember, ChatRoom).join(
            ChatRoom,
            ChatRoom.id == ChatRoomMember.room_id
        ).filter(
            ChatRoomMember.user_id == user_id
        ).order_by(
            func.coalesce(ChatRoom.updated_at, ChatRoom.created_at).desc()
        ).all()

        if not memberships:
            return jsonify({'success': True, 'rooms': [], 'count': 0})

        rooms = [room for _, room in memberships]
        room_ids = [room.id for room in rooms]

        # room_id -> member(last_read_at) 매핑
        member_by_room = {member.room_id: member for member, _ in memberships}

        # 2) 채팅방별 마지막 메시지 조회 (배치)
        latest_ts_subq = db.query(
            ChatMessage.room_id.label('room_id'),
            func.max(ChatMessage.created_at).label('max_created_at')
        ).filter(
            ChatMessage.room_id.in_(room_ids)
        ).group_by(
            ChatMessage.room_id
        ).subquery()

        latest_rows = db.query(ChatMessage).join(
            latest_ts_subq,
            and_(
                ChatMessage.room_id == latest_ts_subq.c.room_id,
                ChatMessage.created_at == latest_ts_subq.c.max_created_at
            )
        ).all()

        last_message_by_room = {}
        for msg in latest_rows:
            prev = last_message_by_room.get(msg.room_id)
            if prev is None or (msg.created_at, msg.id) > (prev.created_at, prev.id):
                last_message_by_room[msg.room_id] = msg

        # 3) 채팅방별 읽지 않은 메시지 수 집계 (배치)
        unread_rows = db.query(
            ChatMessage.room_id.label('room_id'),
            func.count(ChatMessage.id).label('unread_count')
        ).join(
            ChatRoomMember,
            and_(
                ChatRoomMember.room_id == ChatMessage.room_id,
                ChatRoomMember.user_id == user_id
            )
        ).filter(
            ChatMessage.room_id.in_(room_ids)
        ).filter(
            or_(
                ChatRoomMember.last_read_at.is_(None),
                ChatMessage.created_at > ChatRoomMember.last_read_at
            )
        ).group_by(
            ChatMessage.room_id
        ).all()

        unread_count_by_room = {room_id: int(count or 0) for room_id, count in unread_rows}

        rooms_list = []
        for room in rooms:
            room_data = room.to_dict()
            room_data['last_message'] = (
                last_message_by_room[room.id].to_dict()
                if room.id in last_message_by_room else None
            )
            # 멤버 정보가 없는 비정상 케이스는 0으로 처리
            room_data['unread_count'] = unread_count_by_room.get(room.id, 0) if member_by_room.get(room.id) else 0
            rooms_list.append(room_data)
        
        return jsonify({
            'success': True,
            'rooms': rooms_list,
            'count': len(rooms_list)
        })
        
    except Exception as e:
        import traceback
        print(f"채팅방 목록 조회 오류: {e}")
        print(traceback.format_exc())
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/chat/rooms', methods=['POST'])
@login_required
def api_chat_rooms_create():
    """채팅방 생성 API (Quest 6)"""
    try:
        db = get_db()
        user_id = session.get('user_id')
        data = request.get_json()
        
        # 필수 필드 검증
        name = data.get('name', '').strip()
        if not name:
            return jsonify({'success': False, 'message': '채팅방 이름은 필수입니다.'}), 400
        
        # 채팅방 생성
        new_room = ChatRoom(
            name=name,
            description=data.get('description', '').strip(),
            order_id=data.get('order_id'),  # 선택사항
            created_by=user_id
        )
        db.add(new_room)
        db.commit()
        db.refresh(new_room)
        
        # 생성자를 멤버로 추가
        member = ChatRoomMember(
            room_id=new_room.id,
            user_id=user_id
        )
        db.add(member)
        
        # 추가 멤버가 있으면 추가
        member_ids = data.get('member_ids', [])
        if member_ids:
            for member_id in member_ids:
                if member_id != user_id:  # 자기 자신은 이미 추가됨
                    new_member = ChatRoomMember(
                        room_id=new_room.id,
                        user_id=member_id
                    )
                    db.add(new_member)
        
        db.commit()
        
        log_access(f"채팅방 생성: {name} (ID: {new_room.id})", user_id)
        
        return jsonify({
            'success': True,
            'message': '채팅방이 생성되었습니다.',
            'room': new_room.to_dict()
        }), 201
        
    except Exception as e:
        db.rollback()
        import traceback
        print(f"채팅방 생성 오류: {e}")
        print(traceback.format_exc())
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/chat/rooms/<int:room_id>', methods=['GET'])
@login_required
def api_chat_rooms_detail(room_id):
    """채팅방 상세 조회 API (Quest 6)"""
    try:
        db = get_db()
        user_id = session.get('user_id')
        
        # 채팅방 조회
        room = db.query(ChatRoom).filter(ChatRoom.id == room_id).first()
        if not room:
            return jsonify({'success': False, 'message': '채팅방을 찾을 수 없습니다.'}), 404
        
        # 멤버 확인
        member = db.query(ChatRoomMember).filter(
            ChatRoomMember.room_id == room_id,
            ChatRoomMember.user_id == user_id
        ).first()
        
        if not member:
            return jsonify({'success': False, 'message': '채팅방에 접근할 권한이 없습니다.'}), 403
        
        # 멤버 목록 조회
        members = db.query(ChatRoomMember).filter(
            ChatRoomMember.room_id == room_id
        ).all()
        
        # 메시지 목록 조회 (최근 50개)
        messages = db.query(ChatMessage).filter(
            ChatMessage.room_id == room_id
        ).order_by(ChatMessage.created_at.desc()).limit(50).all()
        
        # 각 메시지의 읽음 상태 계산
        messages_with_read_status = []
        for msg in messages:
            msg_dict = msg.to_dict()
            
            # 첨부파일 정보 추가
            attachments = db.query(ChatAttachment).filter(
                ChatAttachment.message_id == msg.id
            ).all()
            if attachments:
                msg_dict['attachments'] = [a.to_dict() for a in attachments]
            
            # 자신의 메시지인 경우에만 읽음 상태 계산
            if msg.user_id == user_id:
                # 다른 멤버들이 이 메시지를 읽었는지 확인
                read_count = 0
                total_other_members = 0
                
                for member in members:
                    if member.user_id != user_id:  # 자신 제외
                        total_other_members += 1
                        if member.last_read_at and member.last_read_at >= msg.created_at:
                            read_count += 1
                
                # 읽음 상태 설정
                if total_other_members == 0:
                    msg_dict['read_status'] = 'no_other_members'  # 다른 멤버가 없음
                elif read_count == 0:
                    msg_dict['read_status'] = 'unread'  # 아직 아무도 읽지 않음
                elif read_count == total_other_members:
                    msg_dict['read_status'] = 'all_read'  # 모두 읽음
                else:
                    msg_dict['read_status'] = 'some_read'  # 일부 읽음
                
                msg_dict['read_count'] = read_count
                msg_dict['total_other_members'] = total_other_members
            else:
                msg_dict['read_status'] = None  # 자신의 메시지가 아니면 읽음 상태 없음
                msg_dict['read_count'] = 0
                msg_dict['total_other_members'] = 0
            
            messages_with_read_status.append(msg_dict)
        
        room_data = room.to_dict()
        # 멤버 정보에 사용자 이름 포함
        room_data['members'] = [{
            **m.to_dict(),
            'user_name': m.user.name if m.user else None,
            'user_username': m.user.username if m.user else None
        } for m in members]
        room_data['messages'] = list(reversed(messages_with_read_status))  # 오래된 순으로 정렬
        
        # 주문 정보 조회 (연결된 주문이 있는 경우) - Quest 9
        if room.order_id:
            try:
                order = db.query(Order).filter(Order.id == room.order_id).first()
                if order:
                    order_data = order.to_dict()
                    
                    # 견적 정보도 함께 조회
                    try:
                        wd_db = get_wdcalculator_db()
                        estimates = wd_db.query(EstimateOrderMatch).filter(
                            EstimateOrderMatch.order_id == room.order_id
                        ).all()
                        
                        estimate_list = []
                        for match in estimates:
                            estimate = wd_db.query(Estimate).filter(
                                Estimate.id == match.estimate_id
                            ).first()
                            if estimate:
                                estimate_list.append(estimate.to_dict())
                        
                        order_data['estimates'] = estimate_list
                    except Exception as e:
                        print(f"견적 정보 조회 오류 (무시): {e}")
                        order_data['estimates'] = []
                    
                    room_data['order'] = order_data
            except Exception as e:
                print(f"주문 정보 조회 오류 (무시): {e}")
                room_data['order'] = None
        else:
            room_data['order'] = None
        
        return jsonify({
            'success': True,
            'room': room_data
        })
        
    except Exception as e:
        import traceback
        print(f"채팅방 상세 조회 오류: {e}")
        print(traceback.format_exc())
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/chat/rooms/<int:room_id>', methods=['PUT'])
@login_required
def api_chat_rooms_update(room_id):
    """채팅방 수정 API (Quest 6)"""
    try:
        db = get_db()
        user_id = session.get('user_id')
        data = request.get_json()
        
        # 채팅방 조회
        room = db.query(ChatRoom).filter(ChatRoom.id == room_id).first()
        if not room:
            return jsonify({'success': False, 'message': '채팅방을 찾을 수 없습니다.'}), 404
        
        # 권한 확인 (생성자만 수정 가능)
        if room.created_by != user_id:
            return jsonify({'success': False, 'message': '채팅방을 수정할 권한이 없습니다.'}), 403
        
        # 수정 가능한 필드 업데이트
        if 'name' in data:
            room.name = data['name'].strip()
        if 'description' in data:
            room.description = data.get('description', '').strip()
        if 'order_id' in data:
            room.order_id = data.get('order_id')
        
        room.updated_at = datetime.datetime.now()
        db.commit()
        
        log_access(f"채팅방 수정: {room.name} (ID: {room_id})", user_id)
        
        return jsonify({
            'success': True,
            'message': '채팅방이 수정되었습니다.',
            'room': room.to_dict()
        })
        
    except Exception as e:
        db.rollback()
        import traceback
        print(f"채팅방 수정 오류: {e}")
        print(traceback.format_exc())
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/chat/rooms/<int:room_id>', methods=['DELETE'])
@login_required
def api_chat_rooms_delete(room_id):
    """채팅방 삭제 API (Quest 6)"""
    try:
        db = get_db()
        user_id = session.get('user_id')
        
        # 채팅방 조회
        room = db.query(ChatRoom).filter(ChatRoom.id == room_id).first()
        if not room:
            return jsonify({'success': False, 'message': '채팅방을 찾을 수 없습니다.'}), 404
        
        # 권한 확인 (생성자만 삭제 가능)
        if room.created_by != user_id:
            return jsonify({'success': False, 'message': '채팅방을 삭제할 권한이 없습니다.'}), 403
        
        room_name = room.name
        
        # 채팅방 삭제 (CASCADE로 메시지, 멤버도 자동 삭제됨)
        db.delete(room)
        db.commit()
        
        log_access(f"채팅방 삭제: {room_name} (ID: {room_id})", user_id)
        
        return jsonify({
            'success': True,
            'message': '채팅방이 삭제되었습니다.'
        })
        
    except Exception as e:
        db.rollback()
        import traceback
        print(f"채팅방 삭제 오류: {e}")
        print(traceback.format_exc())
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/chat/rooms/<int:room_id>/members', methods=['POST'])
@login_required
def api_chat_rooms_add_member(room_id):
    """채팅방 멤버 추가 API (Quest 6)"""
    try:
        db = get_db()
        user_id = session.get('user_id')
        data = request.get_json()
        
        # 채팅방 조회
        room = db.query(ChatRoom).filter(ChatRoom.id == room_id).first()
        if not room:
            return jsonify({'success': False, 'message': '채팅방을 찾을 수 없습니다.'}), 404
        
        # 멤버 확인 (멤버만 다른 멤버 추가 가능)
        member = db.query(ChatRoomMember).filter(
            ChatRoomMember.room_id == room_id,
            ChatRoomMember.user_id == user_id
        ).first()
        
        if not member:
            return jsonify({'success': False, 'message': '채팅방에 접근할 권한이 없습니다.'}), 403
        
        # 추가할 사용자 ID
        new_member_id = data.get('user_id')
        if not new_member_id:
            return jsonify({'success': False, 'message': '사용자 ID는 필수입니다.'}), 400
        
        # 사용자 존재 확인
        user = db.query(User).filter(User.id == new_member_id).first()
        if not user:
            return jsonify({'success': False, 'message': '사용자를 찾을 수 없습니다.'}), 404
        
        # 이미 멤버인지 확인
        existing_member = db.query(ChatRoomMember).filter(
            ChatRoomMember.room_id == room_id,
            ChatRoomMember.user_id == new_member_id
        ).first()
        
        if existing_member:
            return jsonify({'success': False, 'message': '이미 채팅방 멤버입니다.'}), 400
        
        # 멤버 추가
        new_member = ChatRoomMember(
            room_id=room_id,
            user_id=new_member_id
        )
        db.add(new_member)
        db.commit()
        
        log_access(f"채팅방 멤버 추가: 방 {room_id}, 사용자 {new_member_id}", user_id)
        
        return jsonify({
            'success': True,
            'message': '멤버가 추가되었습니다.',
            'member': new_member.to_dict()
        }), 201
        
    except Exception as e:
        db.rollback()
        import traceback
        print(f"멤버 추가 오류: {e}")
        print(traceback.format_exc())
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/chat/rooms/<int:room_id>/members/<int:member_user_id>', methods=['DELETE'])
@login_required
def api_chat_rooms_remove_member(room_id, member_user_id):
    """채팅방 멤버 제거 API (Quest 6)"""
    try:
        db = get_db()
        user_id = session.get('user_id')
        
        # 채팅방 조회
        room = db.query(ChatRoom).filter(ChatRoom.id == room_id).first()
        if not room:
            return jsonify({'success': False, 'message': '채팅방을 찾을 수 없습니다.'}), 404
        
        # 권한 확인 (생성자 또는 자기 자신만 제거 가능)
        if room.created_by != user_id and member_user_id != user_id:
            return jsonify({'success': False, 'message': '멤버를 제거할 권한이 없습니다.'}), 403
        
        # 멤버 조회
        member = db.query(ChatRoomMember).filter(
            ChatRoomMember.room_id == room_id,
            ChatRoomMember.user_id == member_user_id
        ).first()
        
        if not member:
            return jsonify({'success': False, 'message': '멤버를 찾을 수 없습니다.'}), 404
        
        # 멤버 제거
        db.delete(member)
        db.commit()
        
        log_access(f"채팅방 멤버 제거: 방 {room_id}, 사용자 {member_user_id}", user_id)
        
        return jsonify({
            'success': True,
            'message': '멤버가 제거되었습니다.'
        })
        
    except Exception as e:
        db.rollback()
        import traceback
        print(f"멤버 제거 오류: {e}")
        print(traceback.format_exc())
        return jsonify({'success': False, 'message': str(e)}), 500

# ============================================
# 주문 연동 API (Quest 8)
# ============================================

@app.route('/api/chat/orders/<int:order_id>', methods=['GET'])
@login_required
def api_chat_order_detail(order_id):
    """채팅방에서 사용할 주문 상세 정보 조회 API (Quest 8)"""
    try:
        db = get_db()
        
        # 주문 조회
        order = db.query(Order).filter(Order.id == order_id).first()
        if not order:
            return jsonify({'success': False, 'message': '주문을 찾을 수 없습니다.'}), 404
        
        # 주문 정보 구성
        order_data = order.to_dict()
        
        # 견적 정보 조회 (견적 계산기 DB)
        try:
            wd_db = get_wdcalculator_db()
            estimates = wd_db.query(EstimateOrderMatch).filter(
                EstimateOrderMatch.order_id == order_id
            ).all()
            
            estimate_list = []
            for match in estimates:
                estimate = wd_db.query(Estimate).filter(
                    Estimate.id == match.estimate_id
                ).first()
                if estimate:
                    estimate_list.append(estimate.to_dict())
            
            order_data['estimates'] = estimate_list
        except Exception as e:
            print(f"견적 정보 조회 오류 (무시): {e}")
            order_data['estimates'] = []
        
        return jsonify({
            'success': True,
            'order': order_data
        })
        
    except Exception as e:
        import traceback
        print(f"주문 정보 조회 오류: {e}")
        print(traceback.format_exc())
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/chat/search-orders', methods=['GET'])
@login_required
def api_chat_search_orders():
    """채팅방에서 주문 검색 API (Quest 8)"""
    try:
        db = get_db()
        query = request.args.get('q', '').strip()  # 검색어
        limit = int(request.args.get('limit', 20))  # 최대 결과 수
        
        if not query:
            return jsonify({
                'success': True,
                'orders': [],
                'count': 0
            })
        
        # 고객명 또는 주문 ID로 검색
        orders = db.query(Order).filter(
            or_(
                Order.customer_name.ilike(f'%{query}%'),
                Order.id == query if query.isdigit() else None
            )
        ).filter(
            Order.deleted_at.is_(None)  # 삭제되지 않은 주문만
        ).order_by(Order.created_at.desc()).limit(limit).all()
        
        orders_list = [{
            'id': order.id,
            'customer_name': order.customer_name,
            'phone': order.phone,
            'address': order.address,
            'product': order.product,
            'status': order.status,
            'received_date': order.received_date,
            'created_at': order.created_at.strftime('%Y-%m-%d %H:%M:%S') if order.created_at else None
        } for order in orders]
        
        return jsonify({
            'success': True,
            'orders': orders_list,
            'count': len(orders_list)
        })
        
    except Exception as e:
        import traceback
        print(f"주문 검색 오류: {e}")
        print(traceback.format_exc())
        return jsonify({'success': False, 'message': str(e)}), 500

# ============================================
# 도면 관리 API
# ============================================

@app.route('/api/orders/<int:order_id>/blueprint', methods=['POST'])
@login_required
def api_upload_blueprint(order_id):
    """도면 이미지 업로드"""
    try:
        if 'file' not in request.files:
            return jsonify({'success': False, 'message': '파일이 없습니다.'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'success': False, 'message': '파일명이 없습니다.'}), 400
        
        # 이미지 파일 검증
        allowed_image_exts = ['png', 'jpg', 'jpeg', 'gif', 'webp']
        file_ext = file.filename.rsplit('.', 1)[1].lower() if '.' in file.filename else ''
        if file_ext not in allowed_image_exts:
            return jsonify({'success': False, 'message': '이미지 파일만 업로드 가능합니다. (png, jpg, jpeg, gif, webp)'}), 400
        
        # 파일 크기 검증 (최대 50MB)
        file.seek(0, os.SEEK_END)
        file_size = file.tell()
        file.seek(0)
        max_size = 50 * 1024 * 1024  # 50MB
        if file_size > max_size:
            return jsonify({'success': False, 'message': f'파일 크기가 너무 큽니다. 최대 50MB까지 업로드 가능합니다.'}), 400
        
        # 주문 존재 확인
        db = get_db()
        order = db.query(Order).filter(Order.id == order_id).first()
        if not order:
            return jsonify({'success': False, 'message': '주문을 찾을 수 없습니다.'}), 404
        
        # 기존 도면이 있으면 삭제 (선택사항 - 덮어쓰기)
        # if order.blueprint_image_url:
        #     # 기존 파일 삭제 로직 (필요시 구현)
        #     pass
        
        # 파일 업로드
        storage = get_storage()
        folder = f"orders/{order_id}/blueprint"
        result = storage.upload_file(file, file.filename, folder)
        
        if not result.get('success'):
            return jsonify({'success': False, 'message': '파일 업로드 실패: ' + result.get('message', '알 수 없는 오류')}), 500
        
        # ⚠️ presigned URL 만료 문제 방지:
        # DB에는 key 기반 view URL을 저장하여, R2/S3에서도 항상 동작하게 한다.
        order.blueprint_image_url = build_file_view_url(result.get('key'))
        db.commit()
        
        return jsonify({
            'success': True,
            'url': result.get('url'),
            'message': '도면이 업로드되었습니다.'
        })
    except Exception as e:
        import traceback
        print(f"도면 업로드 오류: {e}")
        print(traceback.format_exc())
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/orders/<int:order_id>/blueprint', methods=['GET'])
@login_required
def api_get_blueprint(order_id):
    """도면 이미지 조회"""
    try:
        db = get_db()
        order = db.query(Order).filter(Order.id == order_id).first()
        if not order:
            return jsonify({'success': False, 'message': '주문을 찾을 수 없습니다.'}), 404
        
        return jsonify({
            'success': True,
            'url': order.blueprint_image_url if order.blueprint_image_url else None
        })
    except Exception as e:
        import traceback
        print(f"도면 조회 오류: {e}")
        print(traceback.format_exc())
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/orders/<int:order_id>/blueprint', methods=['DELETE'])
@login_required
def api_delete_blueprint(order_id):
    """도면 이미지 삭제"""
    try:
        db = get_db()
        order = db.query(Order).filter(Order.id == order_id).first()
        if not order:
            return jsonify({'success': False, 'message': '주문을 찾을 수 없습니다.'}), 404
        
        # 파일 삭제 (선택사항 - URL만 제거)
        # if order.blueprint_image_url:
        #     storage = get_storage()
        #     # URL에서 key 추출하여 삭제 (구현 필요)
        #     pass
        
        order.blueprint_image_url = None
        db.commit()
        
        return jsonify({'success': True, 'message': '도면이 삭제되었습니다.'})
    except Exception as e:
        import traceback
        print(f"도면 삭제 오류: {e}")
        print(traceback.format_exc())
        return jsonify({'success': False, 'message': str(e)}), 500

# ============================================
# 전체 채팅 검색 API (Quest 1)
# ============================================

@app.route('/api/chat/search', methods=['GET'])
@login_required
def api_chat_search():
    """전체 채팅 검색 API - 모든 채팅방의 메시지, 주문 정보 포함"""
    try:
        db = get_db()
        user_id = session.get('user_id')
        query = request.args.get('q', '').strip()
        limit = int(request.args.get('limit', 50))
        
        if not query or len(query) < 2:
            return jsonify({
                'success': True,
                'results': [],
                'count': 0
            })
        
        # 사용자가 멤버로 있는 채팅방만 검색
        user_rooms = db.query(ChatRoom.id).join(
            ChatRoomMember,
            ChatRoom.id == ChatRoomMember.room_id
        ).filter(ChatRoomMember.user_id == user_id).subquery()
        
        results = []
        
        # 1. 메시지 내용 검색
        messages = db.query(ChatMessage).join(
            user_rooms, ChatMessage.room_id == user_rooms.c.id
        ).filter(
            ChatMessage.content.ilike(f'%{query}%')
        ).limit(limit).all()
        
        for msg in messages:
            room = db.query(ChatRoom).filter(ChatRoom.id == msg.room_id).first()
            results.append({
                'type': 'message',
                'room_id': msg.room_id,
                'room_name': room.name if room else None,
                'message_id': msg.id,
                'content': msg.content,
                'user_name': msg.user.name if msg.user else None,
                'created_at': msg.created_at.strftime('%Y-%m-%d %H:%M:%S') if msg.created_at else None
            })
        
        # 2. 채팅방 이름/설명 검색
        rooms = db.query(ChatRoom).join(
            user_rooms, ChatRoom.id == user_rooms.c.id
        ).filter(
            or_(
                ChatRoom.name.ilike(f'%{query}%'),
                ChatRoom.description.ilike(f'%{query}%')
            )
        ).limit(limit).all()
        
        for room in rooms:
            if not any(r.get('room_id') == room.id and r.get('type') == 'room' for r in results):
                results.append({
                    'type': 'room',
                    'room_id': room.id,
                    'room_name': room.name,
                    'description': room.description,
                    'created_at': room.created_at.strftime('%Y-%m-%d %H:%M:%S') if room.created_at else None
                })
        
        # 3. 연결된 주문 정보 검색
        orders = db.query(Order).join(
            ChatRoom, Order.id == ChatRoom.order_id
        ).join(
            user_rooms, ChatRoom.id == user_rooms.c.id
        ).filter(
            or_(
                Order.customer_name.ilike(f'%{query}%'),
                Order.phone.ilike(f'%{query}%'),
                Order.address.ilike(f'%{query}%')
            )
        ).limit(limit).all()
        
        for order in orders:
            room = db.query(ChatRoom).filter(ChatRoom.order_id == order.id).first()
            if room and not any(r.get('room_id') == room.id and r.get('type') == 'order' for r in results):
                results.append({
                    'type': 'order',
                    'room_id': room.id,
                    'room_name': room.name,
                    'order_id': order.id,
                    'customer_name': order.customer_name,
                    'phone': order.phone,
                    'address': order.address,
                    'product': order.product
                })
        
        # 중복 제거 및 정렬
        seen = set()
        unique_results = []
        for r in results:
            key = (r['type'], r.get('room_id'), r.get('message_id', 0), r.get('order_id', 0))
            if key not in seen:
                seen.add(key)
                unique_results.append(r)
        
        return jsonify({
            'success': True,
            'results': unique_results[:limit],
            'count': len(unique_results)
        })
        
    except Exception as e:
        import traceback
        print(f"채팅 검색 오류: {e}")
        print(traceback.format_exc())
        return jsonify({'success': False, 'message': str(e)}), 500

# ============================================
# 메시지 읽음 상태 업데이트 API
# ============================================

@app.route('/api/chat/rooms/<int:room_id>/mark-read', methods=['POST'])
@login_required
def api_chat_mark_read(room_id):
    """메시지 읽음 상태 업데이트 API (Socket.IO 폴백용)"""
    try:
        db = get_db()
        user_id = session.get('user_id')
        
        # 채팅방 존재 확인
        room = db.query(ChatRoom).filter(ChatRoom.id == room_id).first()
        if not room:
            return jsonify({'success': False, 'message': '채팅방을 찾을 수 없습니다.'}), 404
        
        # 멤버 조회
        member = db.query(ChatRoomMember).filter(
            ChatRoomMember.room_id == room_id,
            ChatRoomMember.user_id == user_id
        ).first()
        
        if not member:
            return jsonify({'success': False, 'message': '채팅방 멤버가 아닙니다.'}), 403
        
        # 읽은 시간 업데이트
        member.last_read_at = datetime.datetime.now()
        db.commit()
        
        return jsonify({'success': True})
            
    except Exception as e:
        db.rollback()
        import traceback
        print(f"읽음 상태 업데이트 오류: {e}")
        print(traceback.format_exc())
        return jsonify({'success': False, 'message': str(e)}), 500

# ============================================
# 채팅용 사용자 목록 API (Phase 2)
# ============================================

@app.route('/api/chat/users', methods=['GET'])
@login_required
def api_chat_users_list():
    """채팅 초대용 사용자 목록 조회 API"""
    try:
        db = get_db()
        current_user_id = session.get('user_id')
        
        # 활성 사용자만 조회 (자기 자신 제외)
        users = db.query(User).filter(
            User.is_active == True,
            User.id != current_user_id
        ).order_by(User.name).all()
        
        users_list = [{
            'id': user.id,
            'name': user.name,
            'username': user.username,
            'role': user.role
        } for user in users]
        
        return jsonify({
            'success': True,
            'users': users_list
        })
    except Exception as e:
        import traceback
        print(f"사용자 목록 조회 오류: {e}")
        print(traceback.format_exc())
        return jsonify({'success': False, 'message': str(e)}), 500

# ============================================
# 채팅 메시지 REST API (Socket.IO 폴백용)
# ============================================

@app.route('/api/chat/messages', methods=['POST'])
@login_required
def api_chat_send_message():
    """메시지 전송 API (Socket.IO 없을 때 폴백)"""
    try:
        db = get_db()
        user_id = session.get('user_id')
        data = request.get_json()
        
        room_id = data.get('room_id')
        message_type = data.get('message_type', 'text')
        content = data.get('content', '').strip()
        file_info = data.get('file_info')
        
        if not room_id:
            return jsonify({'success': False, 'message': '채팅방 ID는 필수입니다.'}), 400
        
        # 채팅방 존재 확인
        room = db.query(ChatRoom).filter(ChatRoom.id == room_id).first()
        if not room:
            return jsonify({'success': False, 'message': '채팅방을 찾을 수 없습니다.'}), 404
        
        # 멤버 확인
        member = db.query(ChatRoomMember).filter(
            ChatRoomMember.room_id == room_id,
            ChatRoomMember.user_id == user_id
        ).first()
        
        if not member:
            return jsonify({'success': False, 'message': '채팅방에 접근할 권한이 없습니다.'}), 403
        
        # 메시지 저장
        new_message = ChatMessage(
            room_id=room_id,
            user_id=user_id,
            message_type=message_type,
            content=content if message_type == 'text' else None,
            file_info=file_info if message_type != 'text' else None
        )
        db.add(new_message)
        db.commit()
        db.refresh(new_message)
        
        # 첨부파일이 있으면 저장
        if file_info and isinstance(file_info, dict):
            attachment = ChatAttachment(
                message_id=new_message.id,
                filename=file_info.get('filename', ''),
                file_type=file_info.get('file_type', 'file'),
                file_size=file_info.get('size', 0),
                storage_key=file_info.get('key', ''),
                storage_url=file_info.get('url', ''),
                thumbnail_url=file_info.get('thumbnail_url')
            )
            db.add(attachment)
            db.commit()

            # 이미지 첨부의 썸네일은 비동기로 생성해 요청 지연을 줄인다.
            if attachment.file_type == 'image' and not attachment.thumbnail_url and attachment.storage_key:
                schedule_chat_thumbnail_generation(attachment.storage_key)
        
        # 사용자 정보 포함하여 메시지 데이터 구성
        user = db.query(User).filter(User.id == user_id).first()
        message_data = new_message.to_dict()
        if user:
            message_data['user_name'] = user.name
            message_data['user_username'] = user.username
        
        # 첨부파일 정보 추가
        attachments = db.query(ChatAttachment).filter(
            ChatAttachment.message_id == new_message.id
        ).all()
        if attachments:
            message_data['attachments'] = [a.to_dict() for a in attachments]
        
        # 채팅방 업데이트 시간 갱신
        room.updated_at = datetime.datetime.now()
        db.commit()
        
        return jsonify({
            'success': True,
            'message': message_data
        })
        
    except Exception as e:
        db.rollback()
        import traceback
        print(f"메시지 전송 오류: {e}")
        print(traceback.format_exc())
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/chat/messages/<int:message_id>', methods=['GET'])
@login_required
def api_chat_get_message(message_id):
    """단일 메시지 조회 API"""
    try:
        db = get_db()
        user_id = session.get('user_id')
        
        message = db.query(ChatMessage).filter(ChatMessage.id == message_id).first()
        
        if not message:
            return jsonify({'success': False, 'message': '메시지를 찾을 수 없습니다.'}), 404
        
        # 권한 확인 (채팅방 멤버인지)
        member = db.query(ChatRoomMember).filter(
            ChatRoomMember.room_id == message.room_id,
            ChatRoomMember.user_id == user_id
        ).first()
        
        if not member:
            return jsonify({'success': False, 'message': '권한이 없습니다.'}), 403
        
        # 메시지 데이터 구성
        message_data = message.to_dict()
        user = db.query(User).filter(User.id == message.user_id).first()
        if user:
            message_data['user_name'] = user.name
        
        # 첨부파일 정보 추가
        attachments = db.query(ChatAttachment).filter(
            ChatAttachment.message_id == message.id
        ).all()
        if attachments:
            message_data['attachments'] = [a.to_dict() for a in attachments]
        
        return jsonify({
            'success': True,
            'message': message_data
        })
    except Exception as e:
        import traceback
        print(f"메시지 조회 오류: {e}")
        print(traceback.format_exc())
        return jsonify({'success': False, 'message': str(e)}), 500


# ============================================
# ERP Beta: Structured Order Data APIs (Step 3)
# ============================================

def _ensure_system_build_steps_table(db):
    """B안: 진행상태는 DB 테이블에만 기록. 테이블이 없으면 생성."""
    db.execute(text("""
    CREATE TABLE IF NOT EXISTS system_build_steps (
        step_key VARCHAR(100) PRIMARY KEY,
        status VARCHAR(30) NOT NULL DEFAULT 'PENDING',
        started_at TIMESTAMP NULL,
        completed_at TIMESTAMP NULL,
        message TEXT NULL,
        meta JSONB NULL
    );
    """))
    db.commit()


def _record_build_step(db, step_key, status, message=None, meta=None):
    # 빌드 체크포인트는 "부가 기능"이므로, 실패해도 API 자체는 죽지 않게 한다.
    try:
        _ensure_system_build_steps_table(db)
        meta_json = json.dumps(meta, ensure_ascii=False) if isinstance(meta, (dict, list)) else None
        now = datetime.datetime.now()
        db.execute(
            text("""
            INSERT INTO system_build_steps (step_key, status, started_at, completed_at, message, meta)
            VALUES (:k, :s, :started, :completed, :m, CAST(:meta AS JSONB))
            ON CONFLICT (step_key)
            DO UPDATE SET
                status = EXCLUDED.status,
                started_at = COALESCE(system_build_steps.started_at, EXCLUDED.started_at),
                completed_at = CASE WHEN EXCLUDED.status IN ('COMPLETED','FAILED') THEN EXCLUDED.completed_at ELSE system_build_steps.completed_at END,
                message = EXCLUDED.message,
                meta = COALESCE(EXCLUDED.meta, system_build_steps.meta);
            """),
            {
                "k": step_key,
                "s": status,
                "started": now if status == "RUNNING" else None,
                "completed": now if status in ["COMPLETED", "FAILED"] else None,
                "m": message,
                "meta": meta_json,
            }
        )
        db.commit()
    except Exception as e:
        try:
            db.rollback()
        except Exception:
            pass
        print(f"[ERP_BETA] build-step log warning: {e}")


@app.route('/api/orders/<int:order_id>/structured', methods=['GET'])
@login_required
def api_get_order_structured(order_id):
    """구조화 데이터 조회(전사 공용)."""
    db = get_db()
    try:
        order = db.query(Order).filter(Order.id == order_id, Order.status != 'DELETED').first()
        if not order:
            return jsonify({'success': False, 'message': '주문을 찾을 수 없습니다.'}), 404

        return jsonify({
            'success': True,
            'order_id': order.id,
            'raw_order_text': order.raw_order_text,
            'structured_data': order.structured_data,
            'structured_schema_version': order.structured_schema_version,
            'structured_confidence': order.structured_confidence,
            'structured_updated_at': order.structured_updated_at.strftime('%Y-%m-%d %H:%M:%S') if order.structured_updated_at else None,
            'received_date': order.received_date or '',
            'received_time': order.received_time or ''
        })
    except Exception as e:
        import traceback
        print(f"[ERP_BETA] structured GET 오류: {e}")
        print(traceback.format_exc())
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/orders/<int:order_id>/structured', methods=['PUT'])
@login_required
@role_required(['ADMIN', 'MANAGER', 'STAFF'])
def api_put_order_structured(order_id):
    """구조화 데이터 저장(전사 공용)."""
    db = get_db()
    step_key = f"ERP_BETA_API_SAVE_{order_id}"
    _record_build_step(db, step_key, "RUNNING", message="Saving structured data")
    try:
        order = db.query(Order).filter(Order.id == order_id, Order.status != 'DELETED').first()
        if not order:
            _record_build_step(db, step_key, "FAILED", message="Order not found")
            return jsonify({'success': False, 'message': '주문을 찾을 수 없습니다.'}), 404

        payload = request.get_json(silent=True) or {}
        structured_data = payload.get('structured_data')
        raw_order_text = payload.get('raw_order_text')
        schema_version = payload.get('structured_schema_version', 1)
        confidence = payload.get('structured_confidence')
        received_date = payload.get('received_date')
        received_time = payload.get('received_time')
        now = datetime.datetime.now()
        draft_cleared = False

        # 최소 검증: dict 형태
        if structured_data is not None and not isinstance(structured_data, dict):
            _record_build_step(db, step_key, "FAILED", message="structured_data must be an object")
            return jsonify({'success': False, 'message': 'structured_data는 JSON 객체여야 합니다.'}), 400

        old_sd = order.structured_data or {}

        if raw_order_text is not None:
            order.raw_order_text = raw_order_text
        if received_date is not None and isinstance(received_date, str) and received_date.strip():
            order.received_date = received_date.strip()
        if received_time is not None and isinstance(received_time, str):
            order.received_time = received_time.strip() or None
        if structured_data is not None:
            # Canonical ERP mode: structured_data를 원천 데이터로 사용
            # workflow/flags/assignments 기본 구조 보정
            if not structured_data.get('workflow'):
                structured_data['workflow'] = {}
            if not structured_data.get('flags'):
                structured_data['flags'] = {}
            if not structured_data.get('assignments'):
                structured_data['assignments'] = {}

            # workflow.stage 변경 감지 + timestamp 기록 + 검증
            try:
                new_stage = (structured_data.get('workflow') or {}).get('stage')
                old_stage = (old_sd.get('workflow') or {}).get('stage')
                if new_stage and new_stage != old_stage:
                    # 단계 전환 검증: Quest 시스템 사용 시 팀 승인 완료 여부 확인
                    # (수동 단계 변경 시에도 Quest 승인 완료 여부를 확인)
                    is_quest_complete, missing_teams = check_quest_approvals_complete(old_sd, old_stage)
                    
                    if not is_quest_complete and missing_teams:
                        # Quest 승인 미완료 시 경고 (강제 전환은 허용하되 경고)
                        stage_label = STAGE_LABELS.get(old_stage, old_stage) if old_stage else '알 수 없음'
                        # 팀 라벨 매핑
                        TEAM_LABELS = {
                            'CS': '라홈팀',
                            'SALES': '영업팀',
                            'MEASURE': '실측팀',
                            'DRAWING': '도면팀',
                            'PRODUCTION': '생산팀',
                            'CONSTRUCTION': '시공팀',
                        }
                        missing_team_labels = [TEAM_LABELS.get(t, t) for t in missing_teams]
                        # 경고만 표시하고 전환은 허용 (수동 전환 가능하도록)
                        print(f"경고: [{stage_label}] 단계의 Quest 승인 미완료 팀: {', '.join(missing_team_labels)}")
                    
                    # 단계 전환 허용 (Quest 승인 완료 여부와 관계없이 수동 전환 가능)
                    (structured_data.get('workflow') or {})['stage_updated_at'] = datetime.datetime.now().isoformat()
                    # 이벤트 기록
                    ev = OrderEvent(
                        order_id=order.id,
                        event_type='STAGE_CHANGED',
                        payload={'from': old_stage, 'to': new_stage, 'manual': True},
                        created_by_user_id=session.get('user_id')
                    )
                    db.add(ev)
                    
                    # 새 단계의 Quest가 없으면 자동 생성
                    quests = structured_data.get('quests') or []
                    has_new_stage_quest = any(
                        isinstance(q, dict) and q.get('stage') == new_stage 
                        for q in quests
                    )
                    if not has_new_stage_quest:
                        new_quest = create_quest_from_template(new_stage, session.get('username') or '', structured_data)
                        if new_quest:
                            if not structured_data.get('quests'):
                                structured_data['quests'] = []
                            structured_data['quests'].append(new_quest)
            except Exception as _e:
                import traceback
                print(f"단계 전환 검증 오류: {_e}")
                print(traceback.format_exc())
                pass

            # 긴급 변경 이벤트
            try:
                new_urgent = bool((structured_data.get('flags') or {}).get('urgent'))
                old_urgent = bool((old_sd.get('flags') or {}).get('urgent'))
                if new_urgent != old_urgent:
                    ev = OrderEvent(
                        order_id=order.id,
                        event_type='URGENT_CHANGED',
                        payload={'from': old_urgent, 'to': new_urgent, 'reason': (structured_data.get('flags') or {}).get('urgent_reason')},
                        created_by_user_id=session.get('user_id')
                    )
                    db.add(ev)
            except Exception:
                pass

            # 일정 변경 이벤트(실측/시공)
            try:
                new_meas = ((structured_data.get('schedule') or {}).get('measurement') or {}).get('date')
                old_meas = ((old_sd.get('schedule') or {}).get('measurement') or {}).get('date')
                if new_meas != old_meas:
                    db.add(OrderEvent(
                        order_id=order.id,
                        event_type='MEASUREMENT_DATE_CHANGED',
                        payload={'from': old_meas, 'to': new_meas},
                        created_by_user_id=session.get('user_id')
                    ))
            except Exception:
                pass

            try:
                new_cons = ((structured_data.get('schedule') or {}).get('construction') or {}).get('date')
                old_cons = ((old_sd.get('schedule') or {}).get('construction') or {}).get('date')
                if new_cons != old_cons:
                    db.add(OrderEvent(
                        order_id=order.id,
                        event_type='CONSTRUCTION_DATE_CHANGED',
                        payload={'from': old_cons, 'to': new_cons},
                        created_by_user_id=session.get('user_id')
                    ))
            except Exception:
                pass

            # 오너팀 변경 이벤트
            try:
                new_team = (structured_data.get('assignments') or {}).get('owner_team')
                old_team = (old_sd.get('assignments') or {}).get('owner_team')
                if new_team != old_team:
                    db.add(OrderEvent(
                        order_id=order.id,
                        event_type='OWNER_TEAM_CHANGED',
                        payload={'from': old_team, 'to': new_team},
                        created_by_user_id=session.get('user_id')
                    ))
            except Exception:
                pass

            # ------------------------------------------------------------
            # SLA 기반 자동 Task 생성(중복 방지)
            # ------------------------------------------------------------
            try:
                apply_auto_tasks(db, order.id, structured_data)
            except Exception as _e:
                # 자동화 실패가 저장 자체를 막지 않음
                print(f"[ERP_BETA] auto-task apply warning: {_e}")

            # Draft 주문이면 draft 플래그 해제 (신규 주문 덮어쓰기 방지)
            try:
                meta = structured_data.get('meta') or {}
                if meta.get('draft') is True:
                    meta['draft'] = False
                    meta['finalized_at'] = now.isoformat()
                    structured_data['meta'] = meta
                    draft_cleared = True
            except Exception:
                pass

            order.structured_data = structured_data
        order.structured_schema_version = int(schema_version) if schema_version else 1
        order.structured_confidence = confidence or (structured_data.get('confidence') if structured_data else None)
        order.structured_updated_at = now

        # ERP Beta draft 세션 제거 (다음 새 주문에서 재사용되지 않도록)
        try:
            existing_id = session.get('erp_draft_order_id')
            if existing_id and int(existing_id) == order.id:
                session.pop('erp_draft_order_id', None)
                draft_cleared = True
        except Exception:
            pass

        # NOTE: 기존 주문 정보(Order 컬럼)는 레거시/호환용으로 유지될 수 있으나,
        # ERP 구현 기준에서는 structured_data가 원천 데이터다.
        # (필요 시 별도 스텝/옵션으로 브릿지 재활성화 가능)

        db.commit()
        _record_build_step(db, step_key, "COMPLETED", message="Saved structured data")

        return jsonify({'success': True, 'draft_cleared': draft_cleared})
    except Exception as e:
        db.rollback()
        import traceback
        print(f"[ERP_BETA] structured PUT 오류: {e}")
        print(traceback.format_exc())
        _record_build_step(db, step_key, "FAILED", message=str(e))
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/orders/parse-text', methods=['POST'])
@login_required
@role_required(['ADMIN', 'MANAGER', 'STAFF'])
def api_parse_order_text():
    """텍스트 붙여넣기 → 구조화 파싱(미리보기용). 저장은 하지 않음."""
    db = get_db()
    _record_build_step(db, "ERP_BETA_API_PARSE_TEXT", "RUNNING", message="Parsing order text")
    try:
        payload = request.get_json(silent=True) or {}
        raw_text = (payload.get('raw_text') or '').strip()
        if not raw_text:
            _record_build_step(db, "ERP_BETA_API_PARSE_TEXT", "FAILED", message="raw_text is empty")
            return jsonify({'success': False, 'message': 'raw_text가 필요합니다.'}), 400

        structured = parse_order_text(raw_text)
        _record_build_step(db, "ERP_BETA_API_PARSE_TEXT", "COMPLETED", message="Parsed order text")
        return jsonify({'success': True, 'structured_data': structured})
    except Exception as e:
        import traceback
        print(f"[ERP_BETA] parse-text 오류: {e}")
        print(traceback.format_exc())
        _record_build_step(db, "ERP_BETA_API_PARSE_TEXT", "FAILED", message=str(e))
        return jsonify({'success': False, 'message': str(e)}), 500


# ============================================
# ERP Beta: Draft Order API (for Add Order screen)
# ============================================
@app.route('/api/orders/erp/draft', methods=['POST'])
@login_required
@role_required(['ADMIN', 'MANAGER', 'STAFF'])
def api_erp_create_draft():
    """
    ERP Beta '새 주문' 화면에서 모든 기능(첨부/태스크/이벤트/구조화 저장)을 즉시 사용하기 위해
    order_id를 먼저 확보하는 draft 주문 생성 API.

    - 세션에 draft가 있으면 재사용(중복 생성 방지)
    - 없으면 placeholder Order를 생성하고 is_erp_beta=True로 마킹
    """
    db = get_db()
    try:
        existing_id = session.get('erp_draft_order_id')
        if existing_id:
            order = db.query(Order).filter(Order.id == int(existing_id), Order.status != 'DELETED').first()
            if order:
                return jsonify({'success': True, 'order_id': order.id, 'reused': True})

        now = datetime.datetime.now()
        today = now.strftime('%Y-%m-%d')
        time_str = now.strftime('%H:%M')

        # 최소 placeholder (NOT NULL 컬럼 충족)
        structured = {
            'workflow': {'stage': 'RECEIVED', 'stage_updated_at': now.isoformat()},
            'flags': {'urgent': False},
            'assignments': {},
            'schedule': {},
            'meta': {'draft': True, 'created_via': 'ADD_ORDER'},
        }

        order = Order(
            received_date=today,
            received_time=time_str,
            customer_name='ERP Beta',
            phone='000-0000-0000',
            address='-',
            product='ERP Beta',
            options=None,
            notes=None,
            status='RECEIVED',
            is_erp_beta=True,
            raw_order_text='',
            structured_data=structured,
            structured_schema_version=1,
            structured_confidence=None,
            structured_updated_at=now,
        )
        db.add(order)
        db.commit()
        db.refresh(order)

        session['erp_draft_order_id'] = order.id
        return jsonify({'success': True, 'order_id': order.id, 'reused': False})
    except Exception as e:
        try:
            db.rollback()
        except Exception:
            pass
        import traceback
        print(f"[ERP_BETA] draft create error: {e}")
        print(traceback.format_exc())
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/chat')
@login_required
def chat():
    """채팅 페이지 (Quest 10)"""
    # current_user는 @app.context_processor에서 자동으로 주입됨
    # menu도 @app.context_processor에서 자동으로 주입됨
    return render_template('chat.html', 
                          socketio_available=SOCKETIO_AVAILABLE and socketio is not None)

# ============================================
# SocketIO 이벤트 핸들러 (Quest 5)
# ============================================

if SOCKETIO_AVAILABLE and socketio:
    @socketio.on('connect')
    def handle_connect():
        """클라이언트 연결 이벤트"""
        user_id = session.get('user_id')
        if user_id:
            print(f"[SocketIO] 사용자 {user_id} 연결됨")
            
            # 사용자 전용 room에 join (모든 페이지에서 메시지 받을 수 있게)
            join_room(f'user_{user_id}')
            print(f"[SocketIO] 사용자 {user_id}가 자신의 전용 room에 입장: user_{user_id}")

            # NOTE:
            # - connect 시 모든 채팅방을 미리 join하면 (수백~수천 방) 연결 비용이 급증하고
            #   재연결 로그가 폭증합니다.
            # - 방 join은 사용자가 실제로 방을 열 때(join_room 이벤트) 수행하는 lazy join 전략을 사용합니다.
            emit('connected', {'user_id': user_id, 'message': '연결되었습니다.'})
        else:
            print("[SocketIO] 인증되지 않은 연결 시도")
            return False  # 연결 거부
    
    @socketio.on('disconnect')
    def handle_disconnect():
        """클라이언트 연결 해제 이벤트"""
        user_id = session.get('user_id')
        if user_id:
            print(f"[SocketIO] 사용자 {user_id} 연결 해제됨")
    
    @socketio.on('join_room')
    def handle_join_room(data):
        """채팅방 입장"""
        user_id = session.get('user_id')
        if not user_id:
            emit('error', {'message': '인증이 필요합니다.'})
            return
        
        room_id = data.get('room_id')
        if room_id:
            join_room(str(room_id))
            print(f"[SocketIO] 사용자 {user_id}가 채팅방 {room_id}에 입장")
            emit('joined_room', {'room_id': room_id, 'user_id': user_id})
            # 방의 다른 사용자들에게 알림
            socketio.emit('user_joined', {
                'room_id': room_id,
                'user_id': user_id
            }, room=str(room_id))
    
    @socketio.on('leave_room')
    def handle_leave_room(data):
        """채팅방 퇴장"""
        user_id = session.get('user_id')
        if not user_id:
            emit('error', {'message': '인증이 필요합니다.'})
            return
        
        room_id = data.get('room_id')
        if room_id:
            leave_room(str(room_id))
            print(f"[SocketIO] 사용자 {user_id}가 채팅방 {room_id}에서 퇴장")
            emit('left_room', {'room_id': room_id, 'user_id': user_id})
            # 방의 다른 사용자들에게 알림
            socketio.emit('user_left', {
                'room_id': room_id,
                'user_id': user_id
            }, room=str(room_id))
    
    @socketio.on('send_message')
    def handle_send_message(data):
        """메시지 전송 (Quest 7)"""
        user_id = session.get('user_id')
        if not user_id:
            emit('error', {'message': '인증이 필요합니다.'})
            return
        
        try:
            db = get_db()
            room_id = data.get('room_id')
            message_type = data.get('message_type', 'text')
            content = data.get('content', '').strip()
            file_info = data.get('file_info')  # 파일 정보 (업로드 후)
            
            if not room_id:
                emit('error', {'message': '채팅방 ID는 필수입니다.'})
                return
            
            # 채팅방 존재 확인
            room = db.query(ChatRoom).filter(ChatRoom.id == room_id).first()
            if not room:
                emit('error', {'message': '채팅방을 찾을 수 없습니다.'})
                return
            
            # 멤버 확인
            member = db.query(ChatRoomMember).filter(
                ChatRoomMember.room_id == room_id,
                ChatRoomMember.user_id == user_id
            ).first()
            
            if not member:
                emit('error', {'message': '채팅방에 접근할 권한이 없습니다.'})
                return
            
            # 메시지 저장
            new_message = ChatMessage(
                room_id=room_id,
                user_id=user_id,
                message_type=message_type,
                content=content if message_type == 'text' else None,
                file_info=file_info if message_type != 'text' else None
            )
            db.add(new_message)
            db.commit()
            db.refresh(new_message)
            
            # 첨부파일이 있으면 저장
            if file_info and isinstance(file_info, dict):
                attachment = ChatAttachment(
                    message_id=new_message.id,
                    filename=file_info.get('filename', ''),
                    file_type=file_info.get('file_type', 'file'),
                    file_size=file_info.get('size', 0),
                    storage_key=file_info.get('key', ''),
                    storage_url=file_info.get('url', ''),
                    thumbnail_url=file_info.get('thumbnail_url')
                )
                db.add(attachment)
                db.commit()

                # 이미지 첨부의 썸네일은 비동기로 생성해 요청 지연을 줄인다.
                if attachment.file_type == 'image' and not attachment.thumbnail_url and attachment.storage_key:
                    schedule_chat_thumbnail_generation(attachment.storage_key)
            
            # 사용자 정보 포함하여 메시지 데이터 구성
            user = db.query(User).filter(User.id == user_id).first()
            message_data = new_message.to_dict()
            if user:
                message_data['user_name'] = user.name
                message_data['user_username'] = user.username
            
            # 첨부파일 정보 추가
            attachments = db.query(ChatAttachment).filter(
                ChatAttachment.message_id == new_message.id
            ).all()
            if attachments:
                message_data['attachments'] = [a.to_dict() for a in attachments]
            
            # 채팅방의 모든 사용자에게 메시지 브로드캐스트
            # 1. Room 기반 브로드캐스트 (채팅 페이지에서 실시간 업데이트용)
            socketio.emit('new_message', message_data, room=str(room_id))
            
            # 2. User 기반 브로드캐스트 (모든 페이지에서 알림 수신용)
            # 채팅방 멤버 목록 조회
            members = db.query(ChatRoomMember).filter(
                ChatRoomMember.room_id == room_id
            ).all()
            
            # 각 멤버에게 직접 메시지 전송 (발신자 제외)
            print(f"[SocketIO] 채팅방 {room_id} 멤버 수: {len(members)}")
            for member in members:
                if member.user_id != user_id:  # 자신이 보낸 메시지는 자신에게 알림 불필요
                    user_room = f'user_{member.user_id}'
                    print(f"[SocketIO] 📤 사용자 {member.user_id}에게 메시지 전송 시도 (room: {user_room})")
                    socketio.emit('new_message', message_data, room=user_room)
                    print(f"[SocketIO] ✅ 사용자 {member.user_id}에게 메시지 전송 완료 (room: {user_room})")
                else:
                    print(f"[SocketIO] ⏭️ 발신자 {member.user_id}는 알림에서 제외")
            
            # 채팅방 업데이트 시간 갱신
            room.updated_at = datetime.datetime.now()
            db.commit()
            
            print(f"[SocketIO] 메시지 전송: 사용자 {user_id} -> 방 {room_id}")
            
        except Exception as e:
            db.rollback()
            import traceback
            print(f"메시지 전송 오류: {e}")
            print(traceback.format_exc())
            emit('error', {'message': f'메시지 전송 중 오류가 발생했습니다: {str(e)}'})
    
    @socketio.on('typing')
    def handle_typing(data):
        """타이핑 중 알림 (Quest 7)"""
        user_id = session.get('user_id')
        if not user_id:
            return
        
        room_id = data.get('room_id')
        is_typing = data.get('is_typing', False)
        
        if room_id:
            # 방의 다른 사용자들에게 타이핑 상태 전송 (자신 제외)
            socketio.emit('user_typing', {
                'room_id': room_id,
                'user_id': user_id,
                'is_typing': is_typing
            }, room=str(room_id), skip_sid=request.sid)
    
    @socketio.on('mark_read')
    def handle_mark_read(data):
        """메시지 읽음 표시 (Quest 7)"""
        user_id = session.get('user_id')
        if not user_id:
            return
        
        try:
            db = get_db()
            room_id = data.get('room_id')
            
            if room_id:
                # 마지막 읽은 시간 업데이트
                member = db.query(ChatRoomMember).filter(
                    ChatRoomMember.room_id == room_id,
                    ChatRoomMember.user_id == user_id
                ).first()
                
                if member:
                    member.last_read_at = datetime.datetime.now()
                    db.commit()
                    
                    # 읽음 상태를 다른 사용자들에게 알림
                    socketio.emit('message_read', {
                        'room_id': room_id,
                        'user_id': user_id
                    }, room=str(room_id))
        
        except Exception as e:
            db.rollback()
            print(f"읽음 표시 오류: {e}")


@app.route('/admin/test-r2')
@login_required
@role_required(['ADMIN', 'MANAGER'])
def admin_test_r2():
    """R2 연결 테스트 및 디버깅"""
    try:
        storage = get_storage()
        
        # 1. 환경 변수 상태 확인 (마스킹)
        env_status = {
            'STORAGE_TYPE': storage.storage_type,
            'HAS_ENDPOINT': bool(storage.endpoint_url),
            'HAS_ACCESS_KEY': bool(storage.access_key_id),
            'HAS_SECRET_KEY': bool(storage.secret_access_key),
            'HAS_BUCKET': bool(storage.bucket_name),
            'ENDPOINT_URL': storage.endpoint_url if storage.endpoint_url else 'None'
        }
        
        if storage.storage_type != 'r2':
            return jsonify({
                'success': False,
                'message': '현재 R2가 활성화되지 않았습니다.',
                'debug_info': env_status
            })

        # 2. 실제 연결 테스트 (List Objects)
        try:
            # 버킷의 파일 목록 1개만 가져와봄
            response = storage.client.list_objects_v2(Bucket=storage.bucket_name, MaxKeys=1)
            
            # 3. 썸네일/이미지 URL 테스트
            test_url = storage.get_download_url('test_connection.txt')
            
            return jsonify({
                'success': True,
                'message': 'R2 연결 성공! (AWS S3 API 통신 확인됨)',
                'bucket_name': storage.bucket_name,
                'key_count': response.get('KeyCount', 0),
                'debug_info': env_status,
                'generated_test_url': test_url
            })
            
        except Exception as e:
            return jsonify({
                'success': False,
                'message': f'R2 통신 실패: {str(e)}',
                'error_type': type(e).__name__,
                'debug_info': env_status
            })

    except Exception as e:
        return jsonify({'success': False, 'message': f'테스트 중 오류: {str(e)}'})

# Production: Auto-initialize Database Tables
# This runs when a WSGI server imports 'app' (e.g. gunicorn/uwsgi),
# and should not run for `python app.py` because __main__ has its own startup flow.
_should_run_auto_init = (__name__ != '__main__')
if _should_run_auto_init:
    try:
        with app.app_context():
            print("[AUTO-INIT] Checking database tables...")
            from db import init_db
            from wdcalculator_db import init_wdcalculator_db
            
            # Initialize Main DB
            init_db()
            ensure_order_attachments_category_column()
            ensure_order_attachments_item_index_column()
            
            # Initialize WDCalculator DB
            init_wdcalculator_db()
            
            print("[AUTO-INIT] Tables checked/created successfully.")
            
            # Check/Create Admin User
            from models import User
            from werkzeug.security import generate_password_hash
            db_session = get_db()
            try:
                admin = db_session.query(User).filter_by(username='admin').first()
                if not admin:
                    print("[AUTO-INIT] Creating default admin user (admin/admin1234)...")
                    new_admin = User(
                        username='admin',
                        password=generate_password_hash('admin1234'),
                        name='관리자',
                        role='ADMIN',
                        is_active=True
                    )
                    db_session.add(new_admin)
                    db_session.commit()
                else:
                    print("[AUTO-INIT] Admin user exists.")
            except Exception as e:
                print(f"[AUTO-INIT] Failed to create admin user: {e}")
                db_session.rollback()
                
    except Exception as e:
        print(f"[AUTO-INIT] Database initialization failed: {e}")

if __name__ == '__main__':
    _use_reloader = (os.environ.get('FLASK_USE_RELOADER', '1') == '1')
    _is_reloader_child = (os.environ.get('WERKZEUG_RUN_MAIN') == 'true')
    _should_run_startup_tasks = (not _use_reloader) or _is_reloader_child

    # 안전한 시작 프로세스 실행 (SystemExit 방지)
    try:
        import logging
        import sys
        
        # 로깅 설정
        logging.basicConfig(
            level=logging.INFO, 
            format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
            handlers=[
                logging.FileHandler('app_startup.log', encoding='utf-8'),
                logging.StreamHandler(sys.stdout)
            ]
        )
        logger = logging.getLogger('FOMS_Startup')

        if _should_run_startup_tasks:
            logger.info("[START] FOMS 애플리케이션 시작 중...")
            startup_success = True

            # 1. 데이터베이스 초기화 시도
            try:
                init_db()
                # get_db()를 사용하는 자동 컬럼 보정은 Flask 앱 컨텍스트 내에서 실행해야 한다.
                with app.app_context():
                    ensure_order_attachments_category_column()
                    ensure_order_attachments_item_index_column()
                logger.info("[OK] FOMS 데이터베이스 초기화 완료")
            except Exception as e:
                logger.error(f"[ERROR] FOMS 데이터베이스 초기화 실패: {str(e)}")
                startup_success = False

            # 1-1. 견적 계산기 독립 데이터베이스 초기화 시도
            try:
                with app.app_context():
                    init_wdcalculator_db()
                logger.info("[OK] 견적 계산기 데이터베이스 초기화 완료")
            except Exception as e:
                logger.warning(f"[WARN] 견적 계산기 데이터베이스 초기화 실패 (견적 기능 제한): {str(e)}")
                # 견적 계산기 DB 실패는 전체 시스템에 영향 없음

            # 2. 안전한 스키마 마이그레이션 시도
            try:
                from safe_schema_migration import run_safe_migration

                # Flask 앱 컨텍스트 내에서 마이그레이션 실행
                with app.app_context():
                    migration_success = run_safe_migration(app.app_context())
                    if migration_success:
                        logger.info("[OK] 스키마 마이그레이션 완료")
                    else:
                        logger.warning("[WARN] 스키마 마이그레이션 실패 - 기존 스키마로 계속 진행")
                        startup_success = False
            except Exception as e:
                logger.error(f"[ERROR] 스키마 마이그레이션 중 예외: {str(e)}")
                startup_success = False

            # 3. 시작 결과 요약
            if startup_success:
                logger.info("[SUCCESS] 모든 시작 프로세스가 성공적으로 완료되었습니다!")
                print("[OK] FOMS 시스템이 준비되었습니다!")
            else:
                logger.warning("[WARN] 일부 시작 프로세스에서 오류가 발생했지만 앱은 정상적으로 시작됩니다.")
                print("[WARN] 일부 기능에 제한이 있을 수 있습니다. 로그를 확인해주세요.")
        else:
            logger.info("[SKIP] 리로더 부모 프로세스에서는 시작 초기화를 건너뜁니다.")

        # 4. Flask 웹 서버 시작 (안전한 설정)
        if _should_run_startup_tasks:
            print("[START] 웹 서버를 시작합니다...")
            print(f"[INFO] SOCKETIO_AVAILABLE: {SOCKETIO_AVAILABLE}")
            print(f"[INFO] socketio 객체 존재: {socketio is not None}")

        if SOCKETIO_AVAILABLE and socketio:
            # SocketIO 사용 시 socketio.run() 사용
            if _should_run_startup_tasks:
                print("[INFO] Socket.IO 모드로 서버를 시작합니다...")
            socketio.run(
                app,
                host='0.0.0.0',
                port=5000,
                debug=True,
                use_reloader=_use_reloader,
                allow_unsafe_werkzeug=True,
            )
        else:
            # 일반 Flask 실행
            if _should_run_startup_tasks:
                print("[WARN] Socket.IO가 비활성화되어 일반 Flask 모드로 시작합니다...")
            app.run(
                host='0.0.0.0',
                port=5000,
                debug=True,
                use_reloader=_use_reloader,
            )
        
    except KeyboardInterrupt:
        print("\n[STOP] 사용자에 의해 서버가 중단되었습니다.")
    except Exception as e:
        print(f"[ERROR] 서버 시작 중 오류: {str(e)}")
        print("[INFO] 로그 파일(app_startup.log)을 확인해주세요.")
        # SystemExit 대신 정상 종료
    finally:
        if _should_run_startup_tasks:
            print("[END] FOMS 시스템을 종료합니다.")
